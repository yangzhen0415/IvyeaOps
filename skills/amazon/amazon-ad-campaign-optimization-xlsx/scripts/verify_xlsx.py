"""
Verify a generated xlsx matches the golden visual baseline.

Two failure tiers:
  HARD failures (return exit 1)：
    H1. Sheet count == 8
    H2. Sheet names match SHEET_NAMES (order matters)
    H3. R1 of every sheet uses navy text (#1F4E78), bold, 14pt
    H4. R1 font name contains '微软雅黑'

  SOFT warnings (printed but exit 0)：
    S1. Sheet row count below MIN_ROWS_PER_SHEET hint
    S2. Sheet merged-cell count below MIN_MERGED_PER_SHEET hint
    S3. Column widths drift > 30% from spec
    S4. Sheet 04 missing protected_keywords block (no green-fill row in upper third)
    S5. Sheet 02 R3 not red text (#C00000)

Usage:
    python verify_xlsx.py path/to/file.xlsx [--strict-soft]

`--strict-soft` promotes soft warnings to exit 1 (used by E2 smoke).
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SHEET_NAMES = [
    "01-现状诊断", "02-核心动作", "03-新Campaign搭建", "04-否定词清单",
    "05-加码清单", "06-预算重分配", "07-执行Checklist", "08-风险提示",
]

EXPECTED_COL_WIDTHS = {
    "01-现状诊断":       [38, 11, 9, 14, 10, 9, 38, 15],
    "02-核心动作":       [9, 22, 28, 35, 35, 12],
    "03-新Campaign搭建": [22, 38, 40, 12],
    "04-否定词清单":     [42, 14, 40, 16],
    "05-加码清单":       [45, 11, 22, 12, 24, 20],
    "06-预算重分配":     [28, 15, 14, 20, 35, 15],
    "07-执行Checklist":  [14, 52, 26, 12, 8],
    "08-风险提示":       [35, 38, 40],
}

# Soft thresholds: any populated plan should at minimum have header rows + 1 data row per section
MIN_ROWS_PER_SHEET = {
    "01-现状诊断": 6,    # title + 2 sections each with header + 1 row
    "02-核心动作": 5,    # title + section + header + 1 action
    "03-新Campaign搭建": 4,  # title + section + header (or empty-state msg)
    "04-否定词清单": 3,
    "05-加码清单": 3,
    "06-预算重分配": 6,  # always has total + before/after table
    "07-执行Checklist": 4,
    "08-风险提示": 8,    # 2 fixed sections × (section + header + ≥1 row + spacer)
}

MIN_MERGED_PER_SHEET = {
    "01-现状诊断": 3,    # title + 2 sections
    "02-核心动作": 2,
    "03-新Campaign搭建": 2,
    "04-否定词清单": 2,
    "05-加码清单": 2,
    "06-预算重分配": 2,
    "07-执行Checklist": 2,
    "08-风险提示": 3,
}

NAVY = "1F4E78"
RED_TEXT = "C00000"
GREEN_BG = "C6EFCE"


def _color_eq(rgb_attr, expect: str) -> bool:
    if not rgb_attr:
        return False
    val = getattr(rgb_attr, "rgb", rgb_attr)
    if not isinstance(val, str):
        return False
    return val.upper().endswith(expect.upper())


# ---------------------------------------------------------------------------
# HARD checks
# ---------------------------------------------------------------------------
def check_hard(wb) -> list[str]:
    fails: list[str] = []

    # H1: sheet count
    if len(wb.sheetnames) != 8:
        fails.append(f"[H1] sheet 数错误：实际 {len(wb.sheetnames)}，应 8")
        return fails  # follow-up checks meaningless

    # H2: sheet names + order
    if wb.sheetnames != SHEET_NAMES:
        fails.append(f"[H2] sheet 名称/顺序不匹配：\n        实际: {wb.sheetnames}\n        预期: {SHEET_NAMES}")

    # H3 + H4: R1 styling per sheet
    for name in SHEET_NAMES:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        a1 = ws.cell(1, 1)
        f = a1.font
        if not f or f.name is None or "微软雅黑" not in f.name:
            fails.append(f"[H4] {name} A1 字体非微软雅黑（实际: {f.name if f else None}）")
        if not f or not f.bold or (f.size or 0) < 13.5:
            fails.append(f"[H3] {name} A1 标题非粗体≥14pt（bold={f.bold if f else None} size={f.size if f else None}）")
        if not f or not _color_eq(f.color, NAVY):
            actual = getattr(f.color, "rgb", None) if f and f.color else None
            fails.append(f"[H3] {name} A1 标题色非 navy #{NAVY}（实际: {actual}）")

    return fails


# ---------------------------------------------------------------------------
# SOFT checks
# ---------------------------------------------------------------------------
def check_soft(wb) -> list[str]:
    warns: list[str] = []

    for name, expected_widths in EXPECTED_COL_WIDTHS.items():
        if name not in wb.sheetnames:
            continue
        ws: Worksheet = wb[name]

        # S1: row count
        min_rows = MIN_ROWS_PER_SHEET.get(name, 3)
        if ws.max_row < min_rows:
            warns.append(f"[S1] {name} 数据行 ({ws.max_row}) 少于参考下限 {min_rows}")

        # S2: merged count
        min_merged = MIN_MERGED_PER_SHEET.get(name, 2)
        if len(ws.merged_cells.ranges) < min_merged:
            warns.append(f"[S2] {name} 合并区数 ({len(ws.merged_cells.ranges)}) 少于参考下限 {min_merged}")

        # S3: column widths drift
        for i, expected in enumerate(expected_widths, start=1):
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(i)
            actual = ws.column_dimensions[letter].width
            if actual is None:
                warns.append(f"[S3] {name} 列 {letter} 未设置宽度（预期 {expected}）")
            elif abs(actual - expected) / expected > 0.3:
                warns.append(f"[S3] {name} 列 {letter} 宽度 {actual} 偏离预期 {expected} > 30%")

    # S4: Sheet 04 protected box (any green row in first 8 rows)
    if "04-否定词清单" in wb.sheetnames:
        ws = wb["04-否定词清单"]
        has_protected = False
        for r in range(2, min(ws.max_row, 8) + 1):
            cell = ws.cell(r, 1)
            if cell.fill and cell.fill.fgColor and _color_eq(cell.fill.fgColor, GREEN_BG):
                has_protected = True
                break
        if not has_protected:
            # may be intentional (no protected_keywords) — soft warning only
            warns.append("[S4] 04-否定词清单 上半部未发现保护词绿底块（如本次确无保护词可忽略）")

    # S5: Sheet 02 R3 red text (principle line)
    if "02-核心动作" in wb.sheetnames:
        ws = wb["02-核心动作"]
        c = ws.cell(3, 1)
        f = c.font
        if c.value and (not f or not _color_eq(f.color, RED_TEXT)):
            actual = getattr(f.color, "rgb", None) if f and f.color else None
            warns.append(f"[S5] 02-核心动作 R3 原则行非红字 #{RED_TEXT}（实际: {actual}）")

    return warns


# ---------------------------------------------------------------------------
# Entry
# ---------------------------------------------------------------------------
def verify(xlsx_path: Path, strict_soft: bool = False) -> int:
    wb = load_workbook(xlsx_path)
    hard = check_hard(wb)
    soft = check_soft(wb)

    print(f"📋 verify {xlsx_path.name}")
    print(f"  sheets: {len(wb.sheetnames)}")

    if hard:
        print(f"\n❌ HARD failures ({len(hard)}):")
        for f in hard:
            print(f"  {f}")
    else:
        print(f"\n✅ HARD checks passed (4 categories)")

    if soft:
        print(f"\n⚠️  SOFT warnings ({len(soft)}):")
        for w in soft:
            print(f"  {w}")
    else:
        print(f"\n✅ SOFT checks passed")

    if hard:
        return 1
    if soft and strict_soft:
        return 1
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="path to .xlsx file to verify")
    ap.add_argument("--strict-soft", action="store_true",
                    help="treat soft warnings as failures (used by E2 smoke)")
    args = ap.parse_args()

    sys.exit(verify(Path(args.xlsx), strict_soft=args.strict_soft))


if __name__ == "__main__":
    main()
