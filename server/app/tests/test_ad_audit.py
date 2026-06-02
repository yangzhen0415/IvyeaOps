"""Tests for ad-report audit service + routes.

These stay offline — we never spawn a real agent. Subprocess-driven paths
are exercised via direct mutation of the Job state + calling the
post-processing helpers.
"""
from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("IVYEA_OPS_DEV_MODE", "1")
    monkeypatch.setenv("IVYEA_OPS_ALLOWED_ORIGINS", "https://test.example.com")
    monkeypatch.setenv("IVYEA_OPS_SECRET", "test-secret")
    # Reload config so the env vars take effect.
    import importlib
    from app.core import config as cfg_mod
    importlib.reload(cfg_mod)
    from app import main as main_mod
    importlib.reload(main_mod)

    from app.services import ad_audit

    monkeypatch.setattr(ad_audit, "AD_AUDIT_ROOT", tmp_path / "ad-audits")
    ad_audit.AD_AUDIT_ROOT.mkdir(parents=True, exist_ok=True)

    # Bypass cookie auth the same way other routers' tests do.
    from app.core import security as sec_mod
    main_mod.app.dependency_overrides[sec_mod.require_user] = lambda: "tester"

    with TestClient(main_mod.app) as c:
        c.headers.update({"Origin": "https://test.example.com"})
        yield c


# ---------- helpers ----------


def _make_sp_xlsx(path: Path) -> None:
    """Build a tiny SP search-term-report-shaped xlsx for preview tests."""
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Date", "Campaign Name", "Ad Group Name", "Customer Search Term",
        "Match Type", "Impressions", "Clicks", "Spend", "Orders",
    ])
    ws.append(["2026-01-01", "SP-Core", "AG1", "trail camera", "exact", 1200, 40, 22.0, 3])
    ws.append(["2026-01-02", "SP-Core", "AG1", "game camera", "phrase", 900, 20, 15.0, 1])
    ws.append(["2026-01-15", "SP-Core", "AG1", "deer cam", "broad", 400, 8, 6.0, 0])
    wb.save(path)


def _xlsx_bytes(path: Path) -> bytes:
    return path.read_bytes()


# ---------- preview / detection ----------


def test_detect_ad_type_sp_from_columns():
    from app.services.ad_audit import _detect_ad_type

    cols = ["Date", "Customer Search Term", "Match Type", "Impressions"]
    assert _detect_ad_type(cols, "report.xlsx") == "SP"


def test_detect_ad_type_sd_wins_over_filename():
    from app.services.ad_audit import _detect_ad_type

    cols = ["Date", "Viewable Impressions", "Clicks", "Spend"]
    # Filename says SP but columns say SD — columns win.
    assert _detect_ad_type(cols, "sp_report.xlsx") == "SD"


def test_detect_ad_type_filename_fallback():
    from app.services.ad_audit import _detect_ad_type

    # No distinctive column names — filename breaks the tie.
    cols = ["A", "B", "C"]
    assert _detect_ad_type(cols, "report_sb.xlsx") == "SB"
    assert _detect_ad_type(cols, "plain.xlsx") == ""


def test_preview_report_xlsx(tmp_path):
    from app.services.ad_audit import _preview_report

    p = tmp_path / "report.xlsx"
    _make_sp_xlsx(p)
    info = _preview_report(p, "report_sp.xlsx")
    assert info["ad_type"] == "SP"
    assert info["row_count"] == 3
    assert info["date_range"] == "2026-01-01 ~ 2026-01-15"
    assert "Customer Search Term" in info["columns"]


def test_preview_report_csv(tmp_path):
    from app.services.ad_audit import _preview_report

    p = tmp_path / "sb.csv"
    p.write_text(
        "Date,Sponsored Brands Campaign,Impressions,Clicks\n"
        "2026-02-01,Camp-A,500,12\n"
        "2026-02-02,Camp-A,300,8\n",
        encoding="utf-8",
    )
    info = _preview_report(p, "sb.csv")
    assert info["ad_type"] == "SB"
    assert info["row_count"] == 2


# ---------- upload endpoint ----------


def test_upload_rejects_unsupported_ext(client, tmp_path):
    resp = client.post(
        "/api/ad-audit/upload",
        files={"file": ("malware.exe", b"x", "application/octet-stream")},
    )
    assert resp.status_code == 400
    assert "unsupported" in resp.json()["detail"].lower()


def test_upload_xlsx_creates_job(client, tmp_path):
    p = tmp_path / "sp.xlsx"
    _make_sp_xlsx(p)
    with p.open("rb") as f:
        resp = client.post(
            "/api/ad-audit/upload",
            files={"file": ("sp.xlsx", f.read(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"marketplace": "US"},
        )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["ad_type"] == "SP"
    assert body["row_count"] == 3
    assert body["status"] == "uploaded"
    assert len(body["job_id"]) == 12


def test_upload_rejects_oversize(client, monkeypatch, tmp_path):
    from app.services import ad_audit

    monkeypatch.setattr(ad_audit, "MAX_UPLOAD_BYTES", 100)  # tiny cap
    p = tmp_path / "big.xlsx"
    _make_sp_xlsx(p)  # > 100 bytes
    with p.open("rb") as f:
        resp = client.post(
            "/api/ad-audit/upload",
            files={"file": ("big.xlsx", f.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 400
    assert "too large" in resp.json()["detail"].lower()


# ---------- multi-source upload ----------


def test_upload_append_creates_second_source(client, tmp_path):
    """Second upload with job_id appends a source to the same job."""
    p1 = tmp_path / "sp_exact.xlsx"
    _make_sp_xlsx(p1)
    up1 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("sp_exact.xlsx", p1.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert up1.status_code == 200
    job_id = up1.json()["job_id"]
    assert len(up1.json()["sources"]) == 1

    # Second file — give it different content so dedup doesn't kick in.
    p2 = tmp_path / "sp_auto.csv"
    p2.write_text(
        "Date,Customer Search Term,Match Type,Impressions,Clicks,Spend,Orders\n"
        "2026-01-01,trail cam,auto,800,20,10.0,1\n"
        "2026-01-02,game camera,auto,500,10,5.0,0\n",
        encoding="utf-8",
    )
    up2 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("sp_auto.csv", p2.read_bytes(), "text/csv")},
        data={"job_id": job_id},
    )
    assert up2.status_code == 200, up2.text
    body = up2.json()
    assert body["job_id"] == job_id
    assert len(body["sources"]) == 2
    # Top-level mirrors the FIRST source (for legacy list-view compatibility).
    assert body["file_name"] == "sp_exact.xlsx"


def test_upload_append_dedupes_identical_content(client, tmp_path):
    """Re-uploading the exact same bytes returns 400 instead of adding a duplicate."""
    p = tmp_path / "sp.xlsx"
    _make_sp_xlsx(p)
    content = p.read_bytes()
    up1 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("sp.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    job_id = up1.json()["job_id"]
    dup = client.post(
        "/api/ad-audit/upload",
        files={"file": ("sp.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"job_id": job_id},
    )
    assert dup.status_code == 400
    assert "已经上传过" in dup.json()["detail"] or "duplicate" in dup.json()["detail"].lower()


def test_upload_append_enforces_source_cap(client, tmp_path, monkeypatch):
    """Append beyond MAX_SOURCES returns 400."""
    from app.services import ad_audit
    monkeypatch.setattr(ad_audit, "MAX_SOURCES", 2)

    p1 = tmp_path / "a.xlsx"
    _make_sp_xlsx(p1)
    up1 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("a.xlsx", p1.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    job_id = up1.json()["job_id"]

    p2 = tmp_path / "b.csv"
    p2.write_text("Date,Customer Search Term,Impressions\n2026-01-01,x,1\n", encoding="utf-8")
    up2 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("b.csv", p2.read_bytes(), "text/csv")},
        data={"job_id": job_id},
    )
    assert up2.status_code == 200

    p3 = tmp_path / "c.csv"
    p3.write_text("Date,Customer Search Term,Impressions\n2026-01-02,y,2\n", encoding="utf-8")
    up3 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("c.csv", p3.read_bytes(), "text/csv")},
        data={"job_id": job_id},
    )
    assert up3.status_code == 400
    assert "2" in up3.json()["detail"]  # mentions the cap


def test_remove_and_update_source(client, tmp_path):
    """DELETE removes a source; PATCH updates campaign_name / daily budget."""
    p1 = tmp_path / "a.xlsx"
    _make_sp_xlsx(p1)
    up1 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("a.xlsx", p1.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    job_id = up1.json()["job_id"]
    source_id_1 = up1.json()["sources"][0]["source_id"]

    # Add a second source
    p2 = tmp_path / "b.csv"
    p2.write_text("Date,Customer Search Term,Impressions\n2026-01-01,x,1\n", encoding="utf-8")
    up2 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("b.csv", p2.read_bytes(), "text/csv")},
        data={"job_id": job_id},
    )
    source_id_2 = up2.json()["sources"][1]["source_id"]
    assert len(up2.json()["sources"]) == 2

    # Rename + budget
    patch = client.patch(
        f"/api/ad-audit/{job_id}/source/{source_id_1}",
        json={"campaign_name": "SP-Exact-Core", "daily_budget_usd": 50.0},
    )
    assert patch.status_code == 200, patch.text
    renamed = [s for s in patch.json()["sources"] if s["source_id"] == source_id_1][0]
    assert renamed["campaign_name"] == "SP-Exact-Core"
    assert renamed["daily_budget_usd"] == 50.0

    # Clear budget
    patch2 = client.patch(
        f"/api/ad-audit/{job_id}/source/{source_id_1}",
        json={"clear_daily_budget": True},
    )
    assert patch2.status_code == 200
    cleared = [s for s in patch2.json()["sources"] if s["source_id"] == source_id_1][0]
    assert cleared["daily_budget_usd"] is None

    # Remove the second source
    rm = client.delete(f"/api/ad-audit/{job_id}/source/{source_id_2}")
    assert rm.status_code == 200
    assert len(rm.json()["sources"]) == 1
    assert rm.json()["sources"][0]["source_id"] == source_id_1

    # Cannot remove the last source
    rm2 = client.delete(f"/api/ad-audit/{job_id}/source/{source_id_1}")
    assert rm2.status_code == 400
    assert "至少" in rm2.json()["detail"] or "last" in rm2.json()["detail"].lower()


def test_start_multi_source_requires_asin(client, tmp_path):
    """start_job rejects multi-source submissions without an ASIN."""
    p1 = tmp_path / "a.xlsx"
    _make_sp_xlsx(p1)
    up1 = client.post(
        "/api/ad-audit/upload",
        files={"file": ("a.xlsx", p1.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    job_id = up1.json()["job_id"]
    p2 = tmp_path / "b.csv"
    p2.write_text("Date,Customer Search Term,Impressions\n2026-01-01,x,1\n", encoding="utf-8")
    client.post(
        "/api/ad-audit/upload",
        files={"file": ("b.csv", p2.read_bytes(), "text/csv")},
        data={"job_id": job_id},
    )

    # Missing ASIN — should be rejected at start.
    resp = client.post("/api/ad-audit/start", json={
        "job_id": job_id, "goal": "profit", "asin": "",
    })
    # 400 because ASIN missing + runner probably available.
    # (Some CI envs have no runner → 400 too. Either way, NOT 200.)
    assert resp.status_code == 400


# ---------- start: state / validation ----------


def test_start_rejects_bad_goal(client, tmp_path):
    p = tmp_path / "sp.xlsx"
    _make_sp_xlsx(p)
    up = client.post(
        "/api/ad-audit/upload",
        files={"file": ("sp.xlsx", p.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    job_id = up.json()["job_id"]
    resp = client.post("/api/ad-audit/start", json={
        "job_id": job_id, "goal": "nonsense",
    })
    assert resp.status_code == 400


def test_start_rejects_missing_job(client):
    resp = client.post("/api/ad-audit/start", json={
        "job_id": "missing_job", "goal": "profit",
    })
    assert resp.status_code == 404


# ---------- prompt assembly ----------


def test_prompt_includes_goal_posture_and_protected():
    from app.services.ad_audit import AdJob, _build_prompt

    job = AdJob(
        job_id="j1",
        file_name="sp.xlsx",
        file_ext="xlsx",
        ad_type="SP",
        marketplace="US",
        date_range="2026-01-01 ~ 2026-01-15",
        row_count=500,
        goal="new_launch",
        protected_keywords=["trail camera", "cellular trail camera"],
        asin="B0TEST12345",
        product_notes="太阳能 4G 蜂窝 trail cam",
    )
    source = {
        "source_id": "s1", "campaign_name": "SP-Auto", "file_name": "sp.xlsx",
        "ad_type": "SP", "date_range": "2026-01-01 ~ 2026-01-15",
        "row_count": 500, "daily_budget_usd": None,
    }
    out = _build_prompt(job, [(source, Path("/tmp/raw.xlsx"))])
    assert "new_launch" in out
    assert "新品冲量" in out
    assert "trail camera" in out
    assert "B0TEST12345" in out
    # Global rules injected.
    assert "SBV" in out
    assert "Vine" in out
    # JSON schema stub present.
    assert "protected_keywords_status" in out
    assert "action_summary" in out
    # Source block rendered.
    assert "SP-Auto" in out
    assert "/tmp/raw.xlsx" in out


def test_prompt_empty_protected_says_per_data():
    from app.services.ad_audit import AdJob, _build_prompt

    job = AdJob(job_id="j2", ad_type="SP", goal="profit")
    source = {
        "source_id": "s1", "campaign_name": "solo", "file_name": "raw.xlsx",
        "ad_type": "SP", "date_range": "", "row_count": 0, "daily_budget_usd": None,
    }
    out = _build_prompt(job, [(source, Path("/tmp/raw.xlsx"))])
    # Must make clear core keywords may be negated when no protected list.
    assert "完全按数据判断" in out


def test_prompt_multi_source_injects_cross_campaign_and_budgets():
    """Multi-source prompt adds cross-campaign instruction + budget lines."""
    from app.services.ad_audit import AdJob, _build_prompt

    job = AdJob(
        job_id="m1",
        ad_type="SP",
        marketplace="US",
        goal="profit",
        asin="B0MULTI0001",
    )
    src_a = {
        "source_id": "a", "campaign_name": "SP-Exact",
        "file_name": "exact.xlsx", "ad_type": "SP",
        "date_range": "2026-01-01 ~ 2026-01-15", "row_count": 120,
        "daily_budget_usd": 30.0,
    }
    src_b = {
        "source_id": "b", "campaign_name": "SP-Auto",
        "file_name": "auto.xlsx", "ad_type": "SP",
        "date_range": "2026-01-01 ~ 2026-01-15", "row_count": 230,
        "daily_budget_usd": None,
    }
    out = _build_prompt(
        job,
        [(src_a, Path("/tmp/a.xlsx")), (src_b, Path("/tmp/b.xlsx"))],
    )
    # Both campaigns mentioned.
    assert "SP-Exact" in out and "SP-Auto" in out
    # Cross-campaign instruction kicks in when >1 source.
    assert "cross_campaign_insights" in out
    assert "跨活动对比" in out
    # Budget-only source shows $ line; budgetless source doesn't.
    assert "$30.00" in out
    # Budget note injected because at least one source has a budget.
    assert "预算重分配" in out


# ---------- xlsx generation ----------


def test_download_xlsx_builds_sheets_with_fills(tmp_path, monkeypatch):
    from app.services import ad_audit

    monkeypatch.setattr(ad_audit, "AD_AUDIT_ROOT", tmp_path / "r")
    ad_audit.AD_AUDIT_ROOT.mkdir(parents=True)
    jd = ad_audit.AD_AUDIT_ROOT / "jobx"
    jd.mkdir()

    structured = {
        "overview": {
            "ad_type": "SP", "marketplace": "US", "date_range": "2026-01-01 ~ 2026-01-15",
            "impressions": 10000, "clicks": 320, "spend": 200.5, "orders": 18,
            "sales": 1580.0, "acos": "12.7%", "ctr": "3.2%", "cvr": "5.6%",
            "one_line_verdict": "整体健康，低效词明显",
        },
        "campaign_efficiency": [
            {"campaign_name": "SP-Core-Exact", "type": "SP-Exact",
             "spend": 100, "spend_share": "50%", "orders": 12, "order_share": "67%",
             "cost_per_order": "$8.33", "acos": "10%",
             "efficiency_tag": "healthy", "verdict": "核心盘盈利良好"},
            {"campaign_name": "SP-Auto-Broad", "type": "SP-Auto",
             "spend": 100, "spend_share": "50%", "orders": 6, "order_share": "33%",
             "cost_per_order": "$16.67", "acos": "45%",
             "efficiency_tag": "black_hole", "verdict": "50%预算只产33%单量"},
        ],
        "protected_keywords_status": [
            {"keyword": "trail camera", "status": "good", "impressions": 4000, "clicks": 130,
             "spend": 80, "orders": 10, "acos": "10%", "note": "表现稳定"},
            {"keyword": "game camera", "status": "bad", "impressions": 600, "clicks": 25,
             "spend": 40, "orders": 0, "acos": "-", "note": "位置下滑"},
        ],
        "high_performers": [
            {"keyword": "solar trail camera", "match_type": "exact",
             "impressions": 2000, "clicks": 80, "spend": 40, "orders": 6,
             "acos": "8%", "action": "boost",
             "current_bid": "$1.00", "suggested_bid": "$1.20",
             "bid_change_pct": "+20%", "reason": "高 ROI"},
        ],
        "low_performers": [
            {"keyword": "bluetooth speaker", "match_type": "broad",
             "impressions": 800, "clicks": 30, "spend": 25, "orders": 0,
             "acos": "-", "action": "cut",
             "current_bid": "$0.80", "suggested_bid": "否定",
             "bid_change_pct": "-100%", "reason": "不相关"},
        ],
        "new_keyword_candidates": [
            {"keyword": "cellular trail camera 4g", "source_search_term": "cellular trail camera 4g lte",
             "impressions": 150, "orders": 2, "suggested_bid": "$0.90", "reason": "转化良好"},
        ],
        "negative_suggestions": [
            {"term": "bluetooth", "type": "immediate", "reason": "零转化",
             "wasted_spend_usd": 18.40, "window_days": 21},
            {"term": "security", "type": "watch", "reason": "观察两周"},
        ],
        "negative_wasted_total_usd": 67.30,
        "new_campaigns": [
            {"name": "SP-Exact-Core-ToS", "type": "SP", "match_type": "exact",
             "daily_budget_usd": 30, "bid_strategy": "down_only",
             "placement_modifiers": {"top_of_search": "+50%",
                                     "rest_of_search": "0%", "product_pages": "0%"},
             "keywords_with_bid": [{"keyword": "trail camera", "bid_usd": 2.50}],
             "sync_actions": ["SP-Auto 里加否定精准"],
             "verdict": "核心词独立打搜索首页"},
        ],
        "placement_diagnosis": [
            {"placement": "Top of Search", "impressions": 5000, "clicks": 200,
             "spend": 120, "orders": 12, "acos": "10%", "ctr": "4%", "cvr": "6%",
             "suggested_modifier": "+150%", "action": "加码 150%"},
        ],
        "action_summary": [
            {"level": "P0", "day": "Day 1", "eta_minutes": 10,
             "location_path": "广告活动 → SP-Auto → 否定关键词",
             "action": "砍掉 bluetooth speaker 组", "evidence": "0 单",
             "expected_impact": "节省 $25/周"},
            {"level": "P1", "day": "Day 2", "eta_minutes": 15,
             "location_path": "广告活动 → SP-Core → 关键词",
             "action": "守 trail camera 位置", "evidence": "CTR 3.2%",
             "expected_impact": "稳定流量"},
        ],
        "data_notes": "异常值：bluetooth speaker 花费偏高但零转化",
        "meta": {"analyzed_at": "2026-05-12", "row_count": 500, "threshold_posture": "profit"},
    }
    (jd / "report.json").write_text(json.dumps(structured, ensure_ascii=False), encoding="utf-8")
    (jd / "meta.json").write_text(json.dumps({
        "job_id": "jobx", "ad_type": "SP", "marketplace": "US",
        "date_range": "2026-01-01 ~ 2026-01-15", "goal": "profit",
        "asin": "", "protected_keywords": ["trail camera", "game camera"],
        "row_count": 500, "file_name": "sp.xlsx",
        "created_at": "2026-05-12T10:00:00+00:00",
        "finished_at": "2026-05-12T10:05:00+00:00",
        "runner_used": "hermes", "status": "done",
    }), encoding="utf-8")

    xp = ad_audit.download_path("jobx", "xlsx")
    assert xp is not None and xp.is_file()

    wb = load_workbook(xp)
    # 12 sheets after the landable-proposal upgrade (was 10)
    expected = [
        "总览", "Campaign 效率对比", "关键词守位",
        "高效词 Top-20", "低效词 Top-20", "新增关键词候选",
        "否词建议", "新 Campaign 搭建", "位置诊断",
        "执行 Checklist", "原始数据摘要", "元信息",
    ]
    assert wb.sheetnames == expected

    # Overview
    assert wb["总览"]["B1"].value == "值"

    # Protected keyword row has a status fill.
    pk_ws = wb["关键词守位"]
    assert pk_ws["A2"].value == "trail camera"
    assert pk_ws["B2"].value == "good"
    assert pk_ws["B2"].fill.fgColor.rgb.endswith("DFF5E1")  # pale green
    assert pk_ws["B3"].fill.fgColor.rgb.endswith("FBD4D4")  # bad → red

    # 高效词: new layout — action in col G, bid_change_pct in col J.
    hp_ws = wb["高效词 Top-20"]
    assert hp_ws["A2"].value == "solar trail camera"
    assert hp_ws["G2"].value == "boost"
    assert hp_ws["G2"].fill.fgColor.rgb.endswith("D6E8FF")  # boost blue
    assert hp_ws["H2"].value == "$1.00"        # current_bid
    assert hp_ws["I2"].value == "$1.20"        # suggested_bid
    assert hp_ws["J2"].value == "+20%"         # bid_change_pct
    assert hp_ws["J2"].fill.fgColor.rgb.endswith("DFF5E1")  # direction up → green

    # 低效词: "-100%" bid_change_pct column should be red-filled.
    lp_ws = wb["低效词 Top-20"]
    assert lp_ws["J2"].value == "-100%"
    assert lp_ws["J2"].fill.fgColor.rgb.endswith("FBD4D4")

    # Campaign efficiency fills.
    ce_ws = wb["Campaign 效率对比"]
    # black_hole tag → red fill on col I (efficiency column)
    assert "效率黑洞" in str(ce_ws["I3"].value)
    assert ce_ws["I3"].fill.fgColor.rgb.endswith("FBD4D4")
    # healthy tag → green fill
    assert ce_ws["I2"].fill.fgColor.rgb.endswith("DFF5E1")

    # 否词建议: wasted_spend column appears when any row has a value, plus total row.
    neg_ws = wb["否词建议"]
    headers = [cell.value for cell in neg_ws[1]]
    assert "过去浪费 $" in headers
    # The grand-total row is at the bottom.
    last_row = neg_ws.max_row
    assert neg_ws[f"A{last_row}"].value == "💰 合计预计直接省"

    # 位置诊断: suggested_modifier "+150%" → green fill on col H.
    pl_ws = wb["位置诊断"]
    assert pl_ws["H2"].value == "+150%"
    assert pl_ws["H2"].fill.fgColor.rgb.endswith("DFF5E1")

    # 执行 Checklist: ☐ + day + P0 fill on C col.
    act_ws = wb["执行 Checklist"]
    assert act_ws["A2"].value == "☐"
    assert act_ws["B2"].value == "Day 1"
    assert act_ws["C2"].value == "P0"
    assert act_ws["C2"].fill.fgColor.rgb.endswith("F8BFBF")  # P0 red
    assert act_ws["E2"].value == 10                          # eta_minutes
    assert "SP-Auto" in str(act_ws["F2"].value)              # location_path

    # 新 Campaign sheet.
    nc_ws = wb["新 Campaign 搭建"]
    assert nc_ws["A2"].value == "SP-Exact-Core-ToS"
    assert "trail camera" in str(nc_ws["I2"].value)  # keywords_with_bid joined


def test_list_and_clear_failed(tmp_path, monkeypatch):
    from app.services import ad_audit

    root = tmp_path / "list"
    root.mkdir()
    monkeypatch.setattr(ad_audit, "AD_AUDIT_ROOT", root)
    # Two jobs: one failed, one done.
    for jid, st in [("failedone", "failed"), ("donetwo", "done")]:
        d = root / jid
        d.mkdir()
        (d / "meta.json").write_text(json.dumps({
            "job_id": jid, "status": st, "ad_type": "SP",
            "created_at": "2026-05-12T10:00:00+00:00",
        }), encoding="utf-8")

    rows = ad_audit.list_jobs()
    assert {r["job_id"] for r in rows} == {"failedone", "donetwo"}
    removed = ad_audit.clear_failed()
    assert removed == 1
    assert {p.name for p in root.iterdir()} == {"donetwo"}


def test_sweep_stale_running_flips_ghosts(tmp_path, monkeypatch):
    from app.services import ad_audit

    root = tmp_path / "stale"
    root.mkdir()
    monkeypatch.setattr(ad_audit, "AD_AUDIT_ROOT", root)
    for jid, st in [("ghost1", "running"), ("ghost2", "queued"), ("keep", "done")]:
        d = root / jid
        d.mkdir()
        (d / "meta.json").write_text(json.dumps({
            "job_id": jid, "status": st,
            "created_at": "2026-05-12T10:00:00+00:00",
        }), encoding="utf-8")
    n = ad_audit.sweep_stale_running()
    assert n == 2
    assert json.loads((root / "ghost1" / "meta.json").read_text())["status"] == "failed"
    assert json.loads((root / "keep" / "meta.json").read_text())["status"] == "done"


def test_download_html_renders_self_contained(tmp_path, monkeypatch):
    """HTML export should be a standalone document with inline CSS + color classes."""
    from app.services import ad_audit

    root = tmp_path / "html"
    root.mkdir()
    monkeypatch.setattr(ad_audit, "AD_AUDIT_ROOT", root)
    jd = root / "htmljob"
    jd.mkdir()
    structured = {
        "overview": {
            "ad_type": "SP", "marketplace": "US", "date_range": "2026-01-01~01-31",
            "impressions": 15000, "clicks": 210, "spend": 285.5, "orders": 22,
            "sales": 1320.0, "acos": "21.6%", "ctr": "1.40%", "cvr": "10.5%",
            "one_line_verdict": "核心盈利，需要砍长尾烧钱词",
        },
        "protected_keywords_status": [
            {"keyword": "trail camera", "status": "good", "impressions": 8000,
             "clicks": 120, "spend": 150.0, "orders": 18, "acos": "19.2%",
             "note": "核心词表现健康"},
            {"keyword": "game camera", "status": "bad", "impressions": 4000,
             "clicks": 20, "spend": 45.0, "orders": 1, "acos": "112%",
             "note": "需要降价或否定"},
        ],
        "high_performers": [
            {"keyword": "solar trail camera", "match_type": "exact",
             "impressions": 900, "clicks": 30, "spend": 18.0, "orders": 6,
             "acos": "9.1%", "action": "boost", "suggested_bid": "$1.40",
             "reason": "ACOS 低，建议加码"},
        ],
        "low_performers": [
            {"keyword": "night vision camera", "match_type": "broad",
             "impressions": 1200, "clicks": 28, "spend": 42.0, "orders": 0,
             "acos": "—", "action": "cut", "suggested_bid": "否定",
             "reason": "0 单，泛需求"},
        ],
        "new_keyword_candidates": [
            {"keyword": "solar cellular trail camera", "source_search_term": "solar cellular trail camera",
             "impressions": 400, "orders": 5, "suggested_bid": "$1.20",
             "reason": "CVR 12.5% 已验证"},
        ],
        "negative_suggestions": [
            {"term": "battery", "type": "immediate", "reason": "6 点击 0 单"},
        ],
        "placement_diagnosis": [
            {"placement": "Top of Search", "impressions": 5000, "clicks": 120,
             "spend": 180.0, "orders": 15, "acos": "18%", "ctr": "2.4%",
             "cvr": "12.5%", "action": "可提溢价 20%"},
        ],
        "action_summary": [
            {"level": "P0", "action": "否定 10 个无相关宽词",
             "evidence": "8 词共 56 点击 0 单", "expected_impact": "省 $120/周期"},
            {"level": "P1", "action": "新增 solar cellular trail camera Exact",
             "evidence": "CVR 12.5% 已验证", "expected_impact": "+10 单/月"},
        ],
        "data_notes": "报告无 placement 分桶；建议下次导出 Placement Report。",
        "meta": {"analyzed_at": "2026-05-12T12:00:00+00:00",
                 "row_count": 150, "threshold_posture": "standard"},
    }
    (jd / "report.json").write_text(
        json.dumps(structured, ensure_ascii=False), encoding="utf-8",
    )
    (jd / "meta.json").write_text(
        json.dumps({
            "job_id": "htmljob", "ad_type": "SP", "marketplace": "US",
            "status": "done", "goal": "profit", "asin": "B0TEST1111",
            "protected_keywords": ["trail camera"],
            "product_note": "Solar + 4G trail camera",
            "runner_used": "hermes",
            "created_at": "2026-05-12T10:00:00+00:00",
        }),
        encoding="utf-8",
    )

    hp = ad_audit.download_path("htmljob", "html")
    assert hp is not None and hp.is_file() and hp.name == "report.html"
    text = hp.read_text(encoding="utf-8")
    # Doctype + self-contained style
    assert text.startswith("<!DOCTYPE html>")
    assert "<style>" in text and "</style>" in text
    # Color palette inlined (hex values from html_report)
    assert "#DFF5E1" in text  # good
    assert "#F8BFBF" in text  # P0
    # Core section titles present (emoji + upgraded wording after landable-proposal refit)
    for keyword in ("总览", "守护关键词", "高效词", "低效词",
                    "新增关键词候选", "否词建议", "位置诊断"):
        assert keyword in text, f"missing section keyword: {keyword}"
    # Checklist section renders under either heading depending on whether `day` is set.
    assert ("Checklist" in text) or ("建议汇总" in text)
    # Shield section gets its own decorative class
    assert "shield-sec" in text
    # Data flows through
    assert "trail camera" in text
    assert "solar trail camera" in text
    assert "2026-01-01~01-31" in text
    assert "核心盈利" in text
    # Idempotence: regen uses cached file unless json is newer.
    first_mtime = hp.stat().st_mtime
    hp2 = ad_audit.download_path("htmljob", "html")
    assert hp2 is not None
    assert hp2.stat().st_mtime == first_mtime
