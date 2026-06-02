"""Ad-report audit job manager.

Runs the ``zach-search-term-report-analyzer`` skill via an agent CLI
(hermes / codex / claude) as a background subprocess. Each job analyses a
single Amazon Ads search-term report (SP / SB / SD) that the user uploads.

Workflow
--------
1. ``upload_report`` — user uploads raw ``.xlsx`` / ``.csv``. We store it,
   read the first N rows to detect ``ad_type``, date range, row count, and
   column names. A job dir is created with status=``uploaded``.
2. ``start_job`` — user supplies task context (goal, protected keywords,
   notes, etc). Job flips to ``queued`` → ``running`` and spawns the agent.
3. ``get_job`` / ``download_path`` — poll status, fetch markdown / JSON /
   xlsx artifacts.

Design
------
- Single global lock **shared with ASIN audit** so only one heavy agent
  runs at a time (single-seat operator).
- Each job dir: ``~/.hermes/ivyea-ops-data/ad-audits/<job_id>/``
  - ``meta.json``       (full state + user context)
  - ``preview.json``    (detected ad_type / date range / columns)
  - ``raw.<ext>``       (uploaded report, kept for reproducibility)
  - ``report.md``       (agent markdown output)
  - ``report.json``     (structured analysis block parsed from tail)
  - ``report.xlsx``     (rendered on demand from report.json)
  - ``stdout.log``      (live tail)
- Hard timeout: 30 min. Retention: 30 days (swept on startup).
"""
from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import logging
import re
import shutil
import signal
import time
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from app.services.asin_audit import _job_lock  # share single-job semaphore
from app.services.asin_audit import _split_report_and_json
from app.services.runners import (
    build_child_env,
    _build_runner_cmd,
    resolve_with_pref,
    runner_status,
)

AD_AUDIT_ROOT = Path.home() / ".hermes" / "ivyea-ops-data" / "ad-audits"
AD_AUDIT_ROOT.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger(__name__)

HARD_TIMEOUT_SEC = 30 * 60
RETENTION_SEC = 30 * 24 * 3600
FLUSH_EVERY_BYTES = 4096

# Multi-source upload — one ASIN can have multiple campaigns (Auto/Exact/Phrase
# or SP/SB/SD), each exported as its own search-term report. We cap at 8 to
# keep prompt size bounded and prevent accidental batch-dump.
MAX_SOURCES = 8

# Upload size cap (bytes). Real SP search-term reports rarely exceed 20MB;
# larger files are almost certainly the wrong thing.
MAX_UPLOAD_BYTES = 20 * 1024 * 1024

# Valid ad types for this tool.
AD_TYPES = ("SP", "SB", "SD")

# Valid operational goals — each maps to a different threshold posture in
# the prompt.
GOALS = ("profit", "new_launch", "relaunch", "clearance")

# Output modes: "report" = zach-search-term-report-analyzer (markdown+json),
# "xlsx_plan" = amazon-ad-campaign-optimization-xlsx (8-sheet xlsx).
OUTPUT_MODES = ("report", "xlsx_plan")

# Goal mapping for xlsx_plan mode (maps our goals to the skill's goal enum).
_XLSX_GOAL_MAP = {
    "profit": "profit",
    "new_launch": "launch",
    "relaunch": "volume",
    "clearance": "clearance",
}

_live_jobs: Dict[str, "AdJob"] = {}


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _job_dir(job_id: str) -> Path:
    return AD_AUDIT_ROOT / job_id


# --------------------------------------------------------------------------- #
# Dataclass + meta I/O
# --------------------------------------------------------------------------- #

@dataclass
class AdJob:
    job_id: str
    # Upload / preview fields (populated at upload time).
    # For multi-source jobs, these mirror the FIRST source so legacy callers
    # (list view, status API) still see something sensible. Authoritative data
    # lives in ``sources``.
    file_name: str = ""
    file_ext: str = ""
    file_size: int = 0
    ad_type: str = ""             # SP / SB / SD (detected or user-supplied)
    marketplace: str = "US"
    date_range: str = ""          # e.g. "2025-01-01 ~ 2025-01-31"
    row_count: int = 0
    columns: List[str] = field(default_factory=list)
    # Multi-source reports (one job = one ASIN, multiple campaigns).
    # Empty list = legacy single-file job, fall back to raw.<ext>.
    sources: List[Dict[str, Any]] = field(default_factory=list)
    # Task-context fields (populated at start time).
    goal: str = ""                # profit / new_launch / relaunch / clearance
    output_mode: str = "report"   # report / xlsx_plan
    asin: str = ""                # target ASIN (REQUIRED for multi-source jobs)
    product_notes: str = ""       # free text
    protected_keywords: List[str] = field(default_factory=list)
    # Per-campaign daily budget in USD. Key = campaign_name from source entry.
    # Optional — omitted budgets fall back to % reallocation instead of $.
    daily_budgets: Dict[str, float] = field(default_factory=dict)
    runner_pref: str = "auto"
    runner_used: Optional[str] = None
    # State.
    status: str = "uploaded"      # uploaded|queued|running|done|failed|cancelled
    progress: str = ""
    created_at: str = field(default_factory=_now_iso)
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    error: Optional[str] = None
    pid: Optional[int] = None
    stdout_bytes: int = 0

    def to_meta(self) -> Dict[str, Any]:
        return asdict(self)


def _write_meta(job: AdJob) -> None:
    d = _job_dir(job.job_id)
    d.mkdir(parents=True, exist_ok=True)
    (d / "meta.json").write_text(
        json.dumps(job.to_meta(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _read_meta(job_id: str) -> Optional[Dict[str, Any]]:
    mp = _job_dir(job_id) / "meta.json"
    if not mp.is_file():
        return None
    try:
        return json.loads(mp.read_text(encoding="utf-8"))
    except Exception:
        return None


# --------------------------------------------------------------------------- #
# Upload + preview
# --------------------------------------------------------------------------- #

# Column-name fingerprints. Amazon's exports vary by locale and report
# version, so we match loosely (substring, case-insensitive).
_AD_TYPE_HINTS = {
    "SP": [
        # SP-specific columns.
        "customer search term", "search term",
    ],
    "SB": [
        "sponsored brands", "brand",
    ],
    "SD": [
        "sponsored display", "viewable impressions",
    ],
}


def _detect_ad_type(columns: List[str], filename: str) -> str:
    """Guess SP / SB / SD from columns + filename hints.

    Columns are most authoritative (SD has ``viewable impressions``,
    SB has brand-specific fields). Filename is a secondary hint.
    """
    cols_low = [c.lower() for c in columns]
    joined = " | ".join(cols_low)
    name_low = filename.lower()

    # Strongest signals first.
    if "viewable impressions" in joined or "sponsored display" in joined:
        return "SD"
    if "sponsored brands" in joined:
        return "SB"
    if "customer search term" in joined or "search term" in joined:
        # All three have search-term reports; default SP unless filename hints.
        if "sd" in name_low or "display" in name_low:
            return "SD"
        if "sb" in name_low or "brand" in name_low:
            return "SB"
        return "SP"
    # Fallback — filename hint only.
    for key in ("sd", "sb", "sp"):
        if key in name_low:
            return key.upper()
    return ""


def _extract_date_range(rows: List[List[Any]], columns: List[str]) -> str:
    """Find min/max of any date-looking column in the sample rows."""
    date_idxs: List[int] = []
    for i, name in enumerate(columns):
        nl = name.lower()
        if "date" in nl or "日期" in nl:
            date_idxs.append(i)
    if not date_idxs:
        return ""
    seen: List[str] = []
    for r in rows:
        for idx in date_idxs:
            if idx < len(r):
                v = r[idx]
                if v is None:
                    continue
                s = str(v).strip()
                if re.match(r"^\d{4}-\d{2}-\d{2}", s):
                    seen.append(s[:10])
    if not seen:
        return ""
    seen.sort()
    if seen[0] == seen[-1]:
        return seen[0]
    return f"{seen[0]} ~ {seen[-1]}"


def _read_xlsx_preview(path: Path) -> Tuple[List[str], List[List[Any]], int]:
    """Return (columns, sample_rows, total_rows) from an xlsx file.

    Uses openpyxl read_only to avoid loading huge sheets. Samples up to
    100 rows beyond the header for date-range detection.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        columns: List[str] = []
        sample: List[List[Any]] = []
        total = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                columns = [("" if c is None else str(c)).strip() for c in row]
                continue
            total += 1
            if len(sample) < 100:
                sample.append(list(row))
        return columns, sample, total
    finally:
        wb.close()


def _read_csv_preview(path: Path) -> Tuple[List[str], List[List[Any]], int]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], [], 0
    columns = [c.strip() for c in rows[0]]
    data = rows[1:]
    return columns, data[:100], len(data)


def _preview_report(path: Path, filename: str) -> Dict[str, Any]:
    """Produce the preview dict stored at preview.json."""
    ext = path.suffix.lower().lstrip(".")
    if ext in ("xlsx", "xlsm"):
        columns, sample, total = _read_xlsx_preview(path)
    elif ext == "csv":
        columns, sample, total = _read_csv_preview(path)
    else:
        raise ValueError(f"unsupported file type: .{ext}")
    ad_type = _detect_ad_type(columns, filename)
    date_range = _extract_date_range(sample, columns)
    return {
        "columns": columns,
        "row_count": total,
        "ad_type": ad_type,
        "date_range": date_range,
    }


def _source_entry(
    *,
    source_id: str,
    file_name: str,
    file_ext: str,
    file_size: int,
    file_hash: str,
    preview: Dict[str, Any],
    campaign_name: str = "",
    daily_budget_usd: Optional[float] = None,
) -> Dict[str, Any]:
    """Build one entry for ``AdJob.sources[]``.

    ``campaign_name`` defaults to the bare filename (without extension) so the
    user always sees *something* identifiable in the UI. They can rename it
    via PATCH.
    """
    base = Path(file_name).stem.strip() or file_name
    return {
        "source_id": source_id,
        "file_name": file_name,
        "file_ext": file_ext,
        "file_size": file_size,
        "file_hash": file_hash,
        "ad_type": preview.get("ad_type", ""),
        "date_range": preview.get("date_range", ""),
        "row_count": preview.get("row_count", 0),
        "columns": preview.get("columns", []),
        "campaign_name": (campaign_name or base).strip(),
        "daily_budget_usd": daily_budget_usd,
        "uploaded_at": _now_iso(),
    }


def _mirror_first_source(job: AdJob) -> None:
    """Copy the first source's preview fields to top-level mirror fields.

    Legacy callers (list view, status API) read top-level ``file_name``,
    ``ad_type`` etc. Multi-source jobs keep these in sync with ``sources[0]``.
    """
    if not job.sources:
        return
    s = job.sources[0]
    job.file_name = s["file_name"]
    job.file_ext = s["file_ext"]
    job.file_size = s["file_size"]
    job.ad_type = s["ad_type"]
    job.date_range = s["date_range"]
    job.row_count = s["row_count"]
    job.columns = s["columns"]


def _source_raw_path(job_id: str, source: Dict[str, Any]) -> Path:
    """Absolute path to a source's raw report on disk."""
    return _job_dir(job_id) / f"source_{source['source_id']}.{source['file_ext']}"


def upload_report(
    filename: str,
    content: bytes,
    marketplace: str = "US",
    job_id: Optional[str] = None,
) -> AdJob:
    """Persist an uploaded report.

    - ``job_id=None``: create a new job with one source.
    - ``job_id=<existing>``: append a source to an existing ``uploaded`` job
      (multi-file mode). Must not be started yet. Duplicate files (same SHA-1)
      are rejected. Enforces ``MAX_SOURCES`` cap.

    Raises ``ValueError`` on unsupported extension / oversized file /
    unparsable content / duplicate file / cap exceeded / state mismatch.
    Raises ``KeyError`` if appending to a nonexistent job.
    """
    if not filename:
        raise ValueError("filename is required")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(f"file too large (> {MAX_UPLOAD_BYTES // 1024 // 1024} MB)")

    ext = Path(filename).suffix.lower().lstrip(".")
    if ext not in ("xlsx", "xlsm", "csv"):
        raise ValueError(f"unsupported file type: .{ext} (expected xlsx / csv)")

    file_hash = hashlib.sha1(content).hexdigest()

    # Append mode ------------------------------------------------------------
    if job_id:
        meta = _read_meta(job_id)
        if not meta:
            raise KeyError(f"job {job_id} not found")
        if meta.get("status") != "uploaded":
            raise ValueError(
                f"job {job_id} is in state {meta.get('status')}, "
                "只能在未开始的任务上追加文件"
            )
        sources = list(meta.get("sources") or [])
        if any(s.get("file_hash") == file_hash for s in sources):
            raise ValueError("这份文件已经上传过了（SHA-1 相同）")
        if len(sources) >= MAX_SOURCES:
            raise ValueError(f"最多只能上传 {MAX_SOURCES} 份报告")

        jd = _job_dir(job_id)
        source_id = uuid.uuid4().hex[:8]
        raw_path = jd / f"source_{source_id}.{ext}"
        raw_path.write_bytes(content)
        try:
            preview = _preview_report(raw_path, filename)
        except Exception as e:
            raw_path.unlink(missing_ok=True)
            raise ValueError(f"failed to parse report: {e}") from e

        entry = _source_entry(
            source_id=source_id,
            file_name=filename,
            file_ext=ext,
            file_size=len(content),
            file_hash=file_hash,
            preview=preview,
        )
        sources.append(entry)

        job = AdJob(**{k: v for k, v in meta.items() if k in AdJob.__dataclass_fields__})
        job.sources = sources
        _mirror_first_source(job)
        _write_meta(job)
        return job

    # New job ----------------------------------------------------------------
    new_job_id = uuid.uuid4().hex[:12]
    jd = _job_dir(new_job_id)
    jd.mkdir(parents=True, exist_ok=True)
    source_id = uuid.uuid4().hex[:8]
    raw_path = jd / f"source_{source_id}.{ext}"
    raw_path.write_bytes(content)

    try:
        preview = _preview_report(raw_path, filename)
    except Exception as e:
        # Clean up the partial dir so it doesn't leak.
        shutil.rmtree(jd, ignore_errors=True)
        raise ValueError(f"failed to parse report: {e}") from e

    entry = _source_entry(
        source_id=source_id,
        file_name=filename,
        file_ext=ext,
        file_size=len(content),
        file_hash=file_hash,
        preview=preview,
    )

    job = AdJob(
        job_id=new_job_id,
        marketplace=(marketplace or "US").strip().upper(),
        sources=[entry],
        status="uploaded",
    )
    _mirror_first_source(job)
    (jd / "preview.json").write_text(
        json.dumps(preview, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_meta(job)
    return job


def remove_source(job_id: str, source_id: str) -> AdJob:
    """Delete one source file from an uploaded job.

    Raises ``KeyError`` if job/source missing, ``ValueError`` on state
    mismatch or last-source removal (we keep at least one).
    """
    meta = _read_meta(job_id)
    if not meta:
        raise KeyError(f"job {job_id} not found")
    if meta.get("status") != "uploaded":
        raise ValueError(f"job {job_id} is in state {meta.get('status')}, 不能删除源文件")

    sources = list(meta.get("sources") or [])
    idx = next((i for i, s in enumerate(sources) if s.get("source_id") == source_id), None)
    if idx is None:
        raise KeyError(f"source {source_id} not found")
    if len(sources) <= 1:
        raise ValueError("至少保留一份报告；如要取消任务请删除整个 job")

    removed = sources.pop(idx)
    raw_path = _source_raw_path(job_id, removed)
    raw_path.unlink(missing_ok=True)

    job = AdJob(**{k: v for k, v in meta.items() if k in AdJob.__dataclass_fields__})
    job.sources = sources
    _mirror_first_source(job)
    _write_meta(job)
    return job


def update_source(
    job_id: str,
    source_id: str,
    *,
    campaign_name: Optional[str] = None,
    daily_budget_usd: Optional[float] = None,
    clear_daily_budget: bool = False,
) -> AdJob:
    """Rename a source or adjust its daily budget.

    Pass ``clear_daily_budget=True`` to unset the budget (goes back to %
    reallocation for this campaign).
    """
    meta = _read_meta(job_id)
    if not meta:
        raise KeyError(f"job {job_id} not found")
    if meta.get("status") != "uploaded":
        raise ValueError(f"job {job_id} is in state {meta.get('status')}, 不能修改源文件")

    sources = list(meta.get("sources") or [])
    target = next((s for s in sources if s.get("source_id") == source_id), None)
    if target is None:
        raise KeyError(f"source {source_id} not found")

    if campaign_name is not None:
        name = campaign_name.strip()
        if not name:
            raise ValueError("campaign_name 不能为空")
        target["campaign_name"] = name
    if clear_daily_budget:
        target["daily_budget_usd"] = None
    elif daily_budget_usd is not None:
        if daily_budget_usd < 0:
            raise ValueError("daily_budget_usd 必须 >= 0")
        target["daily_budget_usd"] = float(daily_budget_usd)

    job = AdJob(**{k: v for k, v in meta.items() if k in AdJob.__dataclass_fields__})
    job.sources = sources
    _mirror_first_source(job)
    _write_meta(job)
    return job


# --------------------------------------------------------------------------- #
# Prompt assembly
# --------------------------------------------------------------------------- #

# Business rules that apply regardless of goal — these are stable user
# preferences, not task-specific context.
_GLOBAL_RULES = """
- 不建议 SBV 视频广告相关的开启/加码动作（用户不投 SBV）
- 不建议 Amazon Vine 相关动作（用户不走 Vine）
- 建议优先级：CTR/CVR 杠杆（主图 / Listing / A+）> 长尾词辅推 > 出价调整 > 否词
- 不预设产品配置（4G / WiFi / 蜂窝 / 太阳能 / 电池容量 / IP 等级 等）：
  仅以「产品备注」和上传搜索词报告里的实际线索为准；缺信息写"未指定"，禁止编造规格
- 核心大词不享有 ACOS 豁免：除非进入「守护关键词」列表，否则按数据判断
  该降 bid 就降、该否定就否、该暂停就暂停；**报告里禁止出现"核心大词不应否定/降价"
  这类无条件保护语句**——任何针对大词的保留建议必须给出具体数据证据
- 否词建议时：若「守护关键词」列表非空，这些词不得出现在任何否词建议中（不论大词/小词）；
  若列表为空，则完全按数据判断，核心大词也可以建议降价或否（但需给足证据）
""".strip()


_GOAL_POSTURES = {
    "profit": "盈利为主：ACOS 目标 < 盈亏平衡点；任何词（含核心大词）若高花费低转化，立即砍或大幅降价；位置激进收缩；不为'词的体量'或'品牌相关性'保留亏损位置",
    "new_launch": "新品冲量：放宽 ACOS 硬指标；重视曝光、点击和位置（Top of Search）；守核心词位置优先于 ACOS；鼓励测新词",
    "relaunch": "老品重推：中等宽松；关注 CTR/CVR 拉升而非单纯控 ACOS；评估是否存在结构性 Listing 问题",
    "clearance": "清货：最激进；只看订单数和清货速度；ACOS 放到最后；降价、加码高效词、拓宽匹配",
}


def _build_prompt(job: AdJob, source_paths: List[Tuple[Dict[str, Any], Path]]) -> str:
    """Craft the agent prompt.

    ``source_paths`` is a list of ``(source_entry, absolute_path)`` tuples so
    we can render one block per uploaded report — enabling cross-campaign
    insights when the user uploads multiple SP/SB/SD exports for the same ASIN.

    Structure:
    1. Task context (ASIN, goal, operational context)
    2. Per-source block (one per uploaded file)
    3. Global business rules (stable preferences)
    4. Goal-driven threshold posture
    5. Instruction to load the skill and analyse the attached files
    6. Strict output format (markdown + trailing ```json`` block)
    """
    protected = job.protected_keywords or []
    protected_block = (
        "、".join(protected) if protected else "无（完全按数据判断，核心大词也可否/降）"
    )
    goal_posture = _GOAL_POSTURES.get(job.goal, _GOAL_POSTURES["profit"])

    # Source blocks — each independent so SP/SB/SD column differences don't
    # force cross-format merging. Budgets are per-campaign and only injected
    # if the user supplied them.
    source_blocks: List[str] = []
    any_budget = False
    for idx, (src, path) in enumerate(source_paths, start=1):
        campaign = src.get("campaign_name") or src.get("file_name") or f"source-{idx}"
        budget_raw = src.get("daily_budget_usd")
        budget_line = ""
        if budget_raw is not None:
            any_budget = True
            budget_line = f"\n  - 日预算 (USD): ${float(budget_raw):.2f}"
        source_blocks.append(
            f"""### 报告 #{idx} — {campaign}
  - 文件: {path}
  - 广告类型: {src.get("ad_type") or "未识别"}
  - 日期范围: {src.get("date_range") or "未识别"}
  - 行数: {src.get("row_count", 0)}{budget_line}"""
        )

    multi = len(source_paths) > 1
    cross_campaign_note = (
        "\n4. **跨活动对比**：本次共 "
        f"{len(source_paths)} 份报告代表同一 ASIN 下的不同广告活动，请在 "
        "`cross_campaign_insights[]` 中产出至少 3 条跨活动洞察（效率黑洞/预算重分配/关键词迁移等），"
        "并在 `campaign_efficiency` 中把每份报告作为独立 campaign 对比。"
        if multi
        else ""
    )
    budget_note = (
        "\n5. **预算重分配**：运营给出了部分 campaign 的日预算，在 `action_summary` / "
        "`cross_campaign_insights` 中给出具体 $ 数字的加减建议；未给预算的 campaign 只给百分比方向。"
        if any_budget
        else ""
    )

    sources_section = "\n\n".join(source_blocks)

    return f"""请加载并使用 `zach-search-term-report-analyzer` 技能分析下述 Amazon Ads 搜索词报告。

## 任务上下文

- 站点: {job.marketplace}
- 目标 ASIN: {job.asin or "未指定"}
- 运营目标: {job.goal} — {goal_posture}
- 产品备注（来自运营）: {job.product_notes or "未提供"}
- 守护关键词（不得在否词建议中出现）: {protected_block}
- 共上传 {len(source_paths)} 份搜索词报告

## 报告文件（逐份独立读取，列名差异由技能处理）

{sources_section}

## 全局业务规则（用户的稳定偏好）

{_GLOBAL_RULES}

## 分析要求

1. 请根据运营目标的阈值姿态，决定各维度的判定宽严程度。
2. 本报告目标是**可直接落地执行**（不是诊断），请对齐 skill `references/landable_proposal_patterns.md` 中的 8 板块落地结构。
3. 按 skill 方法论产出 markdown 报告，覆盖以下 14 个板块（标 NEW 为落地化新增，没有数据就留空板块但必须出现在 JSON 里）：{cross_campaign_note}{budget_note}
   (1) 总览（曝光/点击/花费/订单/ACOS/CTR/CVR 汇总 + 一句话 verdict，verdict 要具体数字 + 问题性质 + 必要动作）
   (2) Campaign 效率对比 **NEW**（按 campaign 聚合 spend / orders / cost_per_order / acos / spend_share / order_share，每个 campaign 一句 verdict）
   (3) 🎯 守护词与核心词位置诊断（若「守护关键词」为空：本节为 Top-5 流量词位置体检，**所有词均按数据判断，包括降 bid / 否定建议**，不要使用"核心词需保护"等无条件保护表述；若非空：守护词必须全部出现且带 status，仅守护词列表内的词享有不否定豁免）
   (4) 高效词 Top-20（加码候选：给 current_bid → suggested_bid → bid_change_pct "+18%" 这种方向符）
   (5) 低效词 Top-20（降 bid / 暂停 / 否定候选）
   (6) 新增关键词候选（从搜索词挖出的可新建 Exact 的词）
   (7) 否词建议 **加强**（给出 wasted_spend_usd 过去 21 天浪费金额 + window_days；末尾合计 negative_wasted_total_usd 预估直接省金额；自动排除「守护关键词」同义词族）
   (8) 新 Campaign 搭建 **NEW**（"抄作业版"：name / type / match_type / daily_budget_usd / bid_strategy / placement_modifiers / keywords_with_bid / sync_actions / verdict，只在 relaunch / new_launch 目标下输出；profit 目标下可为空）
   (9) 位置诊断（ToS / RoS / Product Pages 对比 + suggested_modifier 如 "+150%" "-100%"）
   (10) 执行 Checklist **升级**（action_summary 每条带 level + day + eta_minutes + location_path 面包屑路径 + 具体 action + evidence + expected_impact）
   (11) 跨活动洞察 **NEW**（仅在多报告场景输出：`cross_campaign_insights[]`，每条含 `insight_type` (black_hole_campaign / budget_reallocation / keyword_migration / match_type_gap / placement_shift) / `summary` / `detail` / `evidence` / `action` / `priority` P0-P2；单报告场景留空数组即可）
   (12) 数据备注
   (13) 元信息

## 输出格式（必须严格遵守）

- 先按上述板块输出完整 markdown 报告
- 所有 `action` 字段、`efficiency_tag` 字段使用下列枚举值（前端按值换色/徽章）：
  - `protected_keywords_status.status`: good / warn / bad
  - `high_performers.action`: boost / watch
  - `low_performers.action`: cut / pause / lower_bid
  - `negative_suggestions.type`: immediate / watch
  - `campaign_efficiency.efficiency_tag`: black_hole / needs_optimization / healthy / high_efficiency
  - `action_summary.level`: P0 / P1 / P2
- `bid_change_pct` 必须是字符串 `+N%` / `-N%` / `0%`，方向符不能省
- 报告结尾追加一段以 ```json 开头、``` 结尾的代码块，内含以下结构化字段（字段不要缺，缺失用空字符串或空数组）：

```json
{{
  "overview": {{
    "ad_type": "",
    "marketplace": "",
    "date_range": "",
    "impressions": 0,
    "clicks": 0,
    "spend": 0,
    "orders": 0,
    "sales": 0,
    "acos": "",
    "ctr": "",
    "cvr": "",
    "one_line_verdict": ""
  }},
  "campaign_efficiency": [
    {{ "campaign_name": "", "type": "", "spend": 0, "spend_share": "", "orders": 0, "order_share": "", "cost_per_order": "", "acos": "", "efficiency_tag": "black_hole|needs_optimization|healthy|high_efficiency", "verdict": "" }}
  ],
  "protected_keywords_status": [
    {{ "keyword": "", "status": "good|warn|bad", "impressions": 0, "clicks": 0, "spend": 0, "orders": 0, "acos": "", "note": "" }}
  ],
  "high_performers": [
    {{ "keyword": "", "match_type": "", "impressions": 0, "clicks": 0, "spend": 0, "orders": 0, "acos": "", "action": "boost|watch", "current_bid": "", "suggested_bid": "", "bid_change_pct": "", "reason": "" }}
  ],
  "low_performers": [
    {{ "keyword": "", "match_type": "", "impressions": 0, "clicks": 0, "spend": 0, "orders": 0, "acos": "", "action": "cut|pause|lower_bid", "current_bid": "", "suggested_bid": "", "bid_change_pct": "", "reason": "" }}
  ],
  "new_keyword_candidates": [
    {{ "keyword": "", "source_search_term": "", "impressions": 0, "orders": 0, "suggested_bid": "", "reason": "" }}
  ],
  "negative_suggestions": [
    {{ "term": "", "type": "immediate|watch", "reason": "", "wasted_spend_usd": 0, "window_days": 0 }}
  ],
  "negative_wasted_total_usd": 0,
  "new_campaigns": [
    {{ "name": "", "type": "", "match_type": "", "daily_budget_usd": 0, "bid_strategy": "", "placement_modifiers": {{ "top_of_search": "", "rest_of_search": "", "product_pages": "" }}, "keywords_with_bid": [ {{ "keyword": "", "bid_usd": 0 }} ], "sync_actions": [], "verdict": "" }}
  ],
  "placement_diagnosis": [
    {{ "placement": "", "impressions": 0, "clicks": 0, "spend": 0, "orders": 0, "acos": "", "ctr": "", "cvr": "", "suggested_modifier": "", "action": "" }}
  ],
  "action_summary": [
    {{ "level": "P0|P1|P2", "day": "", "eta_minutes": 0, "location_path": "", "action": "", "evidence": "", "expected_impact": "" }}
  ],
  "cross_campaign_insights": [
    {{ "insight_type": "black_hole_campaign|budget_reallocation|keyword_migration|match_type_gap|placement_shift", "summary": "", "detail": "", "evidence": "", "action": "", "priority": "P0|P1|P2" }}
  ],
  "data_notes": "",
  "meta": {{
    "analyzed_at": "",
    "row_count": 0,
    "threshold_posture": ""
  }}
}}
```

字段说明：
- verdict 写法：具体数字 + 问题性质 + 必要动作（反例："ACOS 偏高需关注"；正例："ACOS 62% 超目标 22pt，SP-Auto 的 'for hunting' 类长尾占 45% 花费 0 单，立即加否定短语"）
- 新字段：`campaign_efficiency` / `new_campaigns` / `action_summary.day/eta_minutes/location_path` / `high_performers.current_bid/bid_change_pct` / `negative_suggestions.wasted_spend_usd/window_days` / `negative_wasted_total_usd` / `placement_diagnosis.suggested_modifier`
- `day` 推荐用 "Day 1" / "Day 2" / "Day 3-7" / "Day 8-14" 四档；`eta_minutes` 填整数分钟
- `location_path` 写 Seller Central 面包屑，如 "广告活动 → SP-Core-Exact → 关键词定位 → 否定关键词"
- JSON 键用英文，值用中文。除该 JSON 块外，正文保持 markdown 结构不变。
"""


def _extract_xlsx_path_from_output(raw: str) -> Optional[Path]:
    """Extract the xlsx file path from agent stdout for xlsx_plan mode."""
    # Look for absolute paths ending in .xlsx
    for line in reversed(raw.splitlines()):
        line = line.strip()
        if line.startswith("/") and line.endswith(".xlsx"):
            p = Path(line)
            if p.is_file():
                return p
        # Also handle MEDIA: prefix
        if line.startswith("MEDIA:") and ".xlsx" in line:
            p = Path(line.replace("MEDIA:", "").strip())
            if p.is_file():
                return p
    # Fallback: scan ~/.hermes/cache for recent xlsx
    cache = Path.home() / ".hermes" / "cache"
    if cache.is_dir():
        candidates = sorted(cache.glob("*广告优化方案*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if candidates:
            return candidates[0]
    return None


def _build_xlsx_plan_prompt(job: AdJob, source_paths: List[Tuple[Dict[str, Any], Path]]) -> str:
    """Build prompt for the amazon-ad-campaign-optimization-xlsx skill."""
    protected = job.protected_keywords or []
    xlsx_goal = _XLSX_GOAL_MAP.get(job.goal, "profit")
    csv_files = [str(p) for _, p in source_paths]
    csv_list = json.dumps(csv_files, ensure_ascii=False)
    protected_list = json.dumps(protected, ensure_ascii=False)

    return f"""请加载并使用 `amazon-ad-campaign-optimization-xlsx` 技能，为下述 Amazon 广告搜索词报告生成 8-sheet xlsx 优化方案。

## 必填参数

- csv_files: {csv_list}
- asin: {job.asin}
- marketplace: {job.marketplace}
- start_date: （从报告中自动识别）
- end_date: （从报告中自动识别）
- goal: {xlsx_goal}
- protected_keywords: {protected_list}
- product_notes: "{job.product_notes or "未提供"}"

## 说明

请严格按照 skill 的 Workflow（Step 1~6）执行：解析 CSV → SIF 数据增强 → LLM 出 plan.json → 渲染 xlsx → 自检 → 交付。

输出的 xlsx 文件路径请用 MEDIA: 前缀回传。
"""


# --------------------------------------------------------------------------- #
# Subprocess execution
# --------------------------------------------------------------------------- #

async def _run_agent(job: AdJob) -> None:
    jd = _job_dir(job.job_id)
    jd.mkdir(parents=True, exist_ok=True)
    stdout_log = jd / "stdout.log"

    # Build source path list. Multi-source jobs use ``sources[]``; legacy
    # single-file jobs fall back to the old ``raw.<ext>`` layout.
    source_paths: List[Tuple[Dict[str, Any], Path]] = []
    if job.sources:
        for src in job.sources:
            p = _source_raw_path(job.job_id, src)
            if p.is_file():
                source_paths.append((src, p))
    if not source_paths:
        # Legacy fallback — no sources[] or files missing.
        legacy = jd / f"raw.{job.file_ext}"
        if legacy.is_file():
            synthetic = {
                "source_id": "legacy",
                "campaign_name": job.file_name or "legacy",
                "file_name": job.file_name,
                "ad_type": job.ad_type,
                "date_range": job.date_range,
                "row_count": job.row_count,
                "daily_budget_usd": None,
            }
            source_paths.append((synthetic, legacy))

    if not source_paths:
        job.status = "failed"
        job.error = "no source files on disk"
        job.finished_at = _now_iso()
        _write_meta(job)
        return

    runner, runner_bin, err = resolve_with_pref(job.runner_pref)
    if err or not runner or not runner_bin:
        job.status = "failed"
        job.error = err or "no runner resolved"
        job.finished_at = _now_iso()
        _write_meta(job)
        return

    job.runner_used = runner
    child_env = build_child_env(runner_bin)
    prompt = _build_prompt(job, source_paths) if job.output_mode != "xlsx_plan" else _build_xlsx_plan_prompt(job, source_paths)
    cmd = _build_runner_cmd(runner, runner_bin, prompt)

    job.status = "running"
    job.started_at = _now_iso()
    mcp_note = "（MCP: sorftime + sif_mcp）" if runner == "hermes" else ""
    job.progress = f"已启动 {runner} 解析报告{mcp_note}…"
    _write_meta(job)

    start = time.monotonic()
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
            env=child_env,
        )
    except Exception as spawn_err:
        job.status = "failed"
        job.error = f"failed to spawn {runner}: {type(spawn_err).__name__}: {spawn_err}"
        job.finished_at = _now_iso()
        _write_meta(job)
        return

    job.pid = proc.pid
    _write_meta(job)

    buf_bytes = 0
    buf_chunks: List[bytes] = []

    async def _flush() -> None:
        nonlocal buf_bytes
        if not buf_chunks:
            return
        data = b"".join(buf_chunks)
        buf_chunks.clear()
        with stdout_log.open("ab") as f:
            f.write(data)
        job.stdout_bytes += len(data)
        buf_bytes = 0

    timed_out = False
    try:
        while True:
            if time.monotonic() - start > HARD_TIMEOUT_SEC:
                timed_out = True
                try:
                    proc.send_signal(signal.SIGTERM)
                    await asyncio.wait_for(proc.wait(), timeout=5)
                except asyncio.TimeoutError:
                    proc.kill()
                break
            try:
                chunk = await asyncio.wait_for(proc.stdout.read(2048), timeout=2.0)
            except asyncio.TimeoutError:
                elapsed = int(time.monotonic() - start)
                job.progress = f"分析中… {elapsed}s"
                _write_meta(job)
                continue
            if not chunk:
                break
            buf_chunks.append(chunk)
            buf_bytes += len(chunk)
            if buf_bytes >= FLUSH_EVERY_BYTES:
                await _flush()
                kb = job.stdout_bytes // 1024
                job.progress = f"已生成 ~{kb} KB…"
                _write_meta(job)

        await _flush()
        rc = await proc.wait()
        job.finished_at = _now_iso()

        if timed_out:
            job.status = "failed"
            job.error = f"timeout after {HARD_TIMEOUT_SEC}s"
            _write_meta(job)
            return
        if rc != 0:
            job.status = "failed"
            tail = ""
            if stdout_log.is_file():
                tail = stdout_log.read_text(encoding="utf-8", errors="replace")[-800:]
            job.error = f"{runner} exited with code {rc}: {tail.strip()[-400:]}"
            _write_meta(job)
            return

        # Success — split markdown and structured JSON.
        raw = stdout_log.read_text(encoding="utf-8", errors="replace")
        md_text, structured = _split_report_and_json(raw)
        (jd / "report.md").write_text(md_text, encoding="utf-8")
        if structured is not None:
            (jd / "report.json").write_text(
                json.dumps(structured, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        # For xlsx_plan mode, find the generated xlsx from agent output and
        # copy it into the job directory so download_path can serve it.
        if job.output_mode == "xlsx_plan":
            xlsx_path = _extract_xlsx_path_from_output(raw)
            if xlsx_path and xlsx_path.is_file():
                dest = jd / "plan.xlsx"
                shutil.copy2(str(xlsx_path), str(dest))

        job.status = "done"
        job.progress = "完成"
        _write_meta(job)
    except Exception as e:  # pragma: no cover - defensive
        job.status = "failed"
        job.error = f"{type(e).__name__}: {e}"
        job.finished_at = _now_iso()
        _write_meta(job)
        try:
            if proc.returncode is None:
                proc.kill()
        except Exception:
            pass


async def start_job(
    job_id: str,
    goal: str,
    protected_keywords: Optional[List[str]] = None,
    asin: str = "",
    product_notes: str = "",
    runner_pref: str = "auto",
    daily_budgets: Optional[Dict[str, float]] = None,
    output_mode: str = "report",
) -> AdJob:
    """Kick off analysis for a previously-uploaded job.

    The job must exist on disk with status=``uploaded``. Raises
    ``RuntimeError`` if another audit (ASIN or ad) is already running.

    ``daily_budgets`` is a map of ``source_id -> usd/day``. Applied onto each
    source entry so the prompt can inject them. Values <= 0 are ignored.
    """
    if _job_lock.locked():
        raise RuntimeError("another audit is currently running")

    meta = _read_meta(job_id)
    if not meta:
        raise KeyError(f"job {job_id} not found")
    if meta.get("status") != "uploaded":
        raise RuntimeError(
            f"job {job_id} is in state {meta.get('status')}, cannot start"
        )

    if goal not in GOALS:
        raise ValueError(f"goal must be one of {GOALS}, got {goal!r}")

    if output_mode not in OUTPUT_MODES:
        raise ValueError(f"output_mode must be one of {OUTPUT_MODES}, got {output_mode!r}")

    # xlsx_plan mode requires an ASIN always.
    if output_mode == "xlsx_plan" and not (asin or "").strip():
        raise ValueError("xlsx_plan 模式必须填写目标 ASIN")

    # Multi-source jobs need an ASIN so cross-campaign comparison has a
    # shared subject. Single-file legacy jobs also benefit, but we don't
    # force it to avoid breaking existing workflows.
    sources = list(meta.get("sources") or [])
    asin_clean = (asin or "").strip().upper()
    if len(sources) > 1 and not asin_clean:
        raise ValueError("多份报告合并分析必须填写目标 ASIN")

    runner_pref = (runner_pref or "auto").lower()
    # Pre-flight runner availability check — fail fast with a 400.
    _runner, _path, err = resolve_with_pref(runner_pref)
    if err:
        raise RuntimeError(err)

    # Hydrate the full dataclass from disk meta, then overlay the new fields.
    job = AdJob(**{k: v for k, v in meta.items() if k in AdJob.__dataclass_fields__})
    job.goal = goal
    job.output_mode = output_mode
    job.protected_keywords = [w.strip() for w in (protected_keywords or []) if w and w.strip()]
    job.asin = asin_clean
    job.product_notes = (product_notes or "").strip()
    job.runner_pref = runner_pref

    # Apply daily budgets onto source entries. Keyed by source_id so renaming
    # a campaign doesn't lose the link. Values <= 0 clear the budget.
    if daily_budgets:
        budget_map = {str(k): float(v) for k, v in daily_budgets.items()}
        new_sources = []
        for s in job.sources:
            s2 = dict(s)
            if s2.get("source_id") in budget_map:
                val = budget_map[s2["source_id"]]
                s2["daily_budget_usd"] = val if val > 0 else None
            new_sources.append(s2)
        job.sources = new_sources
        job.daily_budgets = {
            k: v for k, v in budget_map.items() if v > 0
        }

    job.status = "queued"
    _live_jobs[job_id] = job
    _write_meta(job)

    async def _runner_task() -> None:
        async with _job_lock:
            try:
                await _run_agent(job)
            except Exception as e:
                try:
                    job.status = "failed"
                    job.error = f"runner crashed: {type(e).__name__}: {e}"
                    job.finished_at = _now_iso()
                    _write_meta(job)
                except Exception:
                    pass
            finally:
                _live_jobs.pop(job.job_id, None)

    asyncio.create_task(_runner_task())
    return job


# --------------------------------------------------------------------------- #
# Read-side helpers
# --------------------------------------------------------------------------- #

def get_job(job_id: str) -> Optional[Dict[str, Any]]:
    meta = _read_meta(job_id)
    if meta is None:
        return None
    jd = _job_dir(job_id)
    result: Dict[str, Any] = {**meta}
    rj = jd / "report.json"
    rm = jd / "report.md"
    if rj.is_file():
        try:
            result["structured"] = json.loads(rj.read_text(encoding="utf-8"))
        except Exception:
            result["structured"] = None
    else:
        result["structured"] = None
    if rm.is_file():
        text = rm.read_text(encoding="utf-8", errors="replace")
        if len(text) > 200_000:
            text = text[:200_000] + "\n\n…(truncated, download for full)…"
        result["raw_md"] = text
    else:
        sl = jd / "stdout.log"
        if sl.is_file():
            raw = sl.read_text(encoding="utf-8", errors="replace")
            result["raw_md"] = raw[-20_000:]
        else:
            result["raw_md"] = ""
    return result


def list_jobs(limit: int = 20) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not AD_AUDIT_ROOT.is_dir():
        return rows
    for entry in sorted(AD_AUDIT_ROOT.iterdir(), key=lambda p: p.name, reverse=True):
        if not entry.is_dir():
            continue
        meta = _read_meta(entry.name)
        if not meta:
            continue
        rows.append(meta)
        if len(rows) >= limit:
            break
    rows.sort(key=lambda r: r.get("created_at", ""), reverse=True)
    return rows[:limit]


def download_path(job_id: str, fmt: str) -> Optional[Path]:
    jd = _job_dir(job_id)
    if fmt == "md":
        fp = jd / "report.md"
        return fp if fp.is_file() else None
    if fmt == "json":
        fp = jd / "report.json"
        return fp if fp.is_file() else None
    if fmt == "xlsx":
        # xlsx_plan mode: serve the pre-built plan.xlsx directly
        plan_xlsx = jd / "plan.xlsx"
        if plan_xlsx.is_file():
            return plan_xlsx
        rj = jd / "report.json"
        if not rj.is_file():
            return None
        xp = jd / "report.xlsx"
        if not xp.is_file() or xp.stat().st_mtime < rj.stat().st_mtime:
            try:
                structured = json.loads(rj.read_text(encoding="utf-8"))
            except Exception:
                logger.exception("ad-audit %s: failed to parse report.json", job_id)
                return None
            meta = _read_meta(job_id) or {}
            try:
                build_xlsx(xp, structured, meta)
            except Exception:
                logger.exception("ad-audit %s: build_xlsx failed", job_id)
                return None
        return xp if xp.is_file() else None
    if fmt == "html":
        rj = jd / "report.json"
        hp = jd / "report.html"
        if rj.is_file():
            if not hp.is_file() or hp.stat().st_mtime < rj.stat().st_mtime:
                from app.services import html_report
                try:
                    structured = json.loads(rj.read_text(encoding="utf-8"))
                except Exception:
                    logger.exception("ad-audit %s: failed to parse report.json (html)", job_id)
                    return None
                meta = _read_meta(job_id) or {}
                try:
                    html_report.build_ad_html(hp, structured, meta)
                except Exception:
                    logger.exception("ad-audit %s: build_ad_html failed", job_id)
                    return None
        elif not hp.is_file():
            # Fallback: render markdown with full styling
            md_path = jd / "report.md"
            if md_path.is_file():
                from app.services import html_report
                md_text = md_path.read_text(encoding="utf-8", errors="replace")
                meta = _read_meta(job_id) or {}
                try:
                    html_report.build_md_html(hp, meta, md_text)
                except Exception:
                    logger.exception("ad-audit %s: build_md_html failed", job_id)
        return hp if hp.is_file() else None
    return None


def is_busy() -> bool:
    return _job_lock.locked()


def sweep_expired() -> int:
    if not AD_AUDIT_ROOT.is_dir():
        return 0
    cutoff = time.time() - RETENTION_SEC
    removed = 0
    for entry in AD_AUDIT_ROOT.iterdir():
        try:
            if entry.is_dir() and entry.stat().st_mtime < cutoff:
                shutil.rmtree(entry, ignore_errors=True)
                removed += 1
        except Exception:
            continue
    return removed


def sweep_stale_running() -> int:
    """Flip running/queued jobs left on disk at startup to failed."""
    if not AD_AUDIT_ROOT.is_dir():
        return 0
    rewritten = 0
    for entry in AD_AUDIT_ROOT.iterdir():
        if not entry.is_dir():
            continue
        meta = _read_meta(entry.name)
        if not meta:
            continue
        if meta.get("status") not in ("running", "queued"):
            continue
        meta["status"] = "failed"
        meta["error"] = meta.get("error") or "服务重启导致任务中断"
        meta["finished_at"] = meta.get("finished_at") or _now_iso()
        try:
            (entry / "meta.json").write_text(
                json.dumps(meta, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            rewritten += 1
        except Exception:
            continue
    return rewritten


def clear_failed() -> int:
    if not AD_AUDIT_ROOT.is_dir():
        return 0
    removed = 0
    for entry in AD_AUDIT_ROOT.iterdir():
        if not entry.is_dir():
            continue
        meta = _read_meta(entry.name)
        if not meta:
            shutil.rmtree(entry, ignore_errors=True)
            removed += 1
            continue
        if meta.get("status") in ("failed", "cancelled"):
            shutil.rmtree(entry, ignore_errors=True)
            removed += 1
    return removed


def delete_job(job_id: str) -> bool:
    """Delete a single job directory. Returns True if deleted."""
    jd = _job_dir(job_id)
    if not jd.is_dir():
        return False
    meta = _read_meta(job_id)
    if meta and meta.get("status") in ("running", "queued"):
        return False
    shutil.rmtree(jd, ignore_errors=True)
    return True


# --------------------------------------------------------------------------- #
# XLSX report generation — 10 sheets with color-coded cells
# --------------------------------------------------------------------------- #

# Color palette for cells (ARGB, no leading #).
_FILL_GOOD = "DFF5E1"    # pale green
_FILL_WARN = "FFF4CC"    # pale yellow
_FILL_BAD  = "FBD4D4"    # pale red
_FILL_BOOST = "D6E8FF"   # pale blue
_FILL_CUT  = "FFE0C2"    # pale orange
_FILL_PAUSE = "E5E5E5"   # pale grey
_FILL_WATCH = "EFE3FF"   # pale violet
_FILL_NEW  = "D8F2EC"    # pale teal
_FILL_P0   = "F8BFBF"    # red
_FILL_P1   = "FFE7A3"    # yellow
_FILL_P2   = "C8DAF2"    # blue
_FILL_HEADER = "1F2A3A"
# Efficiency tags (campaign-level)
_FILL_BLACKHOLE = "FBD4D4"
_FILL_NEEDS_OPT = "FFF4CC"
_FILL_HEALTHY   = "DFF5E1"
_FILL_HIGH_EFF  = "D6E8FF"
# Shield (protected / strategic)
_FILL_SHIELD    = "D6E8FF"

_STATUS_FILL = {"good": _FILL_GOOD, "warn": _FILL_WARN, "bad": _FILL_BAD}
_ACTION_FILL = {
    "boost": _FILL_BOOST,
    "watch": _FILL_WATCH,
    "cut": _FILL_CUT,
    "pause": _FILL_PAUSE,
    "lower_bid": _FILL_CUT,
    "new": _FILL_NEW,
    "immediate": _FILL_CUT,
}
_LEVEL_FILL = {"P0": _FILL_P0, "P1": _FILL_P1, "P2": _FILL_P2}
_EFF_FILL = {
    "black_hole": _FILL_BLACKHOLE,
    "needs_optimization": _FILL_NEEDS_OPT,
    "healthy": _FILL_HEALTHY,
    "high_efficiency": _FILL_HIGH_EFF,
}
_EFF_LABEL = {
    "black_hole": "❌ 效率黑洞",
    "needs_optimization": "⚠️ 需优化",
    "healthy": "✓ 健康",
    "high_efficiency": "✓✓ 高效",
}


def build_xlsx(
    out_path: Path,
    structured: Dict[str, Any],
    meta: Dict[str, Any],
) -> None:
    """Render the structured analysis into a 10-sheet xlsx with color blocks."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor=_FILL_HEADER)
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    def _add_sheet(
        title: str,
        headers: List[str],
        rows: List[List[Any]],
        fills: Optional[List[Optional[Dict[int, str]]]] = None,
        col_widths: Optional[Dict[int, int]] = None,
    ) -> None:
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if not rows:
            ws.append(["（本次报告未提供）"] + [""] * (len(headers) - 1))
        else:
            for i, row in enumerate(rows):
                ws.append(row)
                if fills and i < len(fills) and fills[i]:
                    xrow = i + 2  # +1 header, +1 for 1-based
                    for col_idx, color in fills[i].items():
                        if 0 <= col_idx < len(headers):
                            ws.cell(row=xrow, column=col_idx + 1).fill = (
                                PatternFill("solid", fgColor=color)
                            )
        for i, _ in enumerate(headers, 1):
            w = (col_widths or {}).get(i - 1, 20)
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP

    # ---- Sheet 1: 总览 ----
    ov = structured.get("overview") or {}
    _add_sheet(
        "总览",
        ["字段", "值"],
        [
            ["广告类型", meta.get("ad_type") or ov.get("ad_type", "")],
            ["站点", meta.get("marketplace") or ov.get("marketplace", "")],
            ["日期范围", meta.get("date_range") or ov.get("date_range", "")],
            ["曝光", ov.get("impressions", "")],
            ["点击", ov.get("clicks", "")],
            ["花费", ov.get("spend", "")],
            ["订单", ov.get("orders", "")],
            ["销售额", ov.get("sales", "")],
            ["ACOS", ov.get("acos", "")],
            ["CTR", ov.get("ctr", "")],
            ["CVR", ov.get("cvr", "")],
            ["一句话结论", ov.get("one_line_verdict", "")],
            ["运行 runner", meta.get("runner_used") or meta.get("runner_pref") or ""],
            ["运营目标", meta.get("goal", "")],
            ["目标 ASIN", meta.get("asin", "")],
        ],
        col_widths={0: 16, 1: 60},
    )

    # ---- Sheet: Campaign 效率对比（NEW，紧跟总览） ----
    camp_eff = structured.get("campaign_efficiency") or []
    camp_rows: List[List[Any]] = []
    camp_fills: List[Optional[Dict[int, str]]] = []
    for c in camp_eff:
        if not isinstance(c, dict):
            continue
        eff = str(c.get("efficiency_tag", "")).lower()
        camp_rows.append([
            c.get("campaign_name", ""),
            c.get("type", ""),
            c.get("spend", ""),
            c.get("spend_share", ""),
            c.get("orders", ""),
            c.get("order_share", ""),
            c.get("cost_per_order", ""),
            c.get("acos", ""),
            _EFF_LABEL.get(eff, c.get("efficiency_tag", "")),
            c.get("verdict", ""),
        ])
        color = _EFF_FILL.get(eff)
        camp_fills.append({8: color} if color else None)
    _add_sheet(
        "Campaign 效率对比",
        ["Campaign", "类型", "花费", "预算占比", "订单", "单量占比",
         "每单成本", "ACOS", "效率", "一句话判断"],
        camp_rows,
        fills=camp_fills,
        col_widths={0: 28, 9: 42},
    )

    # ---- Sheet 2: 关键词守位 ----
    protected = structured.get("protected_keywords_status") or []
    rows_pk: List[List[Any]] = []
    fills_pk: List[Optional[Dict[int, str]]] = []
    for p in protected:
        if not isinstance(p, dict):
            continue
        rows_pk.append([
            p.get("keyword", ""),
            p.get("status", ""),
            p.get("impressions", ""),
            p.get("clicks", ""),
            p.get("spend", ""),
            p.get("orders", ""),
            p.get("acos", ""),
            p.get("note", ""),
        ])
        fill_map: Dict[int, str] = {}
        color = _STATUS_FILL.get(str(p.get("status", "")).lower())
        if color:
            fill_map[1] = color
        fills_pk.append(fill_map or None)
    _add_sheet(
        "关键词守位",
        ["关键词", "状态", "曝光", "点击", "花费", "订单", "ACOS", "说明"],
        rows_pk,
        fills=fills_pk,
        col_widths={0: 30, 7: 40},
    )

    def _kw_rows_and_fills(
        items: List[Any],
        action_key: str = "action",
    ) -> Tuple[List[List[Any]], List[Optional[Dict[int, str]]]]:
        """Build kw table rows: keyword | match | clicks | spend | orders | acos |
        action | current_bid | suggested_bid | bid_change_pct | reason (11 cols).
        action column gets colored by action enum; bid_change_pct cell gets
        colored by direction (up/down/flat)."""
        rows: List[List[Any]] = []
        fills: List[Optional[Dict[int, str]]] = []
        for it in items or []:
            if not isinstance(it, dict):
                continue
            change = str(it.get("bid_change_pct", "") or "").strip()
            rows.append([
                it.get("keyword", ""),
                it.get("match_type", ""),
                it.get("clicks", ""),
                it.get("spend", ""),
                it.get("orders", ""),
                it.get("acos", ""),
                it.get(action_key, ""),
                it.get("current_bid", ""),
                it.get("suggested_bid", ""),
                change,
                it.get("reason", ""),
            ])
            fill_map: Dict[int, str] = {}
            color = _ACTION_FILL.get(str(it.get(action_key, "")).lower())
            if color:
                fill_map[6] = color
            if change.startswith("+") and change not in ("+0%", "+0"):
                fill_map[9] = _FILL_GOOD
            elif change.startswith("-") and change not in ("-0%", "-0"):
                fill_map[9] = _FILL_BAD
            fills.append(fill_map or None)
        return rows, fills

    # ---- Sheet 3: 高效词 Top-20 ----
    rows, fills = _kw_rows_and_fills(structured.get("high_performers"))
    _add_sheet(
        "高效词 Top-20",
        ["关键词", "匹配", "点击", "花费", "订单", "ACOS", "动作",
         "当前出价", "建议出价", "变化率", "理由"],
        rows,
        fills=fills,
        col_widths={0: 30, 10: 40},
    )

    # ---- Sheet 4: 低效词 Top-20 ----
    rows, fills = _kw_rows_and_fills(structured.get("low_performers"))
    _add_sheet(
        "低效词 Top-20",
        ["关键词", "匹配", "点击", "花费", "订单", "ACOS", "动作",
         "当前出价", "建议出价", "变化率", "理由"],
        rows,
        fills=fills,
        col_widths={0: 30, 10: 40},
    )

    # ---- Sheet 5: 新增关键词候选 ----
    new_rows: List[List[Any]] = []
    new_fills: List[Optional[Dict[int, str]]] = []
    for it in structured.get("new_keyword_candidates") or []:
        if not isinstance(it, dict):
            continue
        new_rows.append([
            it.get("keyword", ""),
            it.get("source_search_term", ""),
            it.get("impressions", ""),
            it.get("orders", ""),
            it.get("suggested_bid", ""),
            it.get("reason", ""),
        ])
        new_fills.append({0: _FILL_NEW})
    _add_sheet(
        "新增关键词候选",
        ["候选关键词", "源搜索词", "曝光", "订单", "建议出价", "理由"],
        new_rows,
        fills=new_fills,
        col_widths={0: 30, 1: 30, 5: 40},
    )

    # ---- Sheet 6: 否词建议（带浪费金额归因） ----
    neg_rows: List[List[Any]] = []
    neg_fills: List[Optional[Dict[int, str]]] = []
    negs = structured.get("negative_suggestions") or []
    has_wasted = any(
        isinstance(n, dict) and n.get("wasted_spend_usd") for n in negs
    )
    for it in negs:
        if isinstance(it, dict):
            term = it.get("term", "")
            typ = str(it.get("type", "")).lower()
            reason = it.get("reason", "")
            wasted = it.get("wasted_spend_usd", "") if has_wasted else None
            win = it.get("window_days", "") if has_wasted else None
        elif isinstance(it, str):
            term, typ, reason = it, "immediate", ""
            wasted = None
            win = None
        else:
            continue
        row = [term, typ]
        if has_wasted:
            row.extend([wasted or "", win or ""])
        row.append(reason)
        neg_rows.append(row)
        color = _ACTION_FILL.get(typ)
        fmap = {1: color} if color else {}
        if has_wasted and wasted:
            fmap[2] = _FILL_BAD
        neg_fills.append(fmap or None)
    # Append grand-total row if wasted totals present
    total_save = structured.get("negative_wasted_total_usd") or 0
    if has_wasted and total_save:
        total_row = ["💰 合计预计直接省", "", total_save, "", ""]
        neg_rows.append(total_row)
        neg_fills.append({0: _FILL_GOOD, 2: _FILL_GOOD})
    neg_headers = ["否词", "类型"]
    if has_wasted:
        neg_headers.extend(["过去浪费 $", "窗口天数"])
    neg_headers.append("理由")
    _add_sheet(
        "否词建议",
        neg_headers,
        neg_rows,
        fills=neg_fills,
        col_widths={0: 30, len(neg_headers) - 1: 50},
    )

    # ---- Sheet: 新 Campaign 搭建（NEW） ----
    new_camps = structured.get("new_campaigns") or []
    nc_rows: List[List[Any]] = []
    nc_fills: List[Optional[Dict[int, str]]] = []
    for c in new_camps:
        if not isinstance(c, dict):
            continue
        pm = c.get("placement_modifiers") or {}
        kw_items = c.get("keywords_with_bid") or []
        kw_text = "\n".join(
            f"· {kw.get('keyword','')} → ${kw.get('bid_usd','')}"
            for kw in kw_items if isinstance(kw, dict)
        )
        sync = c.get("sync_actions") or []
        sync_text = "\n".join(f"· {s}" for s in sync)
        nc_rows.append([
            c.get("name", ""),
            c.get("type", ""),
            c.get("match_type", ""),
            c.get("daily_budget_usd", ""),
            c.get("bid_strategy", ""),
            pm.get("top_of_search", ""),
            pm.get("rest_of_search", ""),
            pm.get("product_pages", ""),
            kw_text,
            sync_text,
            c.get("verdict", ""),
        ])
        nc_fills.append({0: _FILL_NEW})
    _add_sheet(
        "新 Campaign 搭建",
        ["Campaign 名称", "类型", "匹配", "日预算 $", "竞价策略",
         "搜索顶部", "搜索其余", "商品页", "关键词 + 出价", "同步动作", "一句话判断"],
        nc_rows,
        fills=nc_fills,
        col_widths={0: 30, 8: 40, 9: 36, 10: 40},
    )

    # ---- Sheet 7: 位置诊断（含 suggested_modifier） ----
    place_rows: List[List[Any]] = []
    place_fills: List[Optional[Dict[int, str]]] = []
    for p in structured.get("placement_diagnosis") or []:
        if not isinstance(p, dict):
            continue
        mod = str(p.get("suggested_modifier", "") or "").strip()
        place_rows.append([
            p.get("placement", ""),
            p.get("clicks", ""),
            p.get("spend", ""),
            p.get("orders", ""),
            p.get("acos", ""),
            p.get("ctr", ""),
            p.get("cvr", ""),
            mod,
            p.get("action", ""),
        ])
        fmap: Dict[int, str] = {}
        if mod.startswith("+") and mod not in ("+0%", "+0"):
            fmap[7] = _FILL_GOOD
        elif mod.startswith("-") and mod not in ("-0%", "-0"):
            fmap[7] = _FILL_BAD
        place_fills.append(fmap or None)
    _add_sheet(
        "位置诊断",
        ["位置", "点击", "花费", "订单", "ACOS", "CTR", "CVR",
         "建议溢价", "调整建议"],
        place_rows,
        fills=place_fills,
        col_widths={0: 18, 8: 42},
    )

    # ---- Sheet: 跨活动洞察（仅多报告场景 NEW） ----
    cross_insights = structured.get("cross_campaign_insights") or []
    if cross_insights:
        ci_rows: List[List[Any]] = []
        ci_fills: List[Optional[Dict[int, str]]] = []
        type_label = {
            "black_hole_campaign": "🕳 黑洞活动",
            "budget_reallocation": "💰 预算重分配",
            "keyword_migration":   "➡ 关键词迁移",
            "match_type_gap":      "◇ 匹配缺口",
            "placement_shift":     "📍 位置调整",
        }
        type_fill = {
            "black_hole_campaign": _FILL_BAD,
            "budget_reallocation": _FILL_GOOD,
            "keyword_migration":   _FILL_NEW,
            "match_type_gap":      _FILL_WATCH,
            "placement_shift":     _FILL_WATCH,
        }
        for it in cross_insights:
            if not isinstance(it, dict):
                continue
            t = str(it.get("insight_type", "")).lower()
            ci_rows.append([
                type_label.get(t, t or "—"),
                it.get("from_campaign", ""),
                it.get("to_campaign", ""),
                it.get("summary", ""),
                it.get("detail", ""),
                it.get("evidence", ""),
                it.get("suggested_action", ""),
            ])
            color = type_fill.get(t)
            ci_fills.append({0: color} if color else None)
        _add_sheet(
            "跨活动洞察",
            ["类型", "来源活动", "目标活动", "摘要", "详情", "证据", "建议动作"],
            ci_rows,
            fills=ci_fills,
            col_widths={0: 18, 1: 24, 2: 24, 3: 36, 4: 40, 5: 40, 6: 40},
        )

    # ---- Sheet 8: 执行 Checklist（含 day / eta / path） ----
    act_rows: List[List[Any]] = []
    act_fills: List[Optional[Dict[int, str]]] = []
    for a in structured.get("action_summary") or []:
        if not isinstance(a, dict):
            continue
        lvl = str(a.get("level", "")).upper()
        act_rows.append([
            "☐",
            a.get("day", "") or "未排期",
            lvl,
            a.get("action", ""),
            a.get("eta_minutes", "") or "",
            a.get("location_path", ""),
            a.get("evidence", ""),
            a.get("expected_impact", ""),
        ])
        color = _LEVEL_FILL.get(lvl)
        act_fills.append({2: color} if color else None)
    _add_sheet(
        "执行 Checklist",
        ["☐", "排期", "优先级", "动作", "耗时 min", "位置路径", "依据", "预期影响"],
        act_rows,
        fills=act_fills,
        col_widths={0: 4, 1: 12, 2: 10, 3: 40, 4: 10, 5: 34, 6: 40, 7: 30},
    )

    # ---- Sheet 9: 原始数据摘要 ----
    data_notes = structured.get("data_notes") or ""
    _add_sheet(
        "原始数据摘要",
        ["说明"],
        [[data_notes]] if data_notes else [],
        col_widths={0: 100},
    )

    # ---- Sheet 10: 元信息 ----
    m = structured.get("meta") or {}
    _add_sheet(
        "元信息",
        ["字段", "值"],
        [
            ["分析时间", m.get("analyzed_at", "") or meta.get("finished_at", "")],
            ["数据量（行）", m.get("row_count", "") or meta.get("row_count", "")],
            ["阈值姿态", m.get("threshold_posture", "") or meta.get("goal", "")],
            ["Job ID", meta.get("job_id", "")],
            ["文件名", meta.get("file_name", "")],
            ["守护关键词", "、".join(meta.get("protected_keywords", []) or [])],
            ["产品备注", meta.get("product_notes", "")],
        ],
        col_widths={0: 20, 1: 60},
    )

    wb.save(out_path)
