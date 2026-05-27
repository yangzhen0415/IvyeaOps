from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
# Default paths are overridden by the router; these are fallback-only.
DEFAULT_QUOTE_FOLDER = ROOT / "报价表收集箱"
DEFAULT_DATA_FILE = ROOT / "data" / "normalized_quotes.json"

MAX_SCAN_ROWS = 3000
MAX_SCAN_COLS = 90

IGNORE_SHEET_KEYWORDS = (
    "目录",
    "注意",
    "附加",
    "发货",
    "船期",
    "偏远",
    "罚款",
    "反倾销",
    "品牌",
    "收货",
    "查验",
    "海外仓",
)

WAREHOUSE_HEADER_KEYS = ("仓库代码简称", "仓库代码", "FBA仓", "产品名称")
PRICE_TIER_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(KG|KGS|CBM|方)\s*\+", re.I)
FBA_TOKEN_RE = re.compile(r"[A-Za-z0-9]{3,6}")
ZIP_RE = re.compile(r"\(([^)]{4,20})\)")
DATE_RE = re.compile(r"(20\d{2})[./-]?(1[0-2]|0?[1-9])[./-]?(3[01]|[12]\d|0?[1-9])")
PRODUCT_CODE_RE = re.compile(r"(?:销售)?产品代码[：:]\s*([^/，,；;]+(?:\s*/\s*[^/，,；;]+)?)")


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " / ").replace("\r", " ").strip()


def compact(value: str) -> str:
    return re.sub(r"\s+", "", value or "")


def as_number(value: Any) -> float | None:
    raw = text(value)
    if not raw:
        return None
    bad = ("-", "/", "无服务", "暂停", "待定", "#REF", "#N/A", "渠道无此仓")
    if any(flag in raw for flag in bad):
        return None
    cleaned = raw.replace(",", "").replace("¥", "").replace("￥", "")
    cleaned = re.sub(r"(RMB|CNY|USD|元|/KG|/kg|/CBM|/cbm)", "", cleaned).strip()
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
        return None
    value = float(cleaned)
    return round(value, 4)


def nice_number(value: float | None) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def looks_like_warehouse_code(token: str) -> bool:
    token = token.upper().strip()
    if not (3 <= len(token) <= 6):
        return False
    if token in {"KG", "CBM", "KGS", "USA", "UPS", "FBA", "WMT", "ETD", "ETA"}:
        return False
    if token.endswith("KG") or token.endswith("CBM"):
        return False
    has_letter = any(ch.isalpha() for ch in token)
    has_digit = any(ch.isdigit() for ch in token)
    return has_letter and (has_digit or len(token) == 4)


def extract_warehouse_code(value: Any) -> str:
    codes = extract_warehouse_codes(value)
    return codes[0] if codes else ""


def extract_warehouse_codes(value: Any) -> list[str]:
    raw = text(value).upper()
    out: list[str] = []
    for token in FBA_TOKEN_RE.findall(raw):
        if looks_like_warehouse_code(token):
            token = token.upper()
            if token not in out:
                out.append(token)
    return out


def extract_zip(value: Any) -> str:
    raw = text(value)
    match = ZIP_RE.search(raw)
    if match:
        return match.group(1).strip()
    return ""


def infer_company(path: Path) -> str:
    name = path.stem
    if "凯琦" in name:
        return "凯琦"
    if "安君" in name:
        return "安君国际"
    if "盈和" in name:
        return "盈和国际"
    return re.sub(r"[-_ ]?\d{4}.*$", "", name).strip() or name


def extract_date(path: Path, rows: list[list[str]] | None = None) -> str:
    haystack = path.stem
    if rows:
        for row in rows[:80]:
            joined = " ".join(row[:20])
            if "生效" in joined or "版本" in joined:
                haystack += " " + joined
    match = DATE_RE.search(haystack)
    if not match:
        return ""
    y, m, d = match.groups()
    return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"


def should_skip_sheet(sheet_name: str) -> bool:
    if sheet_name in {"汇总表", "美国FBA仓库产品匹配表", "快捷查询价格表", "价格快速查询"}:
        return False
    return any(key in sheet_name for key in IGNORE_SHEET_KEYWORDS)


def read_sheet_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    max_col = min(ws.max_column or 0, MAX_SCAN_COLS)
    blank_streak = 0
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row or 0, MAX_SCAN_ROWS),
        max_col=max_col,
        values_only=True,
    ):
        parsed = [text(v) for v in row]
        if any(parsed):
            blank_streak = 0
            rows.append(parsed)
        else:
            blank_streak += 1
            rows.append(parsed)
            if len(rows) > 40 and blank_streak >= 80:
                break
    return rows


def filled_header_rows(header_rows: list[list[str]], width: int) -> list[list[str]]:
    filled: list[list[str]] = []
    for row in header_rows:
        out: list[str] = []
        last = ""
        for idx in range(width):
            val = row[idx] if idx < len(row) else ""
            if val:
                last = val
            out.append(last)
        filled.append(out)
    return filled


def find_warehouse_col(row: list[str], next_rows: list[list[str]]) -> int | None:
    candidates: list[tuple[int, int]] = []
    for idx, value in enumerate(row):
        value_c = compact(value)
        if not value_c:
            continue
        score = 0
        if "仓库代码简称" in value_c:
            score = 100
        elif "仓库代码" in value_c:
            score = 90
        elif "FBA仓" in value_c:
            score = 80
        elif value_c == "产品名称":
            score = 50
        if score:
            candidates.append((score, idx))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    for _, idx in candidates:
        hits = 0
        for r in next_rows[:12]:
            if idx < len(r) and extract_warehouse_code(r[idx]):
                hits += 1
        if hits:
            return idx
    return None


def is_price_col(path: list[str]) -> bool:
    joined = " ".join(path)
    joined_c = compact(joined)
    if not joined_c:
        return False
    if any(key in joined_c for key in ("时效", "航期", "船期", "截单", "预计", "备注", "邮编", "区域")):
        if not PRICE_TIER_RE.search(joined_c) and "报价" not in joined_c:
            return False
    return bool(
        PRICE_TIER_RE.search(joined_c)
        or "1CBM" in joined_c.upper()
        or "包税报价" in joined_c
        or "不包税报价" in joined_c
    )


def infer_tier(path: list[str]) -> str:
    for item in reversed(path):
        item_c = compact(item).upper()
        if PRICE_TIER_RE.search(item_c) or "1CBM" in item_c or "5CBM" in item_c:
            return item.strip()
    return ""


def infer_unit(tier: str, path: list[str]) -> str:
    joined = compact(" ".join([tier, *path])).upper()
    if "CBM" in joined or "方" in joined:
        return "CBM"
    if "KG" in joined:
        return "KG"
    return ""


def infer_tax_type(path: list[str]) -> str:
    joined = compact(" ".join(path))
    if "不含税" in joined or "不包税" in joined:
        return "不含税/不包税"
    if "自税" in joined:
        return "自税"
    if "含税" in joined or "包税" in joined:
        return "含税/包税"
    return ""


def infer_origin(path: list[str]) -> str:
    origin_keys = (
        "华东",
        "华南",
        "福建",
        "青岛",
        "深圳",
        "广州",
        "中山",
        "汕头",
        "义乌",
        "宁波",
        "苏州",
        "厦门",
        "泉州",
        "天津",
        "河北",
    )
    for item in path:
        if any(key in item for key in origin_keys):
            return item.strip()
    return ""


def infer_channel(path: list[str], sheet_name: str, row_product: str = "") -> str:
    if row_product and not extract_warehouse_code(row_product):
        return row_product
    ignored = ("含税", "不含税", "包税", "不包税", "自税", "KG", "CBM", "返回目录", "进入报价表")
    candidates: list[str] = []
    for item in path:
        item = item.strip()
        item_c = compact(item)
        if not item_c:
            continue
        if "材积重" in item_c or "产品代码" in item_c:
            continue
        if any(flag in item_c for flag in ignored):
            continue
        if any(flag in item_c for flag in ("华东", "华南", "福建", "青岛", "深圳", "广州", "义乌", "宁波", "苏州", "厦门", "泉州")):
            continue
        if "仓库代码" in item_c or "FBA仓" in item_c or "产品名称" in item_c:
            continue
        candidates.append(item)
    return candidates[-1] if candidates else sheet_name


def infer_product_code_from_path(path: list[str]) -> str:
    for item in path:
        if "产品代码" not in item:
            continue
        match = PRODUCT_CODE_RE.search(item)
        if not match:
            continue
        tail = match.group(1).strip()
        codes = re.findall(r"(?:[A-Z]{1,4}-[A-Z0-9]{1,8}|CP[A-Z0-9]{4,})", tail, flags=re.I)
        return " / ".join(code.upper() for code in codes[:6]) if codes else tail
    return ""


def find_meta_cols(header_paths: dict[int, list[str]], warehouse_col: int) -> dict[str, list[int]]:
    meta = {"product_code": [], "transit": [], "note": [], "zip": [], "product_name": []}
    for col, path in header_paths.items():
        if col == warehouse_col or is_price_col(path):
            continue
        joined = compact(" ".join(path))
        if "销售产品代码" in joined or "产品代码" in joined:
            meta["product_code"].append(col)
        if any(key in joined for key in ("参考时效", "理赔时效", "时效", "航期", "船期", "ETA", "开船")):
            meta["transit"].append(col)
        if "备注" in joined or "说明" in joined:
            meta["note"].append(col)
        if "邮编" in joined:
            meta["zip"].append(col)
        if "产品名称" in joined:
            meta["product_name"].append(col)
    return meta


def row_value(row: list[str], cols: Iterable[int]) -> str:
    for col in cols:
        if col < len(row) and row[col]:
            return row[col]
    return ""


def parse_section(
    *,
    rows: list[list[str]],
    header_idx: int,
    warehouse_col: int,
    company: str,
    market: str,
    file_path: Path,
    sheet_name: str,
    effective_date: str,
) -> list[dict[str, Any]]:
    width = max((len(r) for r in rows[max(0, header_idx - 4) : header_idx + 3]), default=0)
    width = min(width, MAX_SCAN_COLS)
    header_start = max(0, header_idx - 4)
    header_end = header_idx + 1
    if header_idx + 1 < len(rows):
        next_joined = compact(" ".join(rows[header_idx + 1]))
        if any(key in next_joined for key in ("华东", "华南", "福建", "青岛", "含税", "不含税", "KG+", "CBM+")):
            header_end = header_idx + 2
    raw_header_rows = rows[header_start:header_end]
    header_rows = filled_header_rows(raw_header_rows, width)
    header_paths = {
        col: [r[col] for r in header_rows if col < len(r) and r[col]]
        for col in range(width)
    }
    raw_header_paths = {
        col: [r[col] for r in raw_header_rows if col < len(r) and r[col]]
        for col in range(width)
    }
    price_cols = [col for col, path in header_paths.items() if col != warehouse_col and is_price_col(path)]
    if not price_cols:
        return []
    meta_cols = find_meta_cols(raw_header_paths, warehouse_col)
    records: list[dict[str, Any]] = []
    last_product_code = ""
    last_product_name = ""
    for row_idx in range(header_end, len(rows)):
        row = rows[row_idx]
        if warehouse_col >= len(row):
            continue
        codes = extract_warehouse_codes(row[warehouse_col])
        if not codes:
            continue
        full_code = row[warehouse_col]
        product_code = row_value(row, meta_cols["product_code"])
        if product_code:
            last_product_code = product_code
        else:
            product_code = last_product_code
        row_product = row_value(row, meta_cols["product_name"])
        if row_product and not extract_warehouse_code(row_product):
            last_product_name = row_product
        else:
            row_product = last_product_name
        transit = row_value(row, meta_cols["transit"])
        note = row_value(row, meta_cols["note"])
        postal = extract_zip(full_code) or row_value(row, meta_cols["zip"])
        for col in price_cols:
            if col >= len(row):
                continue
            value = as_number(row[col])
            if value is None:
                continue
            path = header_paths.get(col, [])
            tier = infer_tier(path)
            unit = infer_unit(tier, path)
            for code in codes:
                record = {
                    "company": company,
                    "market": market,
                    "source_file": file_path.name,
                    "source_path": str(file_path),
                    "source_mtime": datetime.fromtimestamp(file_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                    "effective_date": effective_date,
                    "sheet": sheet_name,
                    "warehouse_code": code,
                    "warehouse_full": full_code,
                    "postal_code": postal,
                    "channel": infer_channel(path, sheet_name, row_product),
                    "origin": infer_origin(path),
                    "tier": tier,
                    "unit": unit,
                    "tax_type": infer_tax_type(path),
                    "price": nice_number(value),
                    "price_value": value,
                    "product_code": product_code or infer_product_code_from_path(path),
                    "transit": transit,
                    "note": note,
                    "row": row_idx + 1,
                    "column": col + 1,
                    "template": "matrix",
                }
                records.append(record)
    return records


def dedupe(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[Any, ...]] = set()
    out: list[dict[str, Any]] = []
    for rec in records:
        key = (
            rec["company"],
            rec["source_file"],
            rec["sheet"],
            rec["warehouse_code"],
            rec["channel"],
            rec["origin"],
            rec["tier"],
            rec["tax_type"],
            rec["price"],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)
    return out


def parse_workbook(path: Path, company_override: str = "", market: str = "") -> list[dict[str, Any]]:
    company = company_override.strip() or infer_company(path)
    records: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    workbook_effective = ""
    for ws in workbook.worksheets:
        if should_skip_sheet(ws.title):
            continue
        rows = read_sheet_rows(ws)
        if not rows:
            continue
        effective_date = workbook_effective or extract_date(path, rows)
        if effective_date and not workbook_effective:
            workbook_effective = effective_date
        for idx, row in enumerate(rows):
            warehouse_col = find_warehouse_col(row, rows[idx + 1 : idx + 15])
            if warehouse_col is None:
                continue
            records.extend(
                parse_section(
                    rows=rows,
                    header_idx=idx,
                    warehouse_col=warehouse_col,
                    company=company,
                    market=market,
                    file_path=path,
                    sheet_name=ws.title,
                    effective_date=effective_date,
                )
            )
    return dedupe(records)


def scan_folder(
    folder: Path,
    disabled_files: set[str] | None = None,
    file_metadata: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    disabled_files = disabled_files or set()
    file_metadata = file_metadata or {}
    records: list[dict[str, Any]] = []
    files: list[dict[str, Any]] = []
    for path in sorted(folder.glob("*.xls*")):
        if path.name.startswith("~$"):
            continue
        rel_key = path.relative_to(folder).as_posix()
        meta = file_metadata.get(rel_key) or file_metadata.get(path.name) or {}
        company = str(meta.get("company") or "").strip() or infer_company(path)
        market = str(meta.get("market") or "").strip()
        profile = str(meta.get("profile") or "auto").strip() or "auto"
        source = str(meta.get("source") or ("manual" if str(meta.get("company") or "").strip() else "filename")).strip()
        confidence = meta.get("confidence", 1.0 if source in {"manual", "upload-default", "saved-rule"} else 0.55)
        file_info = {
            "name": path.name,
            "relativePath": rel_key,
            "size": path.stat().st_size,
            "modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            "company": company,
            "market": market,
            "profile": profile,
            "companySource": source,
            "recognitionSource": source,
            "recognitionConfidence": confidence,
            "needsReview": bool(meta.get("needsReview", False)),
            "records": 0,
            "error": "",
            "disabled": False,
        }
        if path.name in disabled_files or rel_key in disabled_files:
            file_info["disabled"] = True
            files.append(file_info)
            continue
        try:
            parsed = parse_workbook(path, company_override=company, market=market)
            file_info["records"] = len(parsed)
            records.extend(parsed)
        except Exception as exc:  # Keep one bad vendor file from blocking the weekly batch.
            file_info["error"] = f"{type(exc).__name__}: {exc}"
        files.append(file_info)
    records = dedupe(records)
    return records, files


def build_index(
    folder: Path = DEFAULT_QUOTE_FOLDER,
    out_file: Path = DEFAULT_DATA_FILE,
    disabled_files: set[str] | None = None,
    file_metadata: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    folder = folder.resolve()
    out_file.parent.mkdir(parents=True, exist_ok=True)
    records, files = scan_folder(folder, disabled_files=disabled_files, file_metadata=file_metadata)
    companies = sorted({r["company"] for r in records})
    warehouses = sorted({r["warehouse_code"] for r in records})
    payload = {
        "built_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "quote_folder": str(folder),
        "record_count": len(records),
        "warehouse_count": len(warehouses),
        "companies": companies,
        "files": files,
        "records": sorted(records, key=lambda r: (r["warehouse_code"], r["price_value"], r["company"], r["channel"])),
    }
    out_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def load_index(data_file: Path = DEFAULT_DATA_FILE) -> dict[str, Any]:
    if not data_file.exists():
        return build_index()
    return json.loads(data_file.read_text(encoding="utf-8"))


def query_records(code: str, data_file: Path = DEFAULT_DATA_FILE) -> list[dict[str, Any]]:
    code = code.strip().upper()
    if not code:
        return []
    data = load_index(data_file)
    return [rec for rec in data.get("records", []) if rec.get("warehouse_code", "").upper() == code]


def export_csv(records: list[dict[str, Any]], out_path: Path) -> None:
    fields = [
        "company",
        "channel",
        "sheet",
        "warehouse_code",
        "warehouse_full",
        "origin",
        "tier",
        "unit",
        "tax_type",
        "price",
        "product_code",
        "transit",
        "note",
        "effective_date",
        "source_file",
    ]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for rec in records:
            writer.writerow({field: rec.get(field, "") for field in fields})


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize and query first-leg freight quote workbooks.")
    sub = parser.add_subparsers(dest="cmd", required=True)
    rebuild = sub.add_parser("rebuild")
    rebuild.add_argument("--folder", type=Path, default=DEFAULT_QUOTE_FOLDER)
    rebuild.add_argument("--out", type=Path, default=DEFAULT_DATA_FILE)
    query = sub.add_parser("query")
    query.add_argument("code")
    query.add_argument("--data", type=Path, default=DEFAULT_DATA_FILE)
    query.add_argument("--csv", type=Path)
    args = parser.parse_args()
    if args.cmd == "rebuild":
        payload = build_index(args.folder, args.out)
        print(json.dumps({k: payload[k] for k in ("built_at", "record_count", "warehouse_count", "companies")}, ensure_ascii=False, indent=2))
    elif args.cmd == "query":
        records = query_records(args.code, args.data)
        if args.csv:
            export_csv(records, args.csv)
            print(f"exported {len(records)} rows -> {args.csv}")
        else:
            print(json.dumps(records[:100], ensure_ascii=False, indent=2))
            print(f"total={len(records)}")


if __name__ == "__main__":
    main()
