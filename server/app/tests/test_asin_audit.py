"""Basic tests for ASIN audit routes.

These don't actually spawn claude — they mock start_job to verify the
routing, validation, and download plumbing.
"""
from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient


@pytest.fixture
def client(tmp_path, monkeypatch):
    # Redirect audit root to a tmp dir so tests don't pollute real data.
    monkeypatch.setenv("IVYEA_OPS_DEV_MODE", "1")
    monkeypatch.setenv("IVYEA_OPS_AUTH_DISABLED", "1")
    # Import lazily so env vars take effect.
    from app.main import app
    from app.services import asin_audit

    asin_audit.AUDIT_ROOT = tmp_path / "audits"
    asin_audit.AUDIT_ROOT.mkdir(parents=True, exist_ok=True)
    with TestClient(app) as c:
        yield c


def test_split_report_and_json_extracts_tail_block():
    from app.services.asin_audit import _split_report_and_json

    raw = (
        "# Report\n\nbody text\n\n"
        "```json\n"
        '{"overview":{"asin":"B01"},"scorecard":[]}\n'
        "```\n"
    )
    md, data = _split_report_and_json(raw)
    assert "```json" not in md
    assert data is not None
    assert data["overview"]["asin"] == "B01"


def test_split_report_and_json_missing_returns_raw():
    from app.services.asin_audit import _split_report_and_json

    raw = "# Report\n\njust markdown no json\n"
    md, data = _split_report_and_json(raw)
    assert md == raw
    assert data is None


def test_list_jobs_returns_empty(tmp_path, monkeypatch):
    from app.services import asin_audit

    monkeypatch.setattr(asin_audit, "AUDIT_ROOT", tmp_path / "empty")
    (tmp_path / "empty").mkdir()
    assert asin_audit.list_jobs() == []


def test_list_jobs_reads_meta(tmp_path, monkeypatch):
    from app.services import asin_audit

    root = tmp_path / "r"
    root.mkdir()
    monkeypatch.setattr(asin_audit, "AUDIT_ROOT", root)
    jd = root / "abc123"
    jd.mkdir()
    (jd / "meta.json").write_text(
        json.dumps(
            {
                "job_id": "abc123",
                "asin": "B01",
                "marketplace": "US",
                "mode": "full",
                "status": "done",
                "created_at": "2026-05-12T10:00:00+00:00",
            }
        ),
        encoding="utf-8",
    )
    rows = asin_audit.list_jobs()
    assert len(rows) == 1
    assert rows[0]["job_id"] == "abc123"
    assert rows[0]["status"] == "done"


def test_download_path_md_and_json(tmp_path, monkeypatch):
    from app.services import asin_audit

    root = tmp_path / "r2"
    root.mkdir()
    monkeypatch.setattr(asin_audit, "AUDIT_ROOT", root)
    jd = root / "xyz"
    jd.mkdir()
    (jd / "report.md").write_text("# md", encoding="utf-8")
    (jd / "report.json").write_text("{}", encoding="utf-8")

    assert asin_audit.download_path("xyz", "md") == jd / "report.md"
    assert asin_audit.download_path("xyz", "json") == jd / "report.json"
    assert asin_audit.download_path("xyz", "csv") is None
    assert asin_audit.download_path("missing", "md") is None


def test_download_xlsx_builds_from_structured_json(tmp_path, monkeypatch):
    """xlsx should be generated on-demand from report.json with all sheets."""
    from openpyxl import load_workbook

    from app.services import asin_audit

    root = tmp_path / "r3"
    root.mkdir()
    monkeypatch.setattr(asin_audit, "AUDIT_ROOT", root)
    jd = root / "xlsxjob"
    jd.mkdir()
    structured = {
        "overview": {
            "asin": "B0TEST1234",
            "marketplace": "US",
            "category": "Trail Cameras",
            "title_summary": "Solar 4G cellular cam",
            "key_specs": "4G, Solar, IP66",
            "top_risk": "主图合规风险",
        },
        "scorecard": [
            {"dimension": "语义检索匹配度", "score": 6, "note": "缺少核心关键词"},
            {"dimension": "COSMO 知识图谱对齐度", "score": 5, "note": "未建立同义簇"},
        ],
        "priorities": [
            {"level": "P0", "issue": "主图含文字", "evidence": "违反主图规则", "action": "重拍主图"},
        ],
        "ad_plan": {
            "campaigns": [
                {"name": "SP-Exact-核心", "type": "SP", "targeting": "Exact",
                 "bid_range": "$0.8-1.2", "budget": "$20/d", "strategy": "守位"},
            ],
            "keywords_exact": [
                {"keyword": "cellular trail camera", "bid": "$1.10", "reason": "核心词"},
            ],
            "keywords_phrase_broad": [],
            "negatives_immediate": [{"term": "battery", "reason": "低转化"}],
            "negatives_watch": ["wifi camera"],
        },
        "rewrites": {
            "title": "Solar Cellular Trail Camera 4G LTE 24MP",
            "bullets": ["要点1", "要点2", "要点3", "要点4", "要点5"],
            "qa": [{"q": "支不支持太阳能？", "a": "内置太阳能板"}],
            "backend_terms": "trail cam solar 4g",
            "image_plan": {"main_image": ["透明底展示"]},
            "aplus_plan": ["场景对比模块"],
            "compliance_reminders": ["不做医疗宣称"],
        },
    }
    (jd / "report.json").write_text(
        json.dumps(structured, ensure_ascii=False), encoding="utf-8"
    )
    (jd / "meta.json").write_text(
        json.dumps({
            "job_id": "xlsxjob", "asin": "B0TEST1234", "marketplace": "US",
            "mode": "full", "status": "done",
            "created_at": "2026-05-12T10:00:00+00:00",
            "finished_at": "2026-05-12T10:05:00+00:00",
            "runner_used": "hermes",
        }),
        encoding="utf-8",
    )

    xp = asin_audit.download_path("xlsxjob", "xlsx")
    assert xp is not None and xp.is_file()
    assert xp.name == "report.xlsx"

    wb = load_workbook(xp)
    expected = ["概览", "七维评分", "语义盲区", "COSMO节点", "Rufus问答",
                "用户行为信号", "竞品差异化", "优先级改进", "广告活动",
                "关键词-Exact", "关键词-Phrase", "否定词", "改写稿"]
    assert wb.sheetnames == expected
    # Spot-check content plumbed through correctly.
    assert wb["概览"]["B2"].value == "B0TEST1234"
    assert wb["七维评分"]["A2"].value == "语义检索匹配度"
    assert wb["广告活动"]["A2"].value == "SP-Exact-核心"
    # keywords_phrase_broad was empty → placeholder row.
    assert wb["关键词-Phrase"]["A2"].value == "（本次报告未提供）"


def test_download_html_renders_self_contained(tmp_path, monkeypatch):
    """ASIN audit HTML export should be standalone, include md body + palette."""
    from app.services import asin_audit

    root = tmp_path / "r_html"
    root.mkdir()
    monkeypatch.setattr(asin_audit, "AUDIT_ROOT", root)
    jd = root / "htmlasin"
    jd.mkdir()
    raw_md = (
        "# ASIN B0TEST1234 审计\n\n"
        "## 概览\n\n产品是 solar cellular trail camera。\n\n"
        "## 亮点\n\n- 4G 全网通\n- IP66 防护\n- 太阳能板\n\n"
        "| 维度 | 分数 |\n| --- | --- |\n| 关键词匹配 | 6 |\n| 图片合规 | 3 |\n"
    )
    structured = {
        "asin": "B0TEST1234",
        "marketplace": "US",
        "category": "Trail Cameras",
        "verdict": "中等健康度，主图合规需优先处理",
        "strengths": ["4G 全网通", "IP66 防护"],
        "recommendations": [
            {"level": "P0", "issue": "主图含文字", "action": "重拍主图"},
        ],
    }
    (jd / "report.md").write_text(raw_md, encoding="utf-8")
    (jd / "report.json").write_text(
        json.dumps(structured, ensure_ascii=False), encoding="utf-8",
    )
    (jd / "meta.json").write_text(
        json.dumps({
            "job_id": "htmlasin", "asin": "B0TEST1234", "marketplace": "US",
            "mode": "full", "status": "done", "runner_used": "hermes",
            "created_at": "2026-05-12T10:00:00+00:00",
            "finished_at": "2026-05-12T10:05:00+00:00",
        }),
        encoding="utf-8",
    )

    hp = asin_audit.download_path("htmlasin", "html")
    assert hp is not None and hp.is_file() and hp.name == "report.html"
    text = hp.read_text(encoding="utf-8")
    assert text.startswith("<!DOCTYPE html>")
    assert "<style>" in text
    # Self-contained: CSS is inlined (look for a known selector)
    assert ".ov-grid" in text and ".md" in text
    # Markdown body rendered to HTML (tables + lists)
    assert "<h1" in text or "<h2" in text
    assert "solar cellular trail camera" in text
    assert "4G 全网通" in text
    # Structured overview tiles + verdict
    assert "B0TEST1234" in text
    assert "主图合规" in text or "主图含文字" in text
    # No fmt=html falls through when neither md nor json exists.
    empty_jd = root / "emptyjob"
    empty_jd.mkdir()
    assert asin_audit.download_path("emptyjob", "html") is None


def test_runner_status_lists_all_three(monkeypatch):
    """runner_status should include auto + the three canonical runners."""
    from app.services import asin_audit

    rows = asin_audit.runner_status()
    names = [r["name"] for r in rows]
    assert names[0] == "auto"
    for n in ("hermes", "codex", "claude"):
        assert n in names


def test_build_runner_cmd_per_runner():
    from app.services.asin_audit import _build_runner_cmd

    assert _build_runner_cmd("hermes", "/bin/hermes", "hi")[:2] == ["/bin/hermes", "-z"]
    assert _build_runner_cmd("codex", "/bin/codex", "hi")[:2] == ["/bin/codex", "exec"]
    cc = _build_runner_cmd("claude", "/bin/claude", "hi")
    assert "--print" in cc and "bypassPermissions" in cc


def test_sweep_stale_running_rewrites_ghost_jobs(tmp_path, monkeypatch):
    """A crash/restart leaves status=running on disk with no live subprocess;
    sweep_stale_running must flip them to failed so the UI stops spinning."""
    from app.services import asin_audit

    root = tmp_path / "stale"
    root.mkdir()
    monkeypatch.setattr(asin_audit, "AUDIT_ROOT", root)

    # Ghost running job.
    ghost = root / "ghostjob"
    ghost.mkdir()
    (ghost / "meta.json").write_text(
        json.dumps({
            "job_id": "ghostjob", "asin": "B0GHOSTXXX", "marketplace": "US",
            "mode": "full", "status": "running",
            "created_at": "2026-05-12T10:00:00+00:00",
            "started_at": "2026-05-12T10:00:05+00:00",
        }),
        encoding="utf-8",
    )

    # Ghost queued job (also stale — was never picked up).
    queued = root / "queuedjob"
    queued.mkdir()
    (queued / "meta.json").write_text(
        json.dumps({
            "job_id": "queuedjob", "asin": "B0QUEUEXXX", "marketplace": "US",
            "mode": "full", "status": "queued",
            "created_at": "2026-05-12T10:01:00+00:00",
        }),
        encoding="utf-8",
    )

    # Already-done job must be left alone.
    done = root / "donejob"
    done.mkdir()
    (done / "meta.json").write_text(
        json.dumps({
            "job_id": "donejob", "asin": "B0DONEXXXX", "marketplace": "US",
            "mode": "full", "status": "done",
            "created_at": "2026-05-12T09:00:00+00:00",
            "finished_at": "2026-05-12T09:10:00+00:00",
        }),
        encoding="utf-8",
    )

    assert asin_audit.sweep_stale_running() == 2

    g = json.loads((ghost / "meta.json").read_text(encoding="utf-8"))
    assert g["status"] == "failed"
    assert "重启" in (g.get("error") or "")
    assert g.get("finished_at")  # was filled in

    q = json.loads((queued / "meta.json").read_text(encoding="utf-8"))
    assert q["status"] == "failed"

    # untouched.
    d = json.loads((done / "meta.json").read_text(encoding="utf-8"))
    assert d["status"] == "done"

    # idempotent: second sweep has nothing to do.
    assert asin_audit.sweep_stale_running() == 0
