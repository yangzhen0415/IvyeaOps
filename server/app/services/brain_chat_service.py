"""Knowledge upload and persistent chat service for the GBrain UI."""
from __future__ import annotations

from app.core.proc import no_window_kwargs

import csv
import io
import json
import os
import re
import shutil
import sqlite3
import subprocess
import textwrap
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from app.core.config import settings
from app.services import gbrain_service as gb

DB_PATH = Path(os.environ.get("IVYEA_OPS_BRAIN_CHAT_DB", str(settings.data_dir / "brain_chat.sqlite3")))
MAX_UPLOAD_BYTES = int(os.environ.get("BRAIN_UPLOAD_MAX_BYTES", str(10 * 1024 * 1024)))
ALLOWED_UPLOAD_EXTS = {".md", ".txt", ".csv", ".json", ".xlsx", ".pdf"}
ALLOWED_CATEGORIES = {"inbox", "amazon", "products", "market", "ads", "compliance", "suppliers"}
MAX_CHAT_CHARS = 8000
MAX_INGEST_TEXT_CHARS = int(os.environ.get("BRAIN_INGEST_TEXT_MAX_CHARS", "200000"))


class BrainChatError(RuntimeError):
    """User-facing error for upload/chat failures."""


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db() -> None:
    with _connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS brain_chat_sessions (
              id TEXT PRIMARY KEY,
              title TEXT NOT NULL,
              mode TEXT NOT NULL DEFAULT 'knowledge',
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              archived INTEGER NOT NULL DEFAULT 0,
              last_message_preview TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS brain_chat_messages (
              id TEXT PRIMARY KEY,
              session_id TEXT NOT NULL,
              role TEXT NOT NULL,
              content TEXT NOT NULL,
              citations_json TEXT NOT NULL DEFAULT '[]',
              created_at TEXT NOT NULL,
              FOREIGN KEY(session_id) REFERENCES brain_chat_sessions(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_brain_chat_messages_session_created
              ON brain_chat_messages(session_id, created_at);
            CREATE INDEX IF NOT EXISTS idx_brain_chat_sessions_updated
              ON brain_chat_sessions(archived, updated_at DESC);
            CREATE TABLE IF NOT EXISTS brain_uploads (
              id TEXT PRIMARY KEY,
              source_file TEXT NOT NULL,
              saved_path TEXT NOT NULL,
              category TEXT NOT NULL,
              size INTEGER NOT NULL,
              import_status TEXT NOT NULL,
              warnings_json TEXT NOT NULL DEFAULT '[]',
              created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_brain_uploads_created
              ON brain_uploads(created_at DESC);
            """
        )


def _row_dict(row: sqlite3.Row) -> dict[str, Any]:
    return dict(row)


def list_sessions(include_archived: bool = False) -> dict[str, Any]:
    init_db()
    sql = "SELECT * FROM brain_chat_sessions"
    params: list[Any] = []
    if not include_archived:
        sql += " WHERE archived = 0"
    sql += " ORDER BY updated_at DESC LIMIT 100"
    with _connect() as conn:
        rows = [_row_dict(r) for r in conn.execute(sql, params).fetchall()]
    for r in rows:
        r["archived"] = bool(r["archived"])
    return {"sessions": rows}


def create_session(title: str | None = None, mode: str = "knowledge") -> dict[str, Any]:
    init_db()
    sid = uuid.uuid4().hex
    ts = _now()
    clean_title = (title or "新知识对话").strip()[:80] or "新知识对话"
    clean_mode = mode if mode in {"knowledge", "general", "amazon_operator"} else "knowledge"
    with _connect() as conn:
        conn.execute(
            "INSERT INTO brain_chat_sessions(id, title, mode, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
            (sid, clean_title, clean_mode, ts, ts),
        )
    return get_session(sid)


def get_session(session_id: str) -> dict[str, Any]:
    init_db()
    with _connect() as conn:
        sess = conn.execute("SELECT * FROM brain_chat_sessions WHERE id = ?", (session_id,)).fetchone()
        if not sess:
            raise BrainChatError("会话不存在")
        msgs = conn.execute(
            "SELECT * FROM brain_chat_messages WHERE session_id = ? ORDER BY created_at ASC",
            (session_id,),
        ).fetchall()
    session = _row_dict(sess)
    session["archived"] = bool(session["archived"])
    messages = []
    for row in msgs:
        item = _row_dict(row)
        try:
            item["citations"] = json.loads(item.pop("citations_json") or "[]")
        except json.JSONDecodeError:
            item["citations"] = []
        messages.append(item)
    return {"session": session, "messages": messages}


def update_session(session_id: str, title: str | None = None, archived: bool | None = None) -> dict[str, Any]:
    init_db()
    updates: list[str] = []
    params: list[Any] = []
    if title is not None:
        updates.append("title = ?")
        params.append(title.strip()[:80] or "未命名会话")
    if archived is not None:
        updates.append("archived = ?")
        params.append(1 if archived else 0)
    if not updates:
        return get_session(session_id)
    updates.append("updated_at = ?")
    params.append(_now())
    params.append(session_id)
    with _connect() as conn:
        cur = conn.execute(f"UPDATE brain_chat_sessions SET {', '.join(updates)} WHERE id = ?", params)
        if cur.rowcount == 0:
            raise BrainChatError("会话不存在")
    return get_session(session_id)


def _insert_message(conn: sqlite3.Connection, session_id: str, role: str, content: str, citations: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    mid = uuid.uuid4().hex
    ts = _now()
    clean = content.strip()
    conn.execute(
        "INSERT INTO brain_chat_messages(id, session_id, role, content, citations_json, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (mid, session_id, role, clean, json.dumps(citations or [], ensure_ascii=False), ts),
    )
    conn.execute(
        "UPDATE brain_chat_sessions SET updated_at = ?, last_message_preview = ? WHERE id = ?",
        (ts, clean.replace("\n", " ")[:120], session_id),
    )
    return {"id": mid, "session_id": session_id, "role": role, "content": clean, "citations": citations or [], "created_at": ts}


def _slugify(value: str) -> str:
    v = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff._-]+", "-", value.strip())
    v = re.sub(r"-+", "-", v).strip("-._")
    return v[:80] or "upload"


def _safe_ingest_dir(directory: str | None) -> str:
    """Normalize an auto-selected knowledge directory under BRAIN_ROOT.

    The pasted-text analyzer is allowed to create new directories, but never
    hidden paths, absolute paths, or path traversal. Suspicious model output
    falls back to inbox instead of being creatively rewritten.
    """
    original = (directory or "inbox").replace("\\", "/").strip()
    if not original or "\x00" in original or original.startswith("/"):
        return "inbox"
    raw = original.strip("/")
    if any(part in {".", ".."} or part.startswith(".") for part in raw.split("/")):
        return "inbox"
    parts: list[str] = []
    for part in raw.split("/"):
        p = _slugify(part)
        if not p or p in {".", ".."} or p.startswith("."):
            continue
        parts.append(p[:48])
    if not parts:
        return "inbox"
    safe = "/".join(parts[:4])
    target = (gb.BRAIN_ROOT / safe).resolve()
    try:
        target.relative_to(gb.BRAIN_ROOT)
    except ValueError:
        return "inbox"
    return safe


def _safe_tags(tags: Any) -> list[str]:
    if not isinstance(tags, list):
        return []
    out: list[str] = []
    for tag in tags[:8]:
        clean = _slugify(str(tag).strip().lower())[:32]
        if clean and clean not in out:
            out.append(clean)
    return out


def _first_heading_or_line(text: str) -> str:
    for line in text.splitlines():
        s = line.strip().strip("# ").strip()
        if 4 <= len(s) <= 80:
            return s
    compact = re.sub(r"\s+", " ", text).strip()
    return compact[:60] or "粘贴知识"


def _fallback_ingest_analysis(text: str) -> dict[str, Any]:
    lower = text.lower()
    directory = "inbox"
    tags = ["paste"]
    content_type = "note"
    if any(x in lower for x in ["acos", "campaign", "广告", "关键词", "否词", "ctr", "cvr"]):
        directory = "amazon/ads"
        tags.append("ads")
        content_type = "amazon_ads"
    elif any(x in lower for x in ["asin", "listing", "a+", "产品", "sku", "变体"]):
        directory = "amazon/products"
        tags.append("product")
        content_type = "amazon_product"
    elif any(x in lower for x in ["售后", "customer", "buyer", "refund", "review", "评价", "延保", "站内信"]):
        directory = "amazon/messages"
        tags.extend(["message", "compliance"])
        content_type = "buyer_message"
    elif any(x in lower for x in ["供应商", "1688", "报价", "采购", "工厂"]):
        directory = "amazon/suppliers"
        tags.append("supplier")
        content_type = "supplier_note"
    elif any(x in lower for x in ["市场", "竞品", "竞争", "类目", "搜索量"]):
        directory = "amazon/market"
        tags.append("market")
        content_type = "market_note"
    title = _first_heading_or_line(text)
    summary_source = re.sub(r"\s+", " ", text).strip()
    return {
        "title": title[:80] or "粘贴知识",
        "directory": directory,
        "tags": tags,
        "summary": summary_source[:180] + ("..." if len(summary_source) > 180 else ""),
        "content_type": content_type,
        "confidence": 0.35,
        "source": "rules_fallback",
    }


def _extract_json_object(text: str) -> dict[str, Any] | None:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        parsed = json.loads(cleaned)
        return parsed if isinstance(parsed, dict) else None
    except json.JSONDecodeError:
        pass
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start >= 0 and end > start:
        try:
            parsed = json.loads(cleaned[start : end + 1])
            return parsed if isinstance(parsed, dict) else None
        except json.JSONDecodeError:
            return None
    return None


def _call_hermes_json(prompt: str, timeout: int = 90) -> dict[str, Any] | None:
    if os.environ.get("BRAIN_INGEST_DISABLE_HERMES", "").lower() in {"1", "true", "yes"}:
        return None
    hermes = _hermes_bin()
    cmd = [hermes, "chat", "-q", prompt, "-Q", "--source", "IvyeaOps-web-brain-ingest", "--max-turns", "1", "--toolsets", ""]
    proc = subprocess.run(
        cmd,
        cwd=str(gb.BRAIN_ROOT),
        env=_hermes_env(),
        text=True,
        capture_output=True,
        timeout=timeout,
        **no_window_kwargs(),
    )
    if proc.returncode != 0:
        detail = (proc.stderr or proc.stdout or "").strip()[-800:]
        raise BrainChatError(f"Hermes 分析失败：{detail or '未知错误'}")
    return _extract_json_object(_strip_hermes_output(proc.stdout))


def analyze_pasted_text(text: str) -> dict[str, Any]:
    clean = (text or "").strip()
    if not clean:
        raise BrainChatError("粘贴内容不能为空")
    if len(clean) > MAX_INGEST_TEXT_CHARS:
        raise BrainChatError(f"粘贴内容过长，最多 {MAX_INGEST_TEXT_CHARS} 字符")
    warnings: list[str] = []
    analysis: dict[str, Any] | None = None
    prompt = textwrap.dedent(
        f"""\
        你是 IvyeaOps 私有知识库的入库分类器。不要调用任何工具，不要解释，只返回严格 JSON。

        请分析用户粘贴的文本，自动生成：
        - title: 适合做 Markdown 标题的中文短标题，最多 40 字
        - directory: 应保存到 /root/brain 下的相对目录。可新建目录，但必须安全、简短，例如 inbox、amazon/ads、amazon/products、amazon/messages、amazon/suppliers、amazon/market、compliance
        - tags: 3-6 个短标签，只用中文、英文、数字或短横线
        - summary: 80-160 字中文摘要
        - content_type: note / amazon_ads / amazon_product / buyer_message / supplier_note / market_note / compliance
        - confidence: 0 到 1

        只返回 JSON 对象，不要使用 Markdown 代码块。

        【待入库文本】
        {clean[:12000]}
        """
    ).strip()
    try:
        analysis = _call_hermes_json(prompt)
    except Exception as e:
        warnings.append(f"Hermes 自动分析失败，已使用规则兜底：{e}")
    if not analysis:
        analysis = _fallback_ingest_analysis(clean)
    else:
        analysis["source"] = "hermes_json"

    fallback = _fallback_ingest_analysis(clean)
    title = str(analysis.get("title") or fallback["title"]).strip()[:80] or fallback["title"]
    directory = _safe_ingest_dir(str(analysis.get("directory") or fallback["directory"]))
    tags = _safe_tags(analysis.get("tags")) or fallback["tags"]
    summary = str(analysis.get("summary") or fallback["summary"]).strip()[:500]
    content_type = _slugify(str(analysis.get("content_type") or fallback["content_type"])).lower()[:40] or "note"
    try:
        confidence = float(analysis.get("confidence", fallback["confidence"]))
    except (TypeError, ValueError):
        confidence = fallback["confidence"]
    return {
        "title": title,
        "directory": directory,
        "tags": tags,
        "summary": summary,
        "content_type": content_type,
        "confidence": max(0.0, min(1.0, confidence)),
        "source": analysis.get("source") or "rules_fallback",
        "warnings": warnings,
    }


def _safe_category(category: str | None) -> str:
    c = (category or "inbox").strip().lower()
    return c if c in ALLOWED_CATEGORIES else "inbox"


def _unique_path(path: Path) -> Path:
    candidate = path
    i = 2
    while candidate.exists():
        candidate = path.with_name(f"{path.stem}-{i}{path.suffix}")
        i += 1
    return candidate


def _frontmatter(title: str, source: str, category: str, ext: str) -> str:
    return textwrap.dedent(
        f"""\
        ---
        type: upload
        source_file: {source}
        uploaded_at: {_now()}
        category: {category}
        original_format: {ext}
        tags:
          - upload
        ---

        # {title}

        ## 来源信息

        - 文件名：{source}
        - 上传时间：{_now()}
        - 原始格式：{ext}
        - 分类：{category}

        """
    )


def _decode_text(data: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _csv_to_markdown(data: bytes) -> tuple[str, list[str]]:
    text = _decode_text(data)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    warnings: list[str] = []
    if len(rows) > 120:
        warnings.append(f"CSV 共 {len(rows)} 行，Markdown 预览只保留前 120 行。")
        rows = rows[:120]
    if not rows:
        return "空 CSV 文件。\n", warnings
    width = min(max(len(r) for r in rows), 12)
    normalized = [(r + [""] * width)[:width] for r in rows]
    head = normalized[0]
    body = normalized[1:]
    lines = ["| " + " | ".join(h or f"列{i+1}" for i, h in enumerate(head)) + " |", "| " + " | ".join(["---"] * width) + " |"]
    for row in body:
        lines.append("| " + " | ".join(str(c).replace("\n", " ")[:200] for c in row) + " |")
    return "\n".join(lines) + "\n", warnings


def _xlsx_to_markdown(data: bytes) -> tuple[str, list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    warnings: list[str] = []
    parts: list[str] = []
    for ws in wb.worksheets[:8]:
        parts.append(f"### Sheet: {ws.title}\n")
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx > 80:
                warnings.append(f"Sheet {ws.title} 超过 80 行，已截断。")
                break
            rows.append(["" if c is None else str(c) for c in row[:12]])
        if not rows:
            parts.append("空 sheet。\n")
            continue
        width = min(max(len(r) for r in rows), 12)
        normalized = [(r + [""] * width)[:width] for r in rows]
        head = normalized[0]
        lines = ["| " + " | ".join(h or f"列{i+1}" for i, h in enumerate(head)) + " |", "| " + " | ".join(["---"] * width) + " |"]
        for row in normalized[1:]:
            lines.append("| " + " | ".join(c.replace("\n", " ")[:200] for c in row) + " |")
        parts.append("\n".join(lines) + "\n")
    return "\n".join(parts), warnings


def _pdf_to_markdown(data: bytes) -> tuple[str, list[str]]:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as e:
        raise BrainChatError("PDF 文本提取需要安装 pypdf；请安装依赖后重试。") from e
    reader = PdfReader(io.BytesIO(data))
    warnings: list[str] = []
    parts: list[str] = []
    for i, page in enumerate(reader.pages[:30], start=1):
        txt = page.extract_text() or ""
        parts.append(f"## Page {i}\n\n{txt.strip()}\n")
    if len(reader.pages) > 30:
        warnings.append(f"PDF 共 {len(reader.pages)} 页，已只提取前 30 页。")
    text = "\n".join(parts).strip()
    if not text:
        warnings.append("没有提取到文本；扫描件 PDF 需要后续 OCR 支持。")
        text = "未提取到可用文本。"
    return text + "\n", warnings


def _convert_to_markdown(filename: str, data: bytes, title: str, category: str) -> tuple[str, list[str]]:
    ext = Path(filename).suffix.lower()
    warnings: list[str] = []
    head = _frontmatter(title, filename, category, ext)
    if ext == ".md":
        body = _decode_text(data)
    elif ext == ".txt":
        body = _decode_text(data)
    elif ext == ".csv":
        body, warnings = _csv_to_markdown(data)
    elif ext == ".json":
        parsed = json.loads(_decode_text(data))
        body = "```json\n" + json.dumps(parsed, ensure_ascii=False, indent=2)[:300000] + "\n```\n"
    elif ext == ".xlsx":
        body, warnings = _xlsx_to_markdown(data)
    elif ext == ".pdf":
        body, warnings = _pdf_to_markdown(data)
    else:
        raise BrainChatError("不支持的文件类型")
    return head + "## 原文转换\n\n" + body.strip() + "\n", warnings


def upload_knowledge(filename: str, data: bytes, category: str | None = None, title: str | None = None, import_after_save: bool = True) -> dict[str, Any]:
    init_db()
    if not filename:
        raise BrainChatError("文件名为空")
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTS:
        raise BrainChatError(f"不支持的文件类型：{ext}")
    if len(data) > MAX_UPLOAD_BYTES:
        raise BrainChatError(f"文件超过 {MAX_UPLOAD_BYTES // 1024 // 1024}MB 限制")
    cat = _safe_category(category)
    clean_title = (title or Path(filename).stem).strip()[:120] or Path(filename).stem
    markdown, warnings = _convert_to_markdown(filename, data, clean_title, cat)
    date_prefix = datetime.now().strftime("%Y-%m-%d")
    if cat == "inbox":
        parent = gb.BRAIN_ROOT / "inbox"
    else:
        parent = gb.BRAIN_ROOT / cat / "uploads"
    parent.mkdir(parents=True, exist_ok=True)
    target = _unique_path(parent / f"{date_prefix}-{_slugify(clean_title)}.md")
    target.write_text(markdown, encoding="utf-8")
    rel = str(target.relative_to(gb.BRAIN_ROOT))
    import_status = "skipped"
    import_raw = ""
    if import_after_save:
        try:
            imp = gb.import_brain()
            import_status = "ok"
            import_raw = imp.get("raw", "")
        except Exception as e:
            import_status = f"failed: {e}"
            warnings.append("文件已保存，但自动导入失败；可稍后手动重新导入。")
    uid = uuid.uuid4().hex
    with _connect() as conn:
        conn.execute(
            "INSERT INTO brain_uploads(id, source_file, saved_path, category, size, import_status, warnings_json, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (uid, filename, rel, cat, len(data), import_status, json.dumps(warnings, ensure_ascii=False), _now()),
        )
    return {
        "id": uid,
        "saved_path": rel,
        "category": cat,
        "size": len(data),
        "markdown_preview": markdown[:4000],
        "import_status": import_status,
        "import_raw": import_raw[:2000],
        "warnings": warnings,
    }


def _pasted_markdown(text: str, analysis: dict[str, Any]) -> str:
    tags = analysis.get("tags") or []
    tag_lines = "\n".join(f"  - {tag}" for tag in tags) or "  - paste"
    return textwrap.dedent(
        f"""\
        ---
        type: pasted_text
        uploaded_at: {_now()}
        category: {analysis.get('directory', 'inbox')}
        content_type: {analysis.get('content_type', 'note')}
        analysis_source: {analysis.get('source', 'unknown')}
        tags:
        {tag_lines}
        ---

        # {analysis.get('title') or '粘贴知识'}

        ## 自动摘要

        {analysis.get('summary') or '暂无摘要。'}

        ## 入库信息

        - 来源：Web 端粘贴文本
        - 入库时间：{_now()}
        - 目录：{analysis.get('directory', 'inbox')}
        - 分类置信度：{analysis.get('confidence', 0)}

        ## 原文

        {text.strip()}
        """
    ).strip() + "\n"


def ingest_pasted_text(text: str, import_after_save: bool = True) -> dict[str, Any]:
    init_db()
    clean = (text or "").strip()
    if not clean:
        raise BrainChatError("粘贴内容不能为空")
    if len(clean) > MAX_INGEST_TEXT_CHARS:
        raise BrainChatError(f"粘贴内容过长，最多 {MAX_INGEST_TEXT_CHARS} 字符")
    analysis = analyze_pasted_text(clean)
    directory = _safe_ingest_dir(str(analysis.get("directory") or "inbox"))
    analysis["directory"] = directory
    parent = (gb.BRAIN_ROOT / directory).resolve()
    try:
        parent.relative_to(gb.BRAIN_ROOT)
    except ValueError as e:
        raise BrainChatError("自动目录不安全，已拒绝保存") from e
    parent.mkdir(parents=True, exist_ok=True)
    date_prefix = datetime.now().strftime("%Y-%m-%d")
    target = _unique_path(parent / f"{date_prefix}-{_slugify(str(analysis.get('title') or '粘贴知识'))}.md")
    markdown = _pasted_markdown(clean, analysis)
    target.write_text(markdown, encoding="utf-8")
    rel = str(target.relative_to(gb.BRAIN_ROOT))
    warnings = list(analysis.get("warnings") or [])
    import_status = "skipped"
    import_raw = ""
    if import_after_save:
        try:
            imp = gb.import_brain()
            import_status = "ok"
            import_raw = imp.get("raw", "")
        except Exception as e:
            import_status = f"failed: {e}"
            warnings.append("内容已保存，但自动导入失败；可稍后手动重新导入。")
    uid = uuid.uuid4().hex
    encoded_size = len(clean.encode("utf-8"))
    with _connect() as conn:
        conn.execute(
            "INSERT INTO brain_uploads(id, source_file, saved_path, category, size, import_status, warnings_json, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (uid, "pasted-text", rel, directory, encoded_size, import_status, json.dumps(warnings, ensure_ascii=False), _now()),
        )
    analysis["warnings"] = warnings
    return {
        "id": uid,
        "saved_path": rel,
        "category": directory,
        "size": encoded_size,
        "analysis": analysis,
        "markdown_preview": markdown[:4000],
        "import_status": import_status,
        "import_raw": import_raw[:2000],
        "warnings": warnings,
    }


def list_uploads(limit: int = 50) -> dict[str, Any]:
    init_db()
    with _connect() as conn:
        rows = conn.execute("SELECT * FROM brain_uploads ORDER BY created_at DESC LIMIT ?", (min(max(limit, 1), 100),)).fetchall()
    items = []
    for row in rows:
        item = _row_dict(row)
        try:
            item["warnings"] = json.loads(item.pop("warnings_json") or "[]")
        except json.JSONDecodeError:
            item["warnings"] = []
        items.append(item)
    return {"uploads": items}


def _hermes_bin() -> str:
    from app.core import integrations
    legacy = os.environ.get("BRAIN_CHAT_HERMES_BIN", "").strip()
    if legacy and Path(legacy).exists():
        return legacy
    resolved = integrations.hermes_bin()
    if resolved:
        return resolved
    raise BrainChatError("Hermes CLI 不可用：没有找到 hermes 可执行文件。")


def _hermes_env() -> dict[str, str]:
    from app.core import integrations
    env = os.environ.copy()
    extra_paths = [*integrations.extra_path_dirs(), "/usr/local/bin", "/usr/bin"]
    env["PATH"] = ":".join(extra_paths + [env.get("PATH", "")])
    env.setdefault("PYTHONUNBUFFERED", "1")
    env.setdefault("HERMES_ACCEPT_HOOKS", "1")
    return env


def chat_model_status() -> dict[str, Any]:
    try:
        hermes = _hermes_bin()
    except BrainChatError:
        return {"configured": False, "provider": "hermes", "base_url": "", "model": "", "hermes_bin": "", "mode": "hermes-cli"}
    return {
        "configured": True,
        "provider": "hermes",
        "base_url": "",
        "model": "Hermes Agent",
        "hermes_bin": hermes,
        "mode": "hermes-cli",
    }


def _messages_to_hermes_prompt(messages: list[dict[str, str]]) -> str:
    system = "\n\n".join(m.get("content", "") for m in messages if m.get("role") == "system").strip()
    user = "\n\n".join(m.get("content", "") for m in messages if m.get("role") == "user").strip()
    return textwrap.dedent(
        f"""\
        你正在作为 IvyeaOps Web 知识库对话的回答引擎。
        重要限制：这不是开发任务，不要执行工具、命令、文件读写、联网搜索或系统操作；只基于下面提供的知识库片段和用户问题生成最终回答。

        {system}

        【用户问题】
        {user}
        """
    ).strip()


def _strip_hermes_output(output: str) -> str:
    lines = [line for line in output.splitlines() if not line.strip().startswith("session_id:")]
    text = "\n".join(lines).strip()
    if not text:
        raise BrainChatError("Hermes 返回了空响应。")
    return text


def _call_llm(messages: list[dict[str, str]]) -> str:
    hermes = _hermes_bin()
    prompt = _messages_to_hermes_prompt(messages)
    cmd = [
        hermes,
        "chat",
        "-q",
        prompt,
        "-Q",
        "--source",
        "IvyeaOps-web-brain",
        "--max-turns",
        "1",
        "--toolsets",
        "",
    ]
    try:
        proc = subprocess.run(
            cmd,
            cwd=str(gb.BRAIN_ROOT),
            env=_hermes_env(),
            text=True,
            capture_output=True,
            timeout=int(os.environ.get("BRAIN_CHAT_HERMES_TIMEOUT", "180")),
            **no_window_kwargs(),
        )
    except subprocess.TimeoutExpired as e:
        raise BrainChatError("Hermes 对话超时，请稍后重试或缩短问题。") from e
    except Exception as e:
        raise BrainChatError(f"Hermes 调用失败：{e}") from e
    if proc.returncode != 0:
        detail = (proc.stderr or proc.stdout or "").strip()[-1200:]
        raise BrainChatError(f"Hermes 调用失败：{detail or '未知错误'}")
    return _strip_hermes_output(proc.stdout)


def _build_prompt(user_message: str, citations: list[dict[str, Any]], mode: str) -> list[dict[str, str]]:
    context = "\n\n".join(
        f"[来源 {i+1}] {c.get('slug') or c.get('path') or 'unknown'}\n{c.get('snippet', '')}"
        for i, c in enumerate(citations)
    ) or "未检索到相关知识。"
    persona = "你是用户的私有知识库助手。"
    if mode == "amazon_operator":
        persona += "你同时以资深 Amazon 运营视角回答，重视合规、CTR/CVR、Listing/A+、长尾词策略；不得建议 Vine 或 SBV 视频广告。"
    else:
        persona += "回答应简洁、可执行，并优先依据知识库。"
    system = f"""{persona}

规则：
1. 先依据【知识库片段】回答；知识库没有明确依据时，直接说明“知识库中没有找到明确依据”。
2. 不编造来源；如引用知识，请在句末用“来源：xxx”这种自然方式说明。
3. 涉及 Amazon 售后/评价/站外引流时，主动规避平台合规风险。
4. 用中文回答，结构清晰，避免空泛。

【知识库片段】
{context}
"""
    return [{"role": "system", "content": system}, {"role": "user", "content": user_message}]


def _search_citations(user_message: str, category: str | None = None) -> list[dict[str, Any]]:
    def add_candidate(value: str, out: list[str]) -> None:
        v = value.strip()
        if v and v not in out:
            out.append(v[: gb.MAX_QUERY_CHARS])

    cleaned = re.sub(r"[？?！!。；;，,：:\n\r\t]+", " ", user_message).strip()
    candidates: list[str] = []
    add_candidate(user_message, candidates)
    add_candidate(cleaned, candidates)

    # GBrain 的 conservative/关键词检索对完整口语句不一定敏感；补充常见运营短语兜底。
    phrase_hints = [
        "广告优化", "广告", "优先级", "投放", "关键词", "否词", "Listing", "A+", "CTR", "CVR",
        "trail camera", "4G", "WiFi", "售后", "合规", "评价", "站外引流", "供应商", "1688",
    ]
    lower_msg = user_message.lower()
    for phrase in phrase_hints:
        if phrase.lower() in lower_msg:
            add_candidate(phrase, candidates)

    # 对英文/数字词保留空格组合，便于 ASIN、SKU、品牌、产品线命中。
    ascii_terms = re.findall(r"[A-Za-z0-9][A-Za-z0-9_-]{1,}", user_message)
    if ascii_terms:
        add_candidate(" ".join(ascii_terms[:8]), candidates)

    scope = (category or "").strip().lower()
    seen: set[str] = set()
    citations: list[dict[str, Any]] = []
    last_error: Exception | None = None
    for query in candidates[:8]:
        try:
            search_result = gb.search(query, "search")
        except Exception as e:
            last_error = e
            continue
        for item in search_result.get("items", [])[:8]:
            # Scope retrieval to a single knowledge category when requested.
            if scope and str(item.get("category") or "").strip().lower() != scope:
                continue
            key = str(item.get("slug") or item.get("path") or item.get("snippet") or "")
            if key and key not in seen:
                seen.add(key)
                citations.append(item)
        if citations:
            break
    if not citations and last_error:
        return [{"slug": "gbrain-search", "score": 0, "snippet": f"检索失败：{last_error}"}]
    return citations[:8]


def send_message(session_id: str, content: str) -> dict[str, Any]:
    init_db()
    msg = (content or "").strip()
    if not msg:
        raise BrainChatError("消息不能为空")
    if len(msg) > MAX_CHAT_CHARS:
        raise BrainChatError(f"消息过长，最多 {MAX_CHAT_CHARS} 字符")
    current = get_session(session_id)["session"]
    with _connect() as conn:
        user_msg = _insert_message(conn, session_id, "user", msg, [])
        # Auto-title: if still default title, use first message as title
        if current.get("title") in ("新知识对话", "新对话", None, ""):
            auto_title = msg.replace("\n", " ").strip()[:40] or "对话"
            conn.execute("UPDATE brain_chat_sessions SET title = ? WHERE id = ?", (auto_title, session_id))
    citations = _search_citations(msg)
    answer = _call_llm(_build_prompt(msg, citations, str(current.get("mode") or "knowledge")))
    with _connect() as conn:
        assistant_msg = _insert_message(conn, session_id, "assistant", answer, citations)
    return {"user_message": user_msg, "assistant_message": assistant_msg, "citations": citations, "model": chat_model_status()}


# ── Streaming chat (SSE) ────────────────────────────────────────────────────
_BRAIN_STREAM_WRAPPER = str(Path(__file__).parent / "brain_stream_wrapper.py")
_HERMES_VENV_PYTHON = str(Path.home() / ".hermes" / "hermes-agent" / "venv" / "bin" / "python")


def _row_to_message(row: sqlite3.Row) -> dict[str, Any]:
    try:
        citations = json.loads(row["citations_json"] or "[]")
    except Exception:
        citations = []
    return {
        "id": row["id"], "session_id": row["session_id"], "role": row["role"],
        "content": row["content"], "citations": citations, "created_at": row["created_at"],
    }


def begin_chat_turn(session_id: str, content: str, regenerate: bool = False, category: str | None = None) -> dict[str, Any]:
    """Prepare a streaming turn: persist the user message (or, for regenerate,
    reuse the last user question and drop the stale answer), retrieve citations
    (optionally scoped to a knowledge category), and build the hermes prompt.
    Returns {user_message, prompt, citations}."""
    init_db()
    current = get_session(session_id)["session"]
    user_msg: dict[str, Any] | None = None

    if regenerate:
        with _connect() as conn:
            rows = conn.execute(
                "SELECT * FROM brain_chat_messages WHERE session_id = ? ORDER BY created_at ASC, rowid ASC",
                (session_id,),
            ).fetchall()
        if not rows:
            raise BrainChatError("没有可重新生成的消息")
        last_user_idx = next((i for i in range(len(rows) - 1, -1, -1) if rows[i]["role"] == "user"), None)
        if last_user_idx is None:
            raise BrainChatError("没有找到可重新生成的用户问题")
        msg = (rows[last_user_idx]["content"] or "").strip()
        user_msg = _row_to_message(rows[last_user_idx])
        stale_ids = [rows[j]["id"] for j in range(last_user_idx + 1, len(rows)) if rows[j]["role"] == "assistant"]
        if stale_ids:
            with _connect() as conn:
                conn.executemany("DELETE FROM brain_chat_messages WHERE id = ?", [(i,) for i in stale_ids])
    else:
        msg = (content or "").strip()
        if not msg:
            raise BrainChatError("消息不能为空")
        if len(msg) > MAX_CHAT_CHARS:
            raise BrainChatError(f"消息过长，最多 {MAX_CHAT_CHARS} 字符")
        with _connect() as conn:
            user_msg = _insert_message(conn, session_id, "user", msg, [])
            if current.get("title") in ("新知识对话", "新对话", None, ""):
                auto_title = msg.replace("\n", " ").strip()[:40] or "对话"
                conn.execute("UPDATE brain_chat_sessions SET title = ? WHERE id = ?", (auto_title, session_id))

    citations = _search_citations(msg, category)
    prompt = _messages_to_hermes_prompt(_build_prompt(msg, citations, str(current.get("mode") or "knowledge")))
    return {"user_message": user_msg, "prompt": prompt, "citations": citations, "regenerated": regenerate}


def commit_chat_answer(session_id: str, answer: str, citations: list[dict[str, Any]]) -> dict[str, Any]:
    """Persist the streamed assistant answer once generation finishes."""
    clean = _strip_hermes_output(answer) if answer.strip() else ""
    if not clean:
        raise BrainChatError("Hermes 返回了空响应。")
    with _connect() as conn:
        return _insert_message(conn, session_id, "assistant", clean, citations)


def delete_message(message_id: str) -> dict[str, Any]:
    """Delete a single chat message (user or assistant)."""
    init_db()
    with _connect() as conn:
        cur = conn.execute("DELETE FROM brain_chat_messages WHERE id = ?", (message_id,))
    return {"deleted": (cur.rowcount or 0) > 0, "id": message_id}


def stream_spec(prompt: str) -> dict[str, Any]:
    """Subprocess spec for the no-tools streaming wrapper (used by the SSE route)."""
    return {
        "argv": [_HERMES_VENV_PYTHON, _BRAIN_STREAM_WRAPPER],
        "stdin": prompt.encode("utf-8"),
        "env": _hermes_env(),
        "cwd": str(gb.BRAIN_ROOT),
    }
