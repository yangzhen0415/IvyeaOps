"""
Render 8-sheet Amazon ad-campaign optimization xlsx from plan.json.

Visual spec is golden — see references/visual_spec.md. All colors / fonts / row heights /
column widths / status enums are hardcoded constants below; LLM never decides visual params.

Usage:
    python build_xlsx.py --plan plan.json --aggregated aggregated.json \
        --risk-template templates/risk_text.md \
        --out /path/to/{ASIN}_广告优化方案_{YYYYMMDD}.xlsx

Sheet renderers live in render_sheet_01..render_sheet_08. C1 establishes:
- Style constants
- Cell helpers (title / section / header / data row writers)
- Sheet skeleton with correct names + column widths + standard title row
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# COLOR PALETTE — must match references/visual_spec.md exactly
# =============================================================================
class C:
    NAVY = "1F4E78"           # title text + header bg
    RED_TEXT = "C00000"       # section red
    DARK_GREEN_TEXT = "006100" # protected-keyword section red
    WHITE = "FFFFFF"          # header text
    RED_BG = "FFCCCC"         # blackhole / P0 / risk
    GREEN_BG = "C6EFCE"       # healthy / efficient / protected / P3 / new
    GOLD_BG = "FFE699"        # observe / P1 / warning
    YELLOW_BG = "FFF2CC"      # general / P2 / total / note
    LIGHT_BLUE_BG = "DDEBF7"  # param / day-tag / before-data
    GRAY_BG = "EDEDED"        # secondary summary


# =============================================================================
# FONT / STYLE FACTORIES
# =============================================================================
FONT_FAMILY = "微软雅黑"


def font(*, size: int = 10, bold: bool = False, color: str | None = None) -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold,
                color=Color(rgb=color) if color else None)


def fill(rgb: str | None) -> PatternFill:
    if not rgb:
        return PatternFill(fill_type=None)
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def align(*, h: str = "left", v: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


_THIN_GRAY = Side(border_style="thin", color="BFBFBF")
DATA_BORDER = Border(left=_THIN_GRAY, right=_THIN_GRAY, top=_THIN_GRAY, bottom=_THIN_GRAY)


# =============================================================================
# ROW HEIGHTS
# =============================================================================
class H:
    TITLE = 28
    SECTION = 24
    HEADER = 32
    DATA = 22
    DATA_MID = 24
    DATA_TALL = 32
    DATA_XTALL = 38
    DATA_RISK = 50
    DATA_ACTION = 85


# =============================================================================
# SHEET CONFIG (name + column widths) — order is significant
# =============================================================================
SHEETS_CONFIG = [
    {"name": "01-现状诊断",       "cols": [38, 11, 9, 14, 10, 9, 38, 15]},
    {"name": "02-核心动作",       "cols": [9, 22, 28, 35, 35, 12]},
    {"name": "03-新Campaign搭建", "cols": [22, 38, 40, 12]},
    {"name": "04-否定词清单",     "cols": [42, 14, 40, 16]},
    {"name": "05-加码清单",       "cols": [45, 11, 22, 12, 24, 20]},
    {"name": "06-预算重分配",     "cols": [28, 15, 14, 20, 35, 15]},
    {"name": "07-执行Checklist",  "cols": [14, 52, 26, 12, 8]},
    {"name": "08-风险提示",       "cols": [35, 38, 40]},
]


# =============================================================================
# STATUS ENUMS — exact emoji + text + bg combos
# =============================================================================
EFFICIENCY_BG = {
    "✓✓ 高效": C.GREEN_BG,
    "✓ 健康": C.GREEN_BG,
    "❌ 效率黑洞": C.RED_BG,
    "⚠️ 一般": C.YELLOW_BG,
}
PRIORITY_BG = {
    "🔴 P0": C.RED_BG,
    "🟠 P1": C.GOLD_BG,
    "🟡 P2": C.YELLOW_BG,
    "🟢 P3": C.GREEN_BG,
}
NEG_TYPE_BG = {
    "❌ 立即否": C.RED_BG,
    "⚠️ 观察": C.GOLD_BG,
}
TREND_BG = {
    "🆕 新增": C.GREEN_BG,
    "⬆️ 增加": C.GREEN_BG,
    "⬇️ 缩减": C.YELLOW_BG,
    "➡️ 保持": C.LIGHT_BLUE_BG,
}
KW_TAG_BG = {
    "核心": C.GREEN_BG,
    "高潜": C.YELLOW_BG,
    "迁移": C.LIGHT_BLUE_BG,
    "拓词": C.GOLD_BG,
}
BLACKHOLE_TAG_BG = {
    "黑洞": C.RED_BG,
    "高潜": C.GOLD_BG,
    "稳定": C.GREEN_BG,
    "已转化": C.GREEN_BG,
    "一般": C.YELLOW_BG,
    "其他汇总": C.GRAY_BG,
    "合计": C.YELLOW_BG,
}


# =============================================================================
# CELL WRITERS
# =============================================================================
def write_title(ws: Worksheet, row: int, text: str, last_col_letter: str) -> None:
    """Sheet's R1 big title: navy text 14pt bold, merged across all columns, height 28."""
    cell = ws.cell(row, 1, text)
    cell.font = font(size=14, bold=True, color=C.NAVY)
    cell.alignment = align(h="left", v="center", wrap=True)
    cell.fill = fill(None)
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    ws.row_dimensions[row].height = H.TITLE


def write_section(ws: Worksheet, row: int, text: str, last_col_letter: str,
                  *, color: str = C.RED_TEXT) -> None:
    """Section header: red (or dark-green) text 12pt bold, merged, height 24."""
    cell = ws.cell(row, 1, text)
    cell.font = font(size=12, bold=True, color=color)
    cell.alignment = align(h="left", v="center", wrap=True)
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    ws.row_dimensions[row].height = H.SECTION


def write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Navy bg + white bold 11pt, height 32. Each cell centered."""
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row, i, h)
        cell.font = font(size=11, bold=True, color=C.WHITE)
        cell.fill = fill(C.NAVY)
        cell.alignment = align(h="center", v="center", wrap=True)
        cell.border = DATA_BORDER
    ws.row_dimensions[row].height = H.HEADER


def write_data_row(ws: Worksheet, row: int, values: list[Any],
                   *, bg: str | None = None,
                   bold: bool = False,
                   row_height: float | None = None,
                   align_per_col: list[str] | None = None,
                   border: bool = True) -> None:
    """
    Write one data row.
    align_per_col: list of "left" / "center" / "right" matching values length;
    if None, defaults to: col 1 = left, others = center.
    """
    n = len(values)
    aligns = align_per_col or (["left"] + ["center"] * (n - 1))
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row, i, v)
        cell.font = font(size=10, bold=bold)
        cell.fill = fill(bg)
        cell.alignment = align(h=aligns[i - 1], v="center", wrap=True)
        if border:
            cell.border = DATA_BORDER
    if row_height is not None:
        ws.row_dimensions[row].height = row_height
    else:
        ws.row_dimensions[row].height = H.DATA_MID


def write_merged_note(ws: Worksheet, row: int, text: str, last_col_letter: str,
                      *, bg: str | None = None,
                      bold: bool = False,
                      row_height: float | None = None) -> None:
    """Single merged cell across the row — for summary lines / inline notes."""
    cell = ws.cell(row, 1, text)
    cell.font = font(size=10, bold=bold)
    cell.fill = fill(bg)
    cell.alignment = align(h="left", v="center", wrap=True)
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    ws.row_dimensions[row].height = row_height or H.DATA_MID


# =============================================================================
# FORMAT HELPERS
# =============================================================================
def fmt_money(v: Any) -> str:
    if v is None or v == "" or v == "—":
        return "—"
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return str(v)


def fmt_pct(v: Any) -> str:
    """Accept 0.137 / 13.7 / '13.7%' / '13.7' — output '13.7%'."""
    if v is None or v == "" or v == "—":
        return "—"
    if isinstance(v, str):
        s = v.strip()
        if s.endswith("%"):
            return s
        try:
            return f"{float(s):.0f}%"
        except ValueError:
            return s
    try:
        f = float(v)
        # heuristic: 0..1 means ratio, otherwise raw percent
        if -2 <= f <= 2:
            return f"{f * 100:.0f}%"
        return f"{f:.0f}%"
    except (TypeError, ValueError):
        return str(v)


def fmt_int(v: Any) -> str:
    if v is None or v == "":
        return "—"
    try:
        return f"{int(v)}"
    except (TypeError, ValueError):
        return str(v)


# =============================================================================
# SHEET SETUP — apply column widths + return last_col_letter
# =============================================================================
def setup_sheet(ws: Worksheet, cfg: dict) -> str:
    for i, w in enumerate(cfg["cols"], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(len(cfg["cols"]))
    ws.sheet_view.showGridLines = False
    return last


# =============================================================================
# SHEET RENDERERS — placeholders for C2–C9
# =============================================================================
def render_sheet_01(ws: Worksheet, last: str, plan: dict, agg: dict) -> None:
    """01-现状诊断 — 段一: Campaign 效率对比；段二: 关键词黑洞 vs 高潜。

    8 columns (A..H) layout:
      Section 1: Campaign | 花费($) | 单量 | 每单成本($) | ACOS | 预算占比 | 单量占比 | 效率
      Section 2: 关键词    | 花费($) | 单量 | CPC($)      | ACOS | CTR     | 短评     | 标签
    """
    write_title(ws, 1, _sheet01_title(plan, agg), last)
    ws.row_dimensions[2].height = 8  # spacer

    # ---- Section 1: Campaign efficiency ----
    write_section(ws, 3, "一、各Campaign效率对比", last)
    write_header_row(ws, 4, [
        "Campaign", "花费($)", "单量", "每单成本($)", "ACOS", "预算占比", "单量占比", "效率",
    ])

    campaigns = plan.get("diagnosis", {}).get("campaign_efficiency") or []
    if not campaigns:
        # Fallback to aggregated data with auto-classification
        campaigns = _auto_campaign_efficiency(agg)

    row = 5
    total_spend = sum(_coerce_float(c.get("spend", 0)) for c in campaigns)
    total_orders = sum(_coerce_int(c.get("orders", 0)) for c in campaigns)
    for c in campaigns:
        verdict = c.get("verdict", "⚠️ 一般")
        bg = EFFICIENCY_BG.get(verdict, C.YELLOW_BG)
        write_data_row(ws, row, [
            c.get("campaign", ""),
            fmt_money(c.get("spend", 0)),
            fmt_int(c.get("orders", 0)),
            fmt_money(c.get("cpo", 0)),
            fmt_pct(c.get("acos", 0)),
            fmt_pct(c.get("budget_share", 0)),
            fmt_pct(c.get("order_share", 0)),
            verdict,
        ], bg=bg)
        row += 1

    # 合计 row (yellow bg, bold)
    if campaigns:
        avg_acos = (total_spend / sum(_coerce_float(c.get("sales", 0)) for c in campaigns)) \
            if any(c.get("sales") for c in campaigns) else 0
        cpo_total = (total_spend / total_orders) if total_orders else 0
        write_data_row(ws, row, [
            "合计",
            fmt_money(total_spend),
            fmt_int(total_orders),
            fmt_money(cpo_total),
            fmt_pct(avg_acos) if avg_acos else "—",
            "100%",
            "100%",
            "—",
        ], bg=C.YELLOW_BG, bold=True)
        row += 1

    row += 1  # spacer

    # ---- Section 2: Keyword blackholes ----
    write_section(ws, row, "二、关键词流量分析（黑洞 vs 高潜）", last)
    row += 1
    write_header_row(ws, row, [
        "关键词", "花费($)", "单量", "CPC($)", "ACOS", "CTR", "短评", "标签",
    ])
    row += 1

    blackholes = plan.get("diagnosis", {}).get("keyword_blackholes") or []
    if not blackholes:
        blackholes = _auto_keyword_blackholes(agg)

    for kw in blackholes:
        tag = kw.get("tag", "一般")
        bg = BLACKHOLE_TAG_BG.get(tag, C.YELLOW_BG)
        cpc = kw.get("cpc")
        if cpc in (None, "", "—"):
            sp = _coerce_float(kw.get("spend", 0))
            cl = _coerce_int(kw.get("clicks", 0))
            cpc = sp / cl if cl else 0
        bold = (tag == "合计")
        write_data_row(ws, row, [
            kw.get("keyword", ""),
            fmt_money(kw.get("spend", 0)),
            fmt_int(kw.get("orders", 0)),
            fmt_money(cpc) if cpc else "—",
            fmt_pct(kw.get("acos", 0)),
            fmt_pct(kw.get("ctr", 0)),
            kw.get("comment", ""),
            tag,
        ], bg=bg, bold=bold)
        row += 1


def _auto_campaign_efficiency(agg: dict) -> list[dict]:
    """Fallback when LLM didn't fill diagnosis.campaign_efficiency."""
    campaigns = agg.get("by_campaign", []) or []
    out = []
    for c in campaigns:
        acos = c.get("acos", 0)
        orders = c.get("orders", 0)
        if orders == 0:
            verdict = "❌ 效率黑洞"
        elif acos > 0.50:
            verdict = "❌ 效率黑洞"
        elif acos > 0.30:
            verdict = "⚠️ 一般"
        elif acos > 0.15:
            verdict = "✓ 健康"
        else:
            verdict = "✓✓ 高效"
        out.append({
            "campaign": c.get("campaign", ""),
            "spend": c.get("spend", 0),
            "orders": orders,
            "cpo": c.get("cpo", 0),
            "acos": acos,
            "budget_share": c.get("spend_share", 0),
            "order_share": c.get("order_share", 0),
            "verdict": verdict,
            "sales": c.get("sales", 0),
        })
    return out


def _auto_keyword_blackholes(agg: dict) -> list[dict]:
    """Fallback: top 8 keywords by spend, classified."""
    kws = agg.get("by_keyword", []) or []
    out = []
    for k in kws[:8]:
        acos = k.get("acos", 0)
        orders = k.get("orders", 0)
        spend = k.get("spend", 0)
        if orders == 0 and spend >= 5:
            tag = "黑洞"
            comment = f"花 ${spend:.2f}，零单"
        elif acos > 0.60:
            tag = "黑洞"
            comment = f"ACOS {acos*100:.0f}%，远超盈亏点"
        elif orders >= 2 and acos < 0.30:
            tag = "已转化"
            comment = f"{orders} 单 / ACOS {acos*100:.0f}%，效率好"
        elif spend >= 3 and orders == 0:
            tag = "高潜"
            comment = f"还在学习，观察"
        else:
            tag = "一般"
            comment = ""
        out.append({
            "keyword": k.get("keyword", ""),
            "spend": spend,
            "orders": orders,
            "ctr": k.get("ctr", 0),
            "acos": acos,
            "clicks": k.get("clicks", 0),
            "comment": comment,
            "tag": tag,
        })
    return out


def _coerce_float(v: Any) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().rstrip("%").replace("$", "").replace(",", "")
        try:
            f = float(s)
            return f / 100 if v.strip().endswith("%") else f
        except ValueError:
            return 0.0
    return 0.0


def _coerce_int(v: Any) -> int:
    return int(round(_coerce_float(v)))


def render_sheet_02(ws: Worksheet, last: str, plan: dict, agg: dict) -> None:
    """02-核心动作 — 优先级 / 动作类别 / 操作位置 / 具体参数 / 预期效果 / 时机。

    LLM 必须按 P0→P3 排序；CTR/CVR 杠杆类动作排在 P0 最前。
    每条动作行高 85（含多行参数+预期）。
    """
    write_title(ws, 1, _sheet02_title(plan), last)
    ws.row_dimensions[2].height = 8

    principle = plan.get("meta", {}).get("core_principle") or "按数据驱动，CTR/CVR 杠杆优先"
    write_section(ws, 3, f"原则：{principle}", last)

    ws.row_dimensions[4].height = 8  # spacer

    write_header_row(ws, 5, [
        "优先级", "动作类别", "操作位置", "具体参数", "预期效果", "时机",
    ])

    actions = plan.get("actions") or []
    row = 6
    for a in actions:
        prio = a.get("priority", "🟡 P2")
        bg = PRIORITY_BG.get(prio, C.YELLOW_BG)
        write_data_row(ws, row, [
            prio,
            a.get("title", ""),
            a.get("where", ""),
            a.get("params", ""),
            a.get("expected", ""),
            a.get("timing", ""),
        ], bg=bg, row_height=H.DATA_ACTION,
            align_per_col=["center", "left", "left", "left", "left", "center"])
        row += 1

    if not actions:
        write_merged_note(ws, row, "（本次无核心动作建议——通常说明 CSV 数据样本太少或 ASIN 表现稳定）",
                          last, bg=C.GRAY_BG, row_height=H.DATA)


def render_sheet_03(ws: Worksheet, last: str, plan: dict, agg: dict) -> None:
    """03-新Campaign搭建 — 每个 new_campaign 一段：红字章节 + 参数表。

    4 columns: 参数项(22) / 设置值(38) / 说明(40) / 必填(12)
    """
    write_title(ws, 1, "🏗️ 新建Campaign详细设置（抄作业版）", last)
    ws.row_dimensions[2].height = 8

    new_campaigns = plan.get("new_campaigns") or []
    if not new_campaigns:
        write_section(ws, 3, "本次无新建 Campaign 建议", last)
        write_merged_note(ws, 4,
            "若 goal=volume_with_top_of_search_push 但此处为空，请检查 LLM 输出。",
            last, bg=C.GRAY_BG, row_height=H.DATA)
        return

    row = 3
    for nc in new_campaigns:
        section_title = nc.get("section_title") or f"Campaign：{nc.get('name', '')}"
        write_section(ws, row, section_title, last)
        row += 1

        write_header_row(ws, row, ["参数项", "设置值", "说明", "重要度"])
        row += 1

        for p in nc.get("params", []):
            req = p.get("required", "选填")
            # row bg by required field
            if req == "⚠️ 同步操作":
                bg = C.RED_BG
            elif req == "必填":
                bg = C.LIGHT_BLUE_BG
            else:
                bg = None
            write_data_row(ws, row, [
                p.get("key", ""),
                p.get("value", ""),
                p.get("note", ""),
                req,
            ], bg=bg, row_height=H.DATA)
            row += 1

        row += 1  # spacer between campaigns


def render_sheet_04(ws: Worksheet, last: str, plan: dict, agg: dict) -> None:
    """04-否定词清单 — 顶部保护词红框 + 各 group 否定明细。

    4 columns: 否定词(42) / 类型(14) / 原因(40) / 浪费金额(16)
    保护词区每条 chip 行合并 A:D，绿底，dark_green section title。
    """
    write_title(ws, 1, "🚫 否定词清单（保护词完全不动）", last)
    ws.row_dimensions[2].height = 8

    negatives = plan.get("negatives") or {}
    protected = negatives.get("protected_keywords") or []
    groups = negatives.get("groups") or []
    summary_line = negatives.get("summary_line") or ""

    row = 3
    # ---- Protected keyword box ----
    if protected:
        write_section(ws, row, "🛡️ 保护词清单（这些词一个都不能否）",
                      last, color=C.DARK_GREEN_TEXT)
        row += 1
        for grp in protected:
            cell = ws.cell(row, 1, f"🛡️  {grp}")
            cell.font = font(size=10, bold=False)
            cell.fill = fill(C.GREEN_BG)
            cell.alignment = align(h="left", v="center", wrap=True)
            cell.border = DATA_BORDER
            ws.merge_cells(f"A{row}:{last}{row}")
            ws.row_dimensions[row].height = H.DATA
            row += 1
    else:
        write_section(ws, row, "⚪ 本次未指定保护词，所有词按数据判断", last)
        row += 1

    row += 1  # spacer

    # ---- Negative groups ----
    if not groups:
        write_merged_note(ws, row, "（本次无否定词建议）", last, bg=C.GRAY_BG, row_height=H.DATA)
        return

    for grp in groups:
        write_section(ws, row, grp.get("section_title", "否定清单"), last)
        row += 1
        write_header_row(ws, row, ["否定词", "类型", "原因", "21天浪费($)"])
        row += 1

        for it in grp.get("items", []):
            ntype = it.get("type", "❌ 立即否")
            bg = NEG_TYPE_BG.get(ntype, C.RED_BG)
            wasted = it.get("wasted_usd", 0)
            wasted_str = fmt_money(wasted) if wasted not in (0, "0", None, "") else "—"
            write_data_row(ws, row, [
                it.get("term", ""),
                ntype,
                it.get("reason", ""),
                wasted_str,
            ], bg=bg, row_height=H.DATA)
            row += 1

        row += 1  # spacer

    # Summary line
    if summary_line:
        write_merged_note(ws, row, summary_line, last, bg=C.YELLOW_BG, bold=True,
                          row_height=H.DATA_MID)


def render_sheet_05(ws: Worksheet, last: str, plan: dict, agg: dict) -> None:
    """05-加码清单 — 各 group 一段：红字章节 + 加码明细表。

    6 columns: 关键词(45) / 当前 bid(11) / 新 bid(22) / 变化(12) / 理由(24) / 预期(20)
    行底色由 tag 决定（核心绿/高潜黄/迁移浅蓝/拓词金）。
    """
    write_title(ws, 1, "📈 已验证转化词加码清单", last)
    ws.row_dimensions[2].height = 8

    boost = plan.get("boost") or {}
    groups = boost.get("groups") or []

    if not groups:
        write_merged_note(ws, 3, "（本次无加码建议——通常说明已转化词数量太少或表现稳定）",
                          last, bg=C.GRAY_BG, row_height=H.DATA)
        return

    row = 3
    for grp in groups:
        write_section(ws, row, grp.get("section_title", "加码清单"), last)
        row += 1
        write_header_row(ws, row, ["关键词", "当前bid", "新bid/位置", "变化", "理由", "预期"])
        row += 1

        for it in grp.get("items", []):
            tag = it.get("tag", "核心")
            bg = KW_TAG_BG.get(tag, C.YELLOW_BG)
            write_data_row(ws, row, [
                it.get("keyword", ""),
                it.get("current_bid", "—"),
                it.get("new_bid", ""),
                it.get("change", ""),
                it.get("reason", ""),
                it.get("expected", ""),
            ], bg=bg, row_height=H.DATA_TALL,
                align_per_col=["left", "center", "left", "center", "left", "left"])
            row += 1

        row += 1  # spacer


def render_sheet_06(ws: Worksheet, last: str, plan: dict, agg: dict) -> None:
    """06-预算重分配 — before/after 对比表 + 合计 + 注意事项。

    6 columns: Campaign(28) / 改前花费(15) / 改前占比(14) / 改后花费(20) /
               改后占比/说明(35) / 趋势(15)
    """
    write_title(ws, 1, "💰 广告预算重分配方案", last)
    ws.row_dimensions[2].height = 8

    write_section(ws, 3, "21天花费对比：改造前 vs 改造后", last)

    write_header_row(ws, 4, [
        "Campaign", "改前花费($)", "改前占比", "改后花费($)", "改后占比/说明", "趋势",
    ])

    bd = plan.get("budget_redistribution") or {}
    rows_data = bd.get("before_after") or []
    notes = bd.get("notes") or []

    row = 5
    total_before = total_after = 0.0
    for r in rows_data:
        trend = r.get("trend", "➡️ 保持")
        bg = TREND_BG.get(trend, C.LIGHT_BLUE_BG)
        bs = _coerce_float(r.get("before_spend", 0))
        as_ = _coerce_float(r.get("after_spend", 0))
        total_before += bs
        total_after += as_

        # E column: combine after_share + optional inline note (e.g. 'ToS+150% 溢价')
        after_share_str = fmt_pct(r.get("after_share", 0))
        inline_note = r.get("note", "")
        e_cell = after_share_str if not inline_note else f"{after_share_str} | {inline_note}"

        write_data_row(ws, row, [
            r.get("campaign", ""),
            fmt_money(bs),
            fmt_pct(r.get("before_share", 0)),
            fmt_money(as_),
            e_cell,
            trend,
        ], bg=bg, row_height=H.DATA_MID,
            align_per_col=["left", "center", "center", "center", "left", "center"])
        row += 1

    if rows_data:
        write_data_row(ws, row, [
            "合计",
            fmt_money(total_before),
            "100%",
            fmt_money(total_after),
            "100%",
            "—",
        ], bg=C.YELLOW_BG, bold=True)
        row += 1

    if notes:
        row += 1
        write_section(ws, row, "注意事项", last)
        row += 1
        for n in notes:
            write_merged_note(ws, row, n, last, bg=C.YELLOW_BG, row_height=H.DATA_MID)
            row += 1


def render_sheet_07(ws: Worksheet, last: str, plan: dict, agg: dict) -> None:
    """07-执行Checklist — Week 1 / Week 2 / 复盘指标。

    5 columns: 日期(14) / 任务(52) / 操作位置(26) / 耗时(12) / 完成(8)
    Day 标记含"周末"或"复盘"用 green_bg，否则 light_blue_bg。
    """
    write_title(ws, 1, "✅ 每日执行Checklist", last)
    ws.row_dimensions[2].height = 8

    cl = plan.get("checklist") or {}
    week1 = cl.get("week1") or []
    week2 = cl.get("week2") or []
    review_metrics = cl.get("review_metrics") or []

    row = 3
    if week1:
        write_section(ws, row, "Week 1 - 核心操作", last)
        row += 1
        write_header_row(ws, row, ["日期", "任务", "操作位置", "耗时", "完成"])
        row += 1
        for d in week1:
            day = d.get("day", "")
            bg = C.GREEN_BG if any(k in day for k in ("周末", "复盘", "Day 7", "周日", "周六")) else C.LIGHT_BLUE_BG
            mins = d.get("minutes", "")
            mins_str = f"{mins}分钟" if isinstance(mins, int) or (isinstance(mins, str) and mins.isdigit()) else str(mins)
            write_data_row(ws, row, [
                day, d.get("task", ""), d.get("where", ""), mins_str, "☐",
            ], bg=bg, row_height=H.DATA_MID,
                align_per_col=["center", "left", "left", "center", "center"])
            row += 1
        row += 1

    if week2:
        write_section(ws, row, "Week 2 - 复盘与调优", last)
        row += 1
        write_header_row(ws, row, ["日期", "任务", "操作位置", "耗时", "完成"])
        row += 1
        for d in week2:
            day = d.get("day", "")
            bg = C.GREEN_BG  # week 2 整周都用绿底（复盘期）
            mins = d.get("minutes", "")
            mins_str = f"{mins}分钟" if isinstance(mins, int) or (isinstance(mins, str) and mins.isdigit()) else str(mins)
            write_data_row(ws, row, [
                day, d.get("task", ""), d.get("where", ""), mins_str, "☐",
            ], bg=bg, row_height=H.DATA_MID,
                align_per_col=["center", "left", "left", "center", "center"])
            row += 1
        row += 1

    if review_metrics:
        write_section(ws, row, "复盘指标对比（Day 14 检查）", last)
        row += 1
        write_header_row(ws, row, ["指标", "基线（改前）", "目标（改后）", "实际", "达成"])
        row += 1
        for m in review_metrics:
            write_data_row(ws, row, [
                m.get("metric", ""),
                m.get("baseline", ""),
                m.get("target", ""),
                "",
                "☐",
            ], bg=C.YELLOW_BG, row_height=H.DATA_MID,
                align_per_col=["left", "center", "center", "center", "center"])
            row += 1

    if not (week1 or week2):
        write_merged_note(ws, 3, "（本次无 checklist 输出——通常说明 LLM 输出未填 checklist 字段）",
                          last, bg=C.GRAY_BG, row_height=H.DATA)


def render_sheet_08(ws: Worksheet, last: str, plan: dict, risk_template: str) -> None:
    """08-风险提示 — 固定文案，从 templates/risk_text.md 解析两段表。

    3 columns: 35 / 38 / 40
    段一红底（绝对不能做），段二金底（出现信号立即调整）。
    占位符 {{核心词}} / {{tos_campaign_name_or_default}} 在写入前替换。
    """
    write_title(ws, 1, "⚠️ 风险提示与注意事项", last)
    ws.row_dimensions[2].height = 8

    rendered = _substitute_risk_placeholders(risk_template, plan)
    sections = _parse_risk_sections(rendered)

    row = 3
    for sec in sections:
        bg = C.RED_BG if sec["kind"] == "forbidden" else C.GOLD_BG
        write_section(ws, row, sec["title"], last)
        row += 1
        write_header_row(ws, row, sec["headers"])
        row += 1
        for r in sec["rows"]:
            write_data_row(ws, row, r, bg=bg, row_height=H.DATA_RISK,
                           align_per_col=["left"] * len(r))
            row += 1
        row += 1  # spacer


def _substitute_risk_placeholders(template: str, plan: dict) -> str:
    """Replace {{核心词}} and {{tos_campaign_name_or_default}}."""
    # Core keyword: protected_keywords[0] first chunk, else first blackhole keyword, else default
    core_kw = "核心大词"
    protected = (plan.get("negatives") or {}).get("protected_keywords") or []
    if protected:
        core_kw = protected[0].split("/")[0].strip()
    else:
        bh = (plan.get("diagnosis") or {}).get("keyword_blackholes") or []
        for k in bh:
            kw = k.get("keyword", "").strip()
            if kw and k.get("tag") not in {"其他汇总", "合计"}:
                core_kw = kw
                break

    # ToS campaign name: from new_campaigns, prefer one containing 'ToS' / 'Top'
    tos_name = "新建顶部专属 Campaign"
    for nc in plan.get("new_campaigns") or []:
        n = nc.get("name", "")
        if any(s in n for s in ("ToS", "Top", "顶部")):
            tos_name = n
            break

    return (template
            .replace("{{核心词}}", core_kw)
            .replace("{{tos_campaign_name_or_default}}", tos_name))


def _parse_risk_sections(md: str) -> list[dict]:
    """Parse two markdown table sections from risk_text.md.

    Returns: [
      {"kind": "forbidden", "title": "绝对不能做的事", "headers": [...], "rows": [[...], ...]},
      {"kind": "signal",    "title": "出现这些信号立即调整", "headers": [...], "rows": [[...], ...]},
    ]
    """
    sections = []
    lines = md.splitlines()
    cur_section = None
    cur_kind = None

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        # Detect H2 section title
        m = re.match(r"^##\s+第[一二]段[：:]\s*(.+?)（", line)
        if m:
            title = m.group(1).strip()
            cur_kind = "forbidden" if "绝对" in title else "signal"
            cur_section = {"kind": cur_kind, "title": title, "headers": [], "rows": []}
            sections.append(cur_section)
            i += 1
            continue

        # Detect markdown table inside current section
        if cur_section is not None and line.startswith("|") and "---" not in line:
            cells = [c.strip() for c in line.strip("|").split("|")]
            if not cur_section["headers"]:
                cur_section["headers"] = cells
            else:
                # Skip placeholder-default table (last section)
                if any("占位符" in c or "默认" in c or "缺省" in c for c in cells):
                    # break out: this is the appendix table
                    cur_section = None
                    i += 1
                    continue
                cur_section["rows"].append(cells)
        i += 1

    # Filter: only keep first 2 sections (the appendix table title is "## 占位符默认值")
    return sections[:2]


# Title helpers (used by renderers + tests)
def _sheet01_title(plan: dict, agg: dict) -> str:
    meta = plan.get("meta", {})
    subject = meta.get("title_subject") or meta.get("sku_or_brand") or meta.get("asin", "")
    date_range = meta.get("date_range", "")
    days = meta.get("days") or agg.get("meta", {}).get("days", "")
    return f"📊 {subject} 广告现状诊断（{date_range} 共{days}天）"


def _sheet02_title(plan: dict) -> str:
    obj = plan.get("meta", {}).get("core_objective", "广告优化")
    return f"🎯 广告优化动作清单（{obj}）"


# =============================================================================
# MAIN
# =============================================================================
RENDERERS = [render_sheet_01, render_sheet_02, render_sheet_03, render_sheet_04,
             render_sheet_05, render_sheet_06, render_sheet_07, render_sheet_08]


def build(plan: dict, agg: dict, risk_template: str, out_path: Path) -> Path:
    wb = Workbook()
    # Drop default sheet
    wb.remove(wb.active)

    for cfg, renderer in zip(SHEETS_CONFIG, RENDERERS):
        ws = wb.create_sheet(title=cfg["name"])
        last = setup_sheet(ws, cfg)
        if renderer is render_sheet_08:
            renderer(ws, last, plan, risk_template)
        else:
            renderer(ws, last, plan, agg)

    out_path = _resolve_path_collision(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _resolve_path_collision(p: Path) -> Path:
    """If p exists, append _v2 / _v3 / ... until free."""
    if not p.exists():
        return p
    base = p.with_suffix("")
    suffix = p.suffix
    i = 2
    while True:
        cand = Path(f"{base}_v{i}{suffix}")
        if not cand.exists():
            return cand
        i += 1


def _default_out_name(plan: dict) -> str:
    meta = plan.get("meta", {})
    asin = meta.get("asin", "UNKNOWN")
    today = datetime.now().strftime("%Y%m%d")
    return f"{asin}_广告优化方案_{today}.xlsx"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--plan", required=True, help="LLM-generated plan.json")
    ap.add_argument("--aggregated", required=True, help="aggregated.json from parse_csvs.py")
    ap.add_argument("--risk-template", required=True,
                    help="templates/risk_text.md (Sheet 08 source)")
    ap.add_argument("--out", default=None, help="output xlsx path")
    args = ap.parse_args()

    plan = json.loads(Path(args.plan).read_text(encoding="utf-8"))
    agg = json.loads(Path(args.aggregated).read_text(encoding="utf-8"))
    risk_template = Path(args.risk_template).read_text(encoding="utf-8")

    out = Path(args.out) if args.out else Path.cwd() / _default_out_name(plan)
    final = build(plan, agg, risk_template, out)
    print(f"✓ Wrote {final}")


if __name__ == "__main__":
    main()
