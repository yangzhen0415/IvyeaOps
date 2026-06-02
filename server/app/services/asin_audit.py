"""ASIN audit job manager.

Runs the `amazon-asin-cosmo-rufus-audit` skill via the `claude` CLI as a
background subprocess. Jobs are tracked in-memory and their artifacts
persisted to disk so we can serve historical results after restarts.

Design:
- One job at a time (asyncio.Lock) — user is a single-seat operator
- Each job gets ~/.hermes/ivyea-ops-data/amazon-audits/<job_id>/
  - meta.json   (status, asin, marketplace, timestamps, error)
  - report.md   (raw claude markdown output — final)
  - report.json (parsed structured section, if claude complied)
  - stdout.log  (live tail)
- Hard timeout: 30 min
- 30-day retention (swept on startup)
"""
from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import shutil
import signal
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from app.services.runners import (  # noqa: F401 — re-exported for tests
    RUNNER_LABELS,
    RUNNER_ORDER,
    _build_runner_cmd,
    _extra_paths,
    _find_bin,
    _resolve_runner,
    build_child_env,
    resolve_with_pref,
    runner_status,
)

AUDIT_ROOT = Path.home() / ".hermes" / "ivyea-ops-data" / "amazon-audits"

_log = logging.getLogger(__name__)
AUDIT_ROOT.mkdir(parents=True, exist_ok=True)


# Hard kill after this many seconds.
HARD_TIMEOUT_SEC = 30 * 60
# Retention: delete job dirs older than this.
RETENTION_SEC = 30 * 24 * 3600
# How often we flush stdout.log to disk (bytes written threshold).
FLUSH_EVERY_BYTES = 4096

# Global single-job lock.
_job_lock = asyncio.Lock()
# id -> live Job object (only during execution; persisted state lives on disk).
_live_jobs: Dict[str, "Job"] = {}


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _job_dir(job_id: str) -> Path:
    return AUDIT_ROOT / job_id


@dataclass
class Job:
    job_id: str
    asin: str
    marketplace: str
    mode: str  # "full" | "rewrite_only"
    status: str = "queued"  # queued|running|done|failed|cancelled
    progress: str = ""
    created_at: str = field(default_factory=_now_iso)
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    error: Optional[str] = None
    pid: Optional[int] = None
    stdout_bytes: int = 0
    # Which runner to use ("auto" = auto-pick, else a specific one).
    # The actually-selected runner is stored in `runner_used` after start.
    runner_pref: str = "auto"
    runner_used: Optional[str] = None

    def to_meta(self) -> Dict[str, Any]:
        return {
            "job_id": self.job_id,
            "asin": self.asin,
            "marketplace": self.marketplace,
            "mode": self.mode,
            "status": self.status,
            "progress": self.progress,
            "created_at": self.created_at,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "error": self.error,
            "runner_pref": self.runner_pref,
            "runner_used": self.runner_used,
        }


def _write_meta(job: Job) -> None:
    d = _job_dir(job.job_id)
    d.mkdir(parents=True, exist_ok=True)
    (d / "meta.json").write_text(
        json.dumps(job.to_meta(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _read_meta(job_id: str) -> Optional[Dict[str, Any]]:
    mp = _job_dir(job_id) / "meta.json"
    if not mp.is_file():
        return None
    try:
        return json.loads(mp.read_text(encoding="utf-8"))
    except Exception:
        return None


def _build_prompt(asin: str, marketplace: str, mode: str) -> str:
    """Craft the claude prompt.

    Key requirement: claude MUST append a ```json ...``` block at the end with
    a stable schema so the UI can render tables.
    """
    full_or_rewrite = (
        "完整 11 板块审计报告" if mode != "rewrite_only" else "精简诊断 + 完整改写稿"
    )
    return f"""请使用 `amazon-asin-cosmo-rufus-audit` 技能，对以下 ASIN 做{full_or_rewrite}。

- ASIN: {asin}
- 站点: {marketplace}
- 评分：1-10 分制
- 缺失字段写 "未获取到"
- 不要猜测，不把行业常识写成已证实信息

**重要 — 输出格式要求（必须严格遵守）**：

1. 先按 skill 的 11 板块结构输出完整 markdown 报告
2. 报告结尾追加一段以 ```json 开头、``` 结尾的代码块，内含如下结构化字段（字段不要缺，缺失用空字符串或空数组）：

```json
{{
  "overview": {{
    "asin": "",
    "marketplace": "",
    "category": "",
    "title_summary": "",
    "key_specs": "",
    "top_risk": ""
  }},
  "scorecard": [
    {{ "dimension": "语义检索匹配度", "score": 0, "note": "" }},
    {{ "dimension": "查询属性覆盖度", "score": 0, "note": "" }},
    {{ "dimension": "COSMO 知识图谱对齐度", "score": 0, "note": "" }},
    {{ "dimension": "隐式查询解析友好度", "score": 0, "note": "" }},
    {{ "dimension": "Rufus 因果链完整度", "score": 0, "note": "" }},
    {{ "dimension": "用户行为信号质量", "score": 0, "note": "" }},
    {{ "dimension": "可解释比较生成能力", "score": 0, "note": "" }}
  ],
  "semantic_blind_spots": [
    {{
      "aspect": "主查询意图覆盖",
      "bullets": [
        {{ "label": "页面事实", "text": "" }},
        {{ "label": "经营证据", "text": "" }}
      ]
    }}
  ],
  "cosmo_nodes": [
    {{
      "node": "Who",
      "label_cn": "谁买",
      "bullets": [
        {{ "label": "页面事实", "text": "" }},
        {{ "label": "推断建议", "text": "" }}
      ]
    }},
    {{ "node": "When/Where", "label_cn": "何时何地", "bullets": [] }},
    {{ "node": "Problem", "label_cn": "解决什么问题", "bullets": [] }},
    {{ "node": "Concern", "label_cn": "顾虑", "bullets": [] }},
    {{ "node": "Outcome", "label_cn": "结果", "bullets": [] }}
  ],
  "rufus_qa": [
    {{ "question": "这是什么", "verdict": "能", "evidence": "" }},
    {{ "question": "适合谁", "verdict": "不能", "evidence": "" }},
    {{ "question": "怎么选", "verdict": "部分能", "evidence": "" }},
    {{ "question": "注意事项", "verdict": "部分能", "evidence": "" }},
    {{ "question": "不适合什么情况", "verdict": "不能", "evidence": "" }},
    {{ "question": "最常见顾虑", "verdict": "不能", "evidence": "" }}
  ],
  "behavior_signals": [
    {{
      "category": "评论量/星级",
      "bullets": [
        {{ "label": "页面事实", "text": "" }}
      ]
    }},
    {{ "category": "差评高频问题", "bullets": [] }},
    {{ "category": "误购/退货风险", "bullets": [] }},
    {{ "category": "经营侧异常", "bullets": [] }}
  ],
  "competitor_diff": [
    {{
      "topic": "竞品共性表达",
      "bullets": [
        {{ "label": "推断建议", "text": "" }}
      ]
    }},
    {{ "topic": "当前页面差异化是否可提取", "bullets": [] }},
    {{ "topic": "为什么选你不选别人", "bullets": [] }},
    {{ "topic": "合规风险", "bullets": [] }}
  ],
  "priorities": [
    {{ "level": "P0", "issue": "", "evidence": "", "action": "" }}
  ],
  "ad_plan": {{
    "objective": "",
    "campaigns": [
      {{ "name": "", "type": "", "targeting": "", "bid_range": "", "budget": "", "strategy": "" }}
    ],
    "keywords_exact": [
      {{ "keyword": "", "bid": "", "reason": "" }}
    ],
    "keywords_phrase_broad": [
      {{ "keyword": "", "bid": "", "reason": "" }}
    ],
    "product_targeting": [
      {{ "keyword": "", "bid": "", "reason": "" }}
    ],
    "negatives_immediate": ["词1", "词2"],
    "negatives_watch": ["词1", "词2"],
    "rules": ""
  }},
  "rewrites": {{
    "title": "",
    "bullets": ["", "", "", "", ""],
    "qa": [
      {{ "q": "", "a": "" }}
    ],
    "backend_terms": "",
    "image_plan": {{
      "main_image": [],
      "aux_images": [],
      "scene_images": []
    }},
    "aplus_plan": [],
    "compliance_reminders": []
  }}
}}
```

**证据标签规则（严格遵守 skill Evidence Labels）：**
- `label` 字段仅使用：`页面事实`、`评论证据`、`经营证据`、`推断建议` 四种之一
- 缺失真实证据的字段用 `推断建议` 明示，不要把猜测写成 `页面事实`
- bullets 数组里每一条 bullet 必须独立带 label，不要把多种证据混写在一条里
- `rufus_qa.verdict` 仅使用：`能` / `部分能` / `不能` 三个枚举值
- `cosmo_nodes` 必须覆盖全 5 节点（Who / When/Where / Problem / Concern / Outcome），某节点无内容时 `bullets: []` 留空即可

除该 JSON 块外，报告正文保持 markdown 结构不变。JSON 内字段用英文键名，值用中文。
"""


async def _run_claude(job: Job) -> None:
    """Spawn claude CLI, stream to stdout.log, kill on timeout."""
    jd = _job_dir(job.job_id)
    jd.mkdir(parents=True, exist_ok=True)
    stdout_log = jd / "stdout.log"

    prompt = _build_prompt(job.asin, job.marketplace, job.mode)

    # Respect an explicit runner preference; fall back to auto-pick.
    pref = (job.runner_pref or "auto").lower()
    if pref == "auto":
        runner, runner_bin = _resolve_runner()
    elif pref in RUNNER_ORDER:
        runner_bin = _find_bin(pref)
        runner = pref if runner_bin else None
    else:
        runner, runner_bin = None, None
        job.error = f"unknown runner: {pref}"

    if not runner_bin or not runner:
        job.status = "failed"
        if not job.error:
            job.error = (
                f"runner '{pref}' not available; tried {', '.join(RUNNER_ORDER)} "
                f"in {', '.join(_extra_paths())}"
            )
        job.finished_at = _now_iso()
        _write_meta(job)
        return

    job.runner_used = runner

    # Make sure the runner's own dir is on PATH so it can spawn helpers.
    child_env = {**os.environ}
    bin_dir = str(Path(runner_bin).parent)
    if bin_dir not in child_env.get("PATH", "").split(os.pathsep):
        child_env["PATH"] = bin_dir + os.pathsep + child_env.get("PATH", "")
    # hermes reads ~/.hermes/ — HOME may be missing under systemd.
    child_env.setdefault("HOME", str(Path.home()))

    cmd = _build_runner_cmd(runner, runner_bin, prompt)

    job.status = "running"
    job.started_at = _now_iso()
    mcp_note = "（MCP: sorftime + sif_mcp）" if runner == "hermes" else ""
    job.progress = f"已启动 {runner} 收集证据{mcp_note}…"
    _write_meta(job)

    start = time.monotonic()
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
            env=child_env,
        )
    except Exception as spawn_err:
        job.status = "failed"
        job.error = f"failed to spawn {runner}: {type(spawn_err).__name__}: {spawn_err}"
        job.finished_at = _now_iso()
        _write_meta(job)
        return
    job.pid = proc.pid
    _write_meta(job)

    buf_bytes = 0
    buf_chunks: List[bytes] = []

    async def _flush() -> None:
        nonlocal buf_bytes
        if not buf_chunks:
            return
        data = b"".join(buf_chunks)
        buf_chunks.clear()
        with stdout_log.open("ab") as f:
            f.write(data)
        job.stdout_bytes += len(data)
        buf_bytes = 0

    timed_out = False
    try:
        while True:
            # 30-min guardrail.
            if time.monotonic() - start > HARD_TIMEOUT_SEC:
                timed_out = True
                try:
                    proc.send_signal(signal.SIGTERM)
                    await asyncio.wait_for(proc.wait(), timeout=5)
                except asyncio.TimeoutError:
                    proc.kill()
                break

            try:
                chunk = await asyncio.wait_for(
                    proc.stdout.read(2048), timeout=2.0
                )
            except asyncio.TimeoutError:
                # Heartbeat update even when silent.
                elapsed = int(time.monotonic() - start)
                job.progress = f"分析中… {elapsed}s"
                _write_meta(job)
                continue

            if not chunk:
                break
            buf_chunks.append(chunk)
            buf_bytes += len(chunk)
            if buf_bytes >= FLUSH_EVERY_BYTES:
                await _flush()
                # Give UI a rough text preview length.
                kb = job.stdout_bytes // 1024
                job.progress = f"已生成 ~{kb} KB…"
                _write_meta(job)

        await _flush()
        rc = await proc.wait()
        job.finished_at = _now_iso()

        if timed_out:
            job.status = "failed"
            job.error = f"timeout after {HARD_TIMEOUT_SEC}s"
            _write_meta(job)
            return

        if rc != 0:
            job.status = "failed"
            tail = ""
            if stdout_log.is_file():
                tail = stdout_log.read_text(encoding="utf-8", errors="replace")[-800:]
            job.error = f"claude exited with code {rc}: {tail.strip()[-400:]}"
            _write_meta(job)
            return

        # Success — split markdown and structured JSON.
        raw = stdout_log.read_text(encoding="utf-8", errors="replace")
        md_text, structured = _split_report_and_json(raw)
        (jd / "report.md").write_text(md_text, encoding="utf-8")
        if structured is not None:
            (jd / "report.json").write_text(
                json.dumps(structured, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        job.status = "done"
        job.progress = "完成"
        _write_meta(job)
    except Exception as e:  # pragma: no cover - defensive
        job.status = "failed"
        job.error = f"{type(e).__name__}: {e}"
        job.finished_at = _now_iso()
        _write_meta(job)
        try:
            if proc.returncode is None:
                proc.kill()
        except Exception:
            pass


_JSON_FENCE_RE = re.compile(
    r"```json\s*\n(?P<body>.*?)\n```",
    re.DOTALL | re.IGNORECASE,
)


def _split_report_and_json(raw: str) -> tuple[str, Optional[Dict[str, Any]]]:
    """Split the tail ```json ... ``` block off and parse it.

    Tolerates multiple json blocks (takes the LAST one, since the skill
    appends the structured summary at the very end of the report).
    Tolerates trailing text after the fence.
    """
    matches = list(_JSON_FENCE_RE.finditer(raw))
    if not matches:
        return raw, None
    last = matches[-1]
    md_part = raw[: last.start()].rstrip() + "\n"
    body = last.group("body").strip()
    try:
        data = json.loads(body)
        return md_part, data
    except Exception:
        return raw, None


async def start_job(
    asin: str,
    marketplace: str,
    mode: str,
    runner_pref: str = "auto",
) -> Job:
    """Create and launch a job. Raises RuntimeError if busy."""
    if _job_lock.locked():
        raise RuntimeError("another audit is currently running")

    runner_pref = (runner_pref or "auto").lower()
    if runner_pref not in ("auto",) + RUNNER_ORDER:
        raise ValueError(f"unknown runner: {runner_pref}")

    # Pre-flight: refuse early if the requested runner isn't actually available,
    # so the user gets a 400 instead of a job that silently fails.
    if runner_pref == "auto":
        picked, _ = _resolve_runner()
        if not picked:
            raise RuntimeError("no agent CLI is available on this host")
    else:
        if not _find_bin(runner_pref):
            raise RuntimeError(f"runner '{runner_pref}' is not available")

    job_id = uuid.uuid4().hex[:12]
    job = Job(
        job_id=job_id,
        asin=asin.strip().upper(),
        marketplace=marketplace.strip().upper() or "US",
        mode=mode,
        runner_pref=runner_pref,
    )
    _live_jobs[job_id] = job
    _write_meta(job)

    async def _runner() -> None:
        async with _job_lock:
            try:
                await _run_claude(job)
            except Exception as e:
                # Defensive — don't let the job silently hang on unexpected errors.
                try:
                    job.status = "failed"
                    job.error = f"runner crashed: {type(e).__name__}: {e}"
                    job.finished_at = _now_iso()
                    _write_meta(job)
                except Exception:
                    pass
            finally:
                _live_jobs.pop(job.job_id, None)

    asyncio.create_task(_runner())
    return job


def get_job(job_id: str) -> Optional[Dict[str, Any]]:
    """Get job state + (when done) structured result.

    Preference: disk state — survives process restart.
    """
    meta = _read_meta(job_id)
    if meta is None:
        return None
    jd = _job_dir(job_id)
    result: Dict[str, Any] = {**meta}
    rj = jd / "report.json"
    rm = jd / "report.md"
    if rj.is_file():
        try:
            result["structured"] = json.loads(rj.read_text(encoding="utf-8"))
        except Exception:
            result["structured"] = None
    else:
        result["structured"] = None
    if rm.is_file():
        text = rm.read_text(encoding="utf-8", errors="replace")
        # Cap payload to avoid massive JSON responses — UI reads ~first 100KB.
        if len(text) > 200_000:
            text = text[:200_000] + "\n\n…(truncated, download for full)…"
        result["raw_md"] = text
    else:
        # Fallback: show live tail while running.
        sl = jd / "stdout.log"
        if sl.is_file():
            raw = sl.read_text(encoding="utf-8", errors="replace")
            result["raw_md"] = raw[-20_000:]
        else:
            result["raw_md"] = ""
    return result


def list_jobs(limit: int = 20) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not AUDIT_ROOT.is_dir():
        return rows
    for entry in sorted(AUDIT_ROOT.iterdir(), key=lambda p: p.name, reverse=True):
        if not entry.is_dir():
            continue
        meta = _read_meta(entry.name)
        if not meta:
            continue
        rows.append(meta)
        if len(rows) >= limit:
            break
    # Sort by created_at desc (id is ordered but not guaranteed).
    rows.sort(key=lambda r: r.get("created_at", ""), reverse=True)
    return rows[:limit]


def download_path(job_id: str, fmt: str) -> Optional[Path]:
    """Return path of the downloadable artifact, if it exists.

    For xlsx, generate on demand (and cache) from report.json.
    """
    jd = _job_dir(job_id)
    if fmt == "md":
        fp = jd / "report.md"
        return fp if fp.is_file() else None
    if fmt == "json":
        fp = jd / "report.json"
        return fp if fp.is_file() else None
    if fmt == "xlsx":
        rj = jd / "report.json"
        if not rj.is_file():
            return None
        xp = jd / "report.xlsx"
        # Regenerate if missing or older than the structured JSON.
        if not xp.is_file() or xp.stat().st_mtime < rj.stat().st_mtime:
            try:
                structured = json.loads(rj.read_text(encoding="utf-8"))
            except Exception:
                _log.exception("xlsx: failed to parse report.json for job %s", job_id)
                return None
            meta = _read_meta(job_id) or {}
            try:
                build_xlsx(xp, structured, meta)
            except Exception:
                _log.exception("xlsx: build_xlsx failed for job %s", job_id)
                return None
        return xp if xp.is_file() else None
    if fmt == "html":
        rm = jd / "report.md"
        rj = jd / "report.json"
        if not rm.is_file() and not rj.is_file():
            return None
        hp = jd / "report.html"
        # Regenerate if missing or older than either source.
        sources_mtime = max(
            (p.stat().st_mtime for p in (rm, rj) if p.is_file()),
            default=0.0,
        )
        if not hp.is_file() or hp.stat().st_mtime < sources_mtime:
            from app.services import html_report
            try:
                structured = None
                if rj.is_file():
                    try:
                        structured = json.loads(rj.read_text(encoding="utf-8"))
                    except Exception:
                        _log.exception(
                            "html: failed to parse report.json for job %s", job_id
                        )
                        structured = None
                raw_md = rm.read_text(encoding="utf-8", errors="replace") if rm.is_file() else ""
                meta = _read_meta(job_id) or {}
                html_report.build_asin_html(hp, meta, structured, raw_md)
            except Exception:
                _log.exception("html: build_asin_html failed for job %s", job_id)
                return None
        return hp if hp.is_file() else None
    return None


# --------------------------------------------------------------------------- #
# XLSX report generation
# --------------------------------------------------------------------------- #

def build_xlsx(
    out_path: Path,
    structured: Dict[str, Any],
    meta: Dict[str, Any],
) -> None:
    """Turn the structured JSON block into a multi-sheet xlsx workbook.

    Layout (每个 sheet 都是一张独立的表格):
      - 概览       (asin / 市场 / 日期 / 一句话判断)
      - 七维评分    (dimension / score / note)  — score 单元格按阈值上色 + 条形
      - 优先级改进  (level / issue / evidence / action) — P0 红 / P1 橙 / P2 蓝
      - 广告活动    (name / type / targeting / bid_range / budget / strategy)
      - 关键词-Exact (keyword / bid / reason)
      - 关键词-Phrase(keyword / bid / reason)
      - 否定词      (term / type / reason) — 立即否 红底 / 观察 橙底
      - 改写稿      (section / content) — 区块标签加粗

    Sheets with no data get a "（本次报告未提供）" row instead of being empty,
    so the workbook always has the same shape — easier to compare jobs.

    Colors mirror the HTML report and the React workbench view.
    """
    # Imported lazily so the server still starts if openpyxl is absent.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # --- Palette (must match html_report.py / workbench.css) ---
    HEADER_FILL = PatternFill("solid", fgColor="1F2A3A")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    # Score thresholds (same as HTML: ≥8 good, ≥5 mid, <5 bad)
    FILL_SCORE_GOOD = PatternFill("solid", fgColor="DFF5E1")  # _C_GOOD
    FILL_SCORE_MID = PatternFill("solid", fgColor="FFF4CC")   # _C_WARN
    FILL_SCORE_BAD = PatternFill("solid", fgColor="FBD4D4")   # _C_BAD

    # Priority levels
    FILL_P0 = PatternFill("solid", fgColor="F8BFBF")   # _C_P0
    FILL_P1 = PatternFill("solid", fgColor="FFE7A3")   # _C_P1
    FILL_P2 = PatternFill("solid", fgColor="C8DAF2")   # _C_P2

    # Negatives
    FILL_NEG_IMM = PatternFill("solid", fgColor="FFE0C2")     # _C_CUT
    FILL_NEG_WATCH = PatternFill("solid", fgColor="EFE3FF")   # _C_WATCH

    # Section label (改写稿)
    FILL_SECTION = PatternFill("solid", fgColor="F9FAFC")

    # Evidence labels (skill Evidence Rules: 4 categories)
    # 页面事实 = 蓝 / 评论证据 = 橙 / 经营证据 = 紫 / 推断建议 = 灰
    FILL_EVI_PAGE = PatternFill("solid", fgColor="D6E4F5")    # 页面事实
    FILL_EVI_REVIEW = PatternFill("solid", fgColor="FFE0C2")  # 评论证据
    FILL_EVI_OPS = PatternFill("solid", fgColor="E8DDF5")     # 经营证据
    FILL_EVI_INFER = PatternFill("solid", fgColor="EEF0F3")   # 推断建议
    EVIDENCE_FILL_MAP = {
        "页面事实": FILL_EVI_PAGE,
        "评论证据": FILL_EVI_REVIEW,
        "经营证据": FILL_EVI_OPS,
        "推断建议": FILL_EVI_INFER,
    }

    # Rufus verdict tri-state
    FILL_VERDICT_OK = PatternFill("solid", fgColor="DFF5E1")     # 能
    FILL_VERDICT_PART = PatternFill("solid", fgColor="FFF4CC")   # 部分能
    FILL_VERDICT_FAIL = PatternFill("solid", fgColor="FBD4D4")   # 不能
    VERDICT_FILL_MAP = {
        "能": FILL_VERDICT_OK,
        "部分能": FILL_VERDICT_PART,
        "不能": FILL_VERDICT_FAIL,
    }
    VERDICT_LABEL_MAP = {
        "能": "✅ 能",
        "部分能": "⚠️ 部分能",
        "不能": "❌ 不能",
    }

    BOLD = Font(bold=True)

    def _add_sheet(
        title: str,
        headers: List[str],
        rows: List[List[Any]],
        widths: Optional[List[int]] = None,
        row_stylers: Optional[List[Any]] = None,
    ) -> Any:
        """Create a sheet with header row + data rows.

        row_stylers: optional list aligned with data rows; each entry is a
        dict mapping 0-based column index -> (PatternFill | None, Font | None).
        """
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if not rows:
            ws.append(["（本次报告未提供）"] + [""] * (len(headers) - 1))
        else:
            for row_idx, row in enumerate(rows):
                ws.append(row)
                styler = (row_stylers or [None] * len(rows))[row_idx]
                if styler:
                    for col_idx, style in styler.items():
                        fill, font = style if isinstance(style, tuple) else (style, None)
                        cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                        if fill is not None:
                            cell.fill = fill
                        if font is not None:
                            cell.font = font
        # Column widths
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        else:
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 24
        # Wrap all data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.alignment and cell.alignment.wrap_text:
                    continue
                cell.alignment = WRAP
        # Freeze header
        ws.freeze_panes = "A2"
        return ws

    # --- 概览 ---
    ov = structured.get("overview") or {}
    overview_rows = [
        ["ASIN", meta.get("asin") or ov.get("asin", "")],
        ["市场", meta.get("marketplace") or ov.get("marketplace", "")],
        ["类目", ov.get("category", "")],
        ["标题摘要", ov.get("title_summary", "")],
        ["核心规格", ov.get("key_specs", "")],
        ["最高风险", ov.get("top_risk", "")],
        ["运行 runner", meta.get("runner_used") or meta.get("runner_pref") or ""],
        ["审计时间", meta.get("finished_at") or meta.get("created_at", "")],
    ]
    # Bold the label column (col 0)
    overview_stylers = [{0: (None, BOLD)} for _ in overview_rows]
    _add_sheet(
        "概览",
        ["字段", "内容"],
        overview_rows,
        widths=[18, 80],
        row_stylers=overview_stylers,
    )

    # --- 七维评分 ---
    scorecard = structured.get("scorecard") or []

    def _score_bar(score: float) -> str:
        """10-segment Unicode progress bar so the xlsx shows a visual cue."""
        s = max(0.0, min(10.0, score))
        filled = int(round(s))
        return "█" * filled + "░" * (10 - filled)

    sc_rows: List[List[Any]] = []
    sc_stylers: List[Dict[int, Any]] = []
    for s in scorecard:
        if not isinstance(s, dict):
            continue
        try:
            raw = s.get("score")
            score_f = float(raw) if raw is not None else 0.0
        except (TypeError, ValueError):
            score_f = 0.0
        score_f = max(0.0, min(10.0, score_f))
        fill = (
            FILL_SCORE_GOOD if score_f >= 8
            else FILL_SCORE_MID if score_f >= 5
            else FILL_SCORE_BAD
        )
        bar = _score_bar(score_f)
        sc_rows.append([
            s.get("dimension", ""),
            score_f,
            bar,
            s.get("note", ""),
        ])
        # Color both score column (1) and bar column (2)
        sc_stylers.append({
            1: (fill, BOLD),
            2: (fill, Font(name="Menlo", color="485468")),
        })
    _add_sheet(
        "七维评分",
        ["维度", "评分 (1-10)", "分数条", "评语"],
        sc_rows,
        widths=[22, 12, 16, 60],
        row_stylers=sc_stylers,
    )

    # --- 3. 语义检索盲区 ---
    def _flatten_grouped_bullets(
        groups: List[Any],
        group_key: str,
    ) -> tuple[List[List[Any]], List[Dict[int, Any]]]:
        """Flatten a list of {group_key: ..., bullets: [{label, text}]} into rows.

        Returns ([group, evidence_label, text], [{col: (fill, font)}]).
        Rows with the same group repeat the group name only on the first row;
        later rows blank it out for readability. Evidence label column is filled
        with its category color.
        """
        rows: List[List[Any]] = []
        stylers: List[Dict[int, Any]] = []
        for g in groups or []:
            if not isinstance(g, dict):
                continue
            group_name = str(g.get(group_key, "") or "—")
            bullets = g.get("bullets") or []
            if not bullets:
                rows.append([group_name, "—", "（未获取到）"])
                stylers.append({0: (None, BOLD), 1: (FILL_EVI_INFER, None)})
                continue
            for i, b in enumerate(bullets):
                if not isinstance(b, dict):
                    # Tolerate a plain string bullet
                    if isinstance(b, str):
                        rows.append([group_name if i == 0 else "", "—", b])
                        stylers.append({
                            0: (None, BOLD if i == 0 else None),
                            1: (FILL_EVI_INFER, None),
                        })
                    continue
                label = str(b.get("label", "") or "—").strip()
                text = str(b.get("text", "") or "")
                fill = EVIDENCE_FILL_MAP.get(label, FILL_EVI_INFER)
                rows.append([group_name if i == 0 else "", label, text])
                stylers.append({
                    0: (None, BOLD if i == 0 else None),
                    1: (fill, BOLD),
                })
        return rows, stylers

    semantic = structured.get("semantic_blind_spots") or []
    sem_rows, sem_stylers = _flatten_grouped_bullets(semantic, "aspect")
    _add_sheet(
        "语义盲区",
        ["归类", "证据类型", "内容"],
        sem_rows,
        widths=[22, 12, 80],
        row_stylers=sem_stylers,
    )

    # --- 4. COSMO 节点诊断 ---
    cosmo = structured.get("cosmo_nodes") or []
    cs_rows: List[List[Any]] = []
    cs_stylers: List[Dict[int, Any]] = []
    for node in cosmo:
        if not isinstance(node, dict):
            continue
        node_en = str(node.get("node", "") or "—")
        node_cn = str(node.get("label_cn", "") or "")
        node_label = f"{node_en}（{node_cn}）" if node_cn else node_en
        bullets = node.get("bullets") or []
        if not bullets:
            cs_rows.append([node_label, "—", "（未获取到）"])
            cs_stylers.append({0: (None, BOLD), 1: (FILL_EVI_INFER, None)})
            continue
        for i, b in enumerate(bullets):
            if not isinstance(b, dict):
                if isinstance(b, str):
                    cs_rows.append([node_label if i == 0 else "", "—", b])
                    cs_stylers.append({
                        0: (None, BOLD if i == 0 else None),
                        1: (FILL_EVI_INFER, None),
                    })
                continue
            label = str(b.get("label", "") or "—").strip()
            text = str(b.get("text", "") or "")
            fill = EVIDENCE_FILL_MAP.get(label, FILL_EVI_INFER)
            cs_rows.append([node_label if i == 0 else "", label, text])
            cs_stylers.append({
                0: (None, BOLD if i == 0 else None),
                1: (fill, BOLD),
            })
    _add_sheet(
        "COSMO节点",
        ["节点", "证据类型", "内容"],
        cs_rows,
        widths=[22, 12, 80],
        row_stylers=cs_stylers,
    )

    # --- 5. Rufus 问答能力测试 ---
    rufus = structured.get("rufus_qa") or []
    ru_rows: List[List[Any]] = []
    ru_stylers: List[Dict[int, Any]] = []
    for q in rufus:
        if not isinstance(q, dict):
            continue
        question = str(q.get("question", "") or "")
        verdict_raw = str(q.get("verdict", "") or "").strip()
        verdict_label = VERDICT_LABEL_MAP.get(verdict_raw, verdict_raw or "—")
        verdict_fill = VERDICT_FILL_MAP.get(verdict_raw)
        evidence = str(q.get("evidence", "") or "")
        ru_rows.append([question, verdict_label, evidence])
        styler = {}
        if verdict_fill is not None:
            styler[1] = (verdict_fill, BOLD)
        ru_stylers.append(styler)
    _add_sheet(
        "Rufus问答",
        ["问题", "判定", "证据/缺口"],
        ru_rows,
        widths=[28, 14, 70],
        row_stylers=ru_stylers,
    )

    # --- 6. 用户行为信号诊断 ---
    behavior = structured.get("behavior_signals") or []
    bh_rows, bh_stylers = _flatten_grouped_bullets(behavior, "category")
    _add_sheet(
        "用户行为信号",
        ["分类", "证据类型", "内容"],
        bh_rows,
        widths=[22, 12, 80],
        row_stylers=bh_stylers,
    )

    # --- 7. 竞品差异化可提取性 ---
    compdiff = structured.get("competitor_diff") or []
    cd_rows, cd_stylers = _flatten_grouped_bullets(compdiff, "topic")
    _add_sheet(
        "竞品差异化",
        ["主题", "证据类型", "内容"],
        cd_rows,
        widths=[22, 12, 80],
        row_stylers=cd_stylers,
    )

    # --- 8. 优先级改进 ---
    priorities = structured.get("priorities") or []
    pr_rows: List[List[Any]] = []
    pr_stylers: List[Dict[int, Any]] = []
    level_fills = {"P0": FILL_P0, "P1": FILL_P1, "P2": FILL_P2}
    level_labels = {"P0": "🔴 P0", "P1": "🟠 P1", "P2": "🟡 P2"}
    for p in priorities:
        if not isinstance(p, dict):
            continue
        lvl_raw = str(p.get("level", "")).upper().strip()
        lvl_label = level_labels.get(lvl_raw, lvl_raw or "—")
        fill = level_fills.get(lvl_raw)
        pr_rows.append([
            lvl_label,
            p.get("issue", ""),
            p.get("evidence", ""),
            p.get("action", ""),
        ])
        if fill is not None:
            pr_stylers.append({0: (fill, BOLD)})
        else:
            pr_stylers.append({})
    _add_sheet(
        "优先级改进",
        ["优先级", "问题", "依据", "行动建议"],
        pr_rows,
        widths=[10, 40, 50, 50],
        row_stylers=pr_stylers,
    )

    # --- 广告活动 ---
    ad_plan = structured.get("ad_plan") or {}
    campaigns = ad_plan.get("campaigns") or []
    camp_rows = [
        [
            c.get("name", ""),
            c.get("type", ""),
            c.get("targeting", ""),
            c.get("bid_range", ""),
            c.get("budget", ""),
            c.get("strategy", ""),
        ]
        for c in campaigns
        if isinstance(c, dict)
    ]
    # Bold campaign name
    camp_stylers = [{0: (None, BOLD)} for _ in camp_rows]
    _add_sheet(
        "广告活动",
        ["Campaign 名", "类型", "定位", "出价区间", "日预算", "策略"],
        camp_rows,
        widths=[26, 10, 36, 16, 12, 50],
        row_stylers=camp_stylers,
    )

    def _kw_rows(lst: List[Any]) -> List[List[Any]]:
        out: List[List[Any]] = []
        for k in lst or []:
            if isinstance(k, dict):
                out.append([k.get("keyword", ""), k.get("bid", ""), k.get("reason", "")])
            elif isinstance(k, str):
                out.append([k, "", ""])
        return out

    ex_rows = _kw_rows(ad_plan.get("keywords_exact"))
    _add_sheet(
        "关键词-Exact",
        ["关键词", "出价", "入选理由"],
        ex_rows,
        widths=[30, 10, 60],
        row_stylers=[{0: (None, Font(name="Menlo"))} for _ in ex_rows],
    )
    ph_rows = _kw_rows(ad_plan.get("keywords_phrase_broad"))
    _add_sheet(
        "关键词-Phrase",
        ["关键词", "出价", "入选理由"],
        ph_rows,
        widths=[30, 10, 60],
        row_stylers=[{0: (None, Font(name="Menlo"))} for _ in ph_rows],
    )

    # --- 否定词 ---
    # 兼容两种 shape：① string ② {term/keyword/word, reason/note}
    def _neg_label(item: Any) -> tuple[str, str]:
        if isinstance(item, str):
            return item, ""
        if isinstance(item, dict):
            term = item.get("term") or item.get("keyword") or item.get("word") or item.get("text") or ""
            reason = item.get("reason") or item.get("note") or ""
            return str(term), str(reason)
        return str(item), ""

    neg_rows: List[List[Any]] = []
    neg_stylers: List[Dict[int, Any]] = []
    for item in ad_plan.get("negatives_immediate") or []:
        term, reason = _neg_label(item)
        if not term:
            continue
        neg_rows.append([term, "❌ 立即否", reason])
        neg_stylers.append({1: (FILL_NEG_IMM, BOLD)})
    for item in ad_plan.get("negatives_watch") or []:
        term, reason = _neg_label(item)
        if not term:
            continue
        neg_rows.append([term, "⚠️ 观察", reason])
        neg_stylers.append({1: (FILL_NEG_WATCH, BOLD)})
    _add_sheet(
        "否定词",
        ["词", "类型", "原因"],
        neg_rows,
        widths=[30, 14, 50],
        row_stylers=neg_stylers,
    )

    # --- 改写稿 ---
    rw = structured.get("rewrites") or {}
    rewrite_rows: List[List[Any]] = []
    if rw.get("title"):
        rewrite_rows.append(["标题", rw.get("title", "")])
    for i, b in enumerate(rw.get("bullets") or [], 1):
        rewrite_rows.append([f"五点 {i}", b])
    qa_list = rw.get("qa") or []
    for i, q in enumerate(qa_list, 1):
        if isinstance(q, dict):
            rewrite_rows.append([f"Q&A {i}", f"Q: {q.get('q','')}\nA: {q.get('a','')}"])
    if rw.get("backend_terms"):
        rewrite_rows.append(["Backend Terms", rw.get("backend_terms", "")])
    img = rw.get("image_plan") or {}
    if img:
        for key, zh in (
            ("main_image", "主图计划"),
            ("aux_images", "辅图计划"),
            ("scene_images", "场景图计划"),
        ):
            vals = img.get(key) or []
            if vals:
                rewrite_rows.append([zh, "\n".join(str(v) for v in vals)])
    aplus = rw.get("aplus_plan") or []
    if aplus:
        rewrite_rows.append(["A+ 方案", "\n".join(str(v) for v in aplus)])
    compliance = rw.get("compliance_reminders") or []
    if compliance:
        rewrite_rows.append(["合规提醒", "\n".join(str(v) for v in compliance)])
    rw_stylers = [{0: (FILL_SECTION, BOLD)} for _ in rewrite_rows]
    _add_sheet(
        "改写稿",
        ["板块", "内容"],
        rewrite_rows,
        widths=[16, 90],
        row_stylers=rw_stylers,
    )

    wb.save(out_path)


def is_busy() -> bool:
    return _job_lock.locked()


def sweep_expired() -> int:
    """Delete audit dirs older than RETENTION_SEC. Called on startup."""
    if not AUDIT_ROOT.is_dir():
        return 0
    cutoff = time.time() - RETENTION_SEC
    removed = 0
    for entry in AUDIT_ROOT.iterdir():
        try:
            if entry.is_dir() and entry.stat().st_mtime < cutoff:
                shutil.rmtree(entry, ignore_errors=True)
                removed += 1
        except Exception:
            continue
    return removed


def sweep_stale_running() -> int:
    """Mark any jobs left as running/queued on disk as failed.

    Called on startup: if the server was killed mid-run, the subprocess is
    gone but meta.json is stuck at status=running, so the UI shows a ghost
    "analyzing" task forever. On boot we don't have any live jobs yet
    (_live_jobs is empty), so anything still running on disk is stale.

    Returns the number of jobs rewritten.
    """
    if not AUDIT_ROOT.is_dir():
        return 0
    rewritten = 0
    for entry in AUDIT_ROOT.iterdir():
        if not entry.is_dir():
            continue
        meta = _read_meta(entry.name)
        if not meta:
            continue
        if meta.get("status") not in ("running", "queued"):
            continue
        meta["status"] = "failed"
        meta["error"] = meta.get("error") or "服务重启导致任务中断"
        meta["finished_at"] = meta.get("finished_at") or _now_iso()
        try:
            (entry / "meta.json").write_text(
                json.dumps(meta, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            rewritten += 1
        except Exception:
            continue
    return rewritten


def clear_failed() -> int:
    """Remove all failed/cancelled job directories. Returns count removed.

    Skips jobs that are still running or queued, regardless of meta.
    """
    if not AUDIT_ROOT.is_dir():
        return 0
    removed = 0
    for entry in AUDIT_ROOT.iterdir():
        if not entry.is_dir():
            continue
        meta = _read_meta(entry.name)
        if not meta:
            # Orphan dir with no meta — treat as garbage, clean it up.
            shutil.rmtree(entry, ignore_errors=True)
            removed += 1
            continue
        if meta.get("status") in ("failed", "cancelled"):
            shutil.rmtree(entry, ignore_errors=True)
            removed += 1
    return removed


def delete_job(job_id: str) -> bool:
    """Delete a single job directory. Returns True if deleted."""
    jd = _job_dir(job_id)
    if not jd.is_dir():
        return False
    # Don't delete running jobs.
    meta = _read_meta(job_id)
    if meta and meta.get("status") in ("running", "queued"):
        return False
    shutil.rmtree(jd, ignore_errors=True)
    return True
